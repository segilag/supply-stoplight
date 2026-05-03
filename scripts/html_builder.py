# -*- coding: utf-8 -*-
"""
html_builder.py  —  Supply Planner v5
Holds the HTML dashboard template and exposes two functions:

    render_html(json_str)         -> str   (full HTML string)
    write_html(json_str, out_path)         (render + write to disk)
"""

from pathlib import Path


HTML_TEMPLATE = r"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>supply.stoplight v5</title>
<link rel="preconnect" href="https://fonts.googleapis.com">
<link href="https://fonts.googleapis.com/css2?family=DM+Mono:wght@400;500&family=Inter:wght@300;400;500;600&display=swap" rel="stylesheet">
<script src="https://cdn.sheetjs.com/xlsx-0.20.3/package/dist/xlsx.full.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.4/dist/chart.umd.min.js"></script>
<style>
:root{
  /* ── Supply Stoplight — DARK (default) ── */
  --bg0:#0a0c10;--bg1:#111318;--bg2:#181c25;--bg3:#1e2330;--bg4:#252b3b;
  --tx1:#f0f2f8;--tx2:#a0a8c0;--tx3:#606880;
  --red:#ff4d4d;--red2:#ff4d4d22;--red3:#ff4d4d44;
  --yel:#f5a623;--yel2:#f5a62322;--yel3:#f5a62344;
  --grn:#3ecf8e;--grn2:#3ecf8e18;--grn3:#3ecf8e40;
  --blu:#2563eb;--blu2:#2563eb18;
  --ora:#3b82f6;--ora2:#3b82f618;--ora3:#3b82f640;
  --sky:#60a5fa;--sky2:#60a5fa18;
  --brd:#1e2540;--r:8px;--r2:5px;
}
/* ── LIGHT Theme override ── */
[data-theme="light"]{
  --bg0:#f0f4fb;--bg1:#ffffff;--bg2:#e8edf8;--bg3:#dde4f5;--bg4:#cdd7ef;
  --tx1:#0f1629;--tx2:#3d5280;--tx3:#7a8fb5;
  --red:#dc2626;--red2:#dc262614;--red3:#dc262628;
  --yel:#c97b0a;--yel2:#c97b0a14;--yel3:#c97b0a28;
  --grn:#0a7a52;--grn2:#0a7a5214;--grn3:#0a7a5228;
  --blu:#1d4ed8;--blu2:#1d4ed814;
  --ora:#2563eb;--ora2:#2563eb14;--ora3:#2563eb28;
  --sky:#3b82f6;--sky2:#3b82f614;
  --brd:#c3d0eb;
}
/* Light-only overrides for hardcoded dark colors */
[data-theme="light"] .inv-table tfoot tr{background:var(--bg3) !important}
[data-theme="light"] .sd.K{background:#fee2e2;color:#991b1b}
[data-theme="light"] .dot-sc{background:#9aa3b2}
[data-theme="light"] .kpi-sc .kpi-v{color:#5a6478}
[data-theme="light"] .badge-sc{background:#e2e6ed;color:#5a6478}
[data-theme="light"] .row-atraso td{background:rgba(220,38,38,0.06)}
/* Theme toggle button */
.theme-toggle{background:none;border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:4px 9px;cursor:pointer;font-size:14px;line-height:1;transition:.15s}
.theme-toggle:hover{border-color:var(--ora);color:var(--ora)}
*{box-sizing:border-box;margin:0;padding:0}
body{background:var(--bg0);color:var(--tx1);font-family:'Inter',sans-serif;font-size:13px;height:100vh;overflow:hidden;display:flex;flex-direction:column}
/* ── Topbar ── */
.topbar{background:var(--bg1);border-bottom:1px solid var(--brd);padding:0 20px;height:48px;display:flex;align-items:center;gap:16px;flex-shrink:0;z-index:10}
.logo{font-family:'DM Mono',monospace;font-size:15px;font-weight:500;color:var(--tx1);letter-spacing:.5px}
.logo span{color:var(--ora);font-weight:600}
.cut-info{font-size:11px;color:var(--tx2);background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);padding:3px 10px;white-space:nowrap;flex:1;min-width:0;overflow:hidden;text-overflow:ellipsis}
.cut-info b{color:var(--tx1)}
.topbar-right{margin-left:auto;display:flex;gap:8px;flex-shrink:0}
.rpt-menu-wrap{position:relative}
.rpt-dropdown{position:absolute;right:0;top:calc(100% + 4px);background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);min-width:220px;z-index:999;box-shadow:0 8px 24px #0009;overflow:hidden}
.rpt-dd-item{padding:10px 16px;font-size:12px;color:var(--tx1);cursor:pointer;display:flex;align-items:center;gap:8px;transition:.1s}
.rpt-dd-item:hover{background:var(--bg3);color:var(--ora)}
.rpt-dd-item-imp{border-top:1px solid var(--brd);color:var(--sky)}
.rpt-dd-sep{height:1px;background:var(--brd);margin:2px 0}
/* Report modal — country pills (theme-aware) */
.rpt-country-bar{display:flex;gap:6px;flex-wrap:wrap;margin-bottom:12px}
.rpt-cpill{background:var(--bg3);border:1px solid var(--brd);border-radius:16px;padding:4px 12px;font-size:11px;color:var(--tx2);cursor:pointer;white-space:nowrap;transition:.15s}
.rpt-cpill:hover{border-color:var(--tx2);color:var(--tx1)}
.rpt-cpill.active{background:var(--ora2);border-color:var(--ora);color:var(--ora);font-weight:600}
/* Report modal — centro multi-select (theme-aware) */
.rpt-msel-wrap{position:relative;width:100%}
.rpt-msel-btn{width:100%;text-align:left;background:var(--bg2);border:1px solid var(--brd);border-radius:6px;padding:8px 12px;font-size:12px;color:var(--tx1);cursor:pointer;display:flex;align-items:center;justify-content:space-between;font-family:'Inter',sans-serif;transition:.15s}
.rpt-msel-btn:hover{border-color:var(--ora)}
.rpt-msel-panel{position:fixed;background:var(--bg1);border:1px solid var(--brd);border-radius:8px;z-index:10001;max-height:280px;display:none;flex-direction:column;box-shadow:0 8px 32px #000a;min-width:280px}
.rpt-msel-panel.open{display:flex}
.rpt-msel-search{display:flex;align-items:center;gap:8px;padding:8px 12px;border-bottom:1px solid var(--brd);flex-shrink:0}
.rpt-msel-search input{flex:1;background:none;border:none;outline:none;color:var(--tx1);font-size:12px;font-family:'Inter',sans-serif}
.rpt-msel-scroll{overflow-y:auto;flex:1}
.rpt-msel-grp-hdr{font-size:10px;font-weight:700;color:var(--tx3);text-transform:uppercase;padding:8px 12px 3px;letter-spacing:.5px}
.rpt-msel-item{display:flex;align-items:center;gap:8px;padding:6px 12px;cursor:pointer;font-size:12px;color:var(--tx2);transition:.1s}
.rpt-msel-item:hover{background:var(--bg3);color:var(--tx1)}
.rpt-msel-item input[type=checkbox]{width:13px;height:13px;cursor:pointer;accent-color:var(--grn);flex-shrink:0}
.rpt-msel-all-row{border-bottom:1px solid var(--brd);color:var(--tx1);font-weight:600}
.rpt-badge-r{background:var(--red2);color:var(--red);border-radius:3px;padding:0 5px;font-size:10px;font-weight:600}
.rpt-badge-y{background:var(--yel2);color:var(--yel);border-radius:3px;padding:0 5px;font-size:10px}
.rpt-mat-cnt{color:var(--tx3);font-size:10px;margin-left:auto;padding-left:4px}
.btn{background:var(--bg3);border:1px solid var(--brd);color:var(--tx1);border-radius:var(--r2);padding:5px 12px;cursor:pointer;font-size:12px;font-family:'Inter',sans-serif;transition:.15s}
.btn:hover{background:var(--bg4);border-color:var(--ora)}
.copy-btn{background:none;border:1px solid var(--brd);color:var(--tx3);border-radius:var(--r2);padding:2px 9px;cursor:pointer;font-size:11px;font-family:'Inter',sans-serif;transition:.15s;display:inline-flex;align-items:center;gap:4px}
.copy-btn:hover{color:var(--tx1);border-color:var(--tx2)}
.tbl-toolbar{display:flex;justify-content:flex-end;padding:4px 4px 2px;gap:6px}
.copy-toast{position:fixed;bottom:28px;right:28px;background:var(--grn);color:#0a1a0f;padding:7px 16px;border-radius:var(--r);font-size:12px;font-weight:600;opacity:0;transform:translateY(6px);transition:.2s;pointer-events:none;z-index:9999}
.copy-toast.show{opacity:1;transform:translateY(0)}
/* ── Country filter bar ── */
.country-bar{background:var(--bg0);border-bottom:1px solid var(--brd);padding:5px 16px;display:flex;gap:6px;flex-shrink:0;flex-wrap:wrap;align-items:center}
.cpill{background:var(--bg2);border:1px solid var(--brd);border-radius:12px;color:var(--tx2);padding:3px 11px;cursor:pointer;font-size:11px;font-family:'Inter',sans-serif;transition:.15s;display:inline-flex;align-items:center;gap:4px;white-space:nowrap}
.cpill:hover{border-color:var(--blu);color:var(--tx1)}
.cpill.active{background:var(--blu2);border-color:var(--blu);color:var(--tx1);font-weight:600}
.cpill .cpill-cnt{background:var(--bg3);border-radius:8px;padding:0 5px;font-size:10px;color:var(--tx2)}
/* ── Centro dropdown filter ── */
.centro-filter-bar{background:var(--bg1);border-bottom:1px solid var(--brd);padding:5px 16px;display:flex;align-items:center;gap:10px;flex-shrink:0}
.centro-filter-lbl{font-size:11px;color:var(--tx2);white-space:nowrap;font-weight:500}
.centro-dd{position:relative}
.centro-dd-btn{background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx1);padding:5px 10px 5px 12px;cursor:pointer;font-size:12px;font-family:'Inter',sans-serif;display:inline-flex;align-items:center;gap:8px;transition:.15s;white-space:nowrap;max-width:420px}
.centro-dd-btn:hover,.centro-dd-btn.open{border-color:var(--blu);background:var(--bg3)}
.centro-dd-btn .dd-label{overflow:hidden;text-overflow:ellipsis;max-width:360px}
.centro-dd-btn .dd-arrow{font-size:9px;color:var(--tx3);flex-shrink:0}
.centro-dd-panel{position:fixed;background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r);z-index:9999;box-shadow:0 8px 32px #000a;width:280px;max-height:440px;display:none;flex-direction:column;overflow:hidden}
.centro-dd-panel.open{display:flex}
.centro-dd-search{background:var(--bg3);border:none;border-bottom:1px solid var(--brd);color:var(--tx1);padding:8px 12px;font-size:12px;font-family:'Inter',sans-serif;outline:none;width:100%;box-sizing:border-box;flex-shrink:0}
.centro-dd-search::placeholder{color:var(--tx3)}
.centro-dd-scroll{overflow-y:auto;flex:1}
.centro-dd-all{padding:8px 12px;border-bottom:1px solid var(--brd);display:flex;align-items:center;gap:8px;cursor:pointer;font-size:12px;color:var(--tx1);user-select:none}
.centro-dd-all:hover{background:var(--bg3)}
.centro-dd-grp-hdr{padding:6px 12px 2px;font-size:10px;font-weight:700;color:var(--tx3);text-transform:uppercase;letter-spacing:.5px;user-select:none}
.centro-dd-item{padding:5px 12px 5px 24px;display:flex;align-items:center;gap:8px;cursor:pointer;font-size:12px;color:var(--tx2);user-select:none;transition:.1s}
.centro-dd-item:hover{background:var(--bg3);color:var(--tx1)}
.centro-dd-item input[type=checkbox]{accent-color:var(--blu);cursor:pointer;flex-shrink:0}
.centro-dd-item .dd-mat-cnt{margin-left:auto;font-size:10px;color:var(--tx3)}
.centro-dd-item .dd-crit{background:var(--red);border-radius:8px;padding:0 5px;font-size:10px;color:#fff}
.badge-r{background:var(--red);border-radius:10px;padding:1px 6px;font-size:10px;color:#fff}
.badge-y{background:var(--yel);border-radius:10px;padding:1px 6px;font-size:10px;color:#111}
/* ── KPI strip ── */
.kpi-strip{background:var(--bg1);border-bottom:1px solid var(--brd);padding:0 20px;display:flex;gap:0;flex-shrink:0;overflow-x:auto}
.kpi-sep{width:1px;background:var(--brd);margin:6px 8px;flex-shrink:0}
.kpi-grp-lbl{display:flex;align-items:center;padding:0 8px 0 4px;color:var(--tx3);font-size:10px;font-weight:700;letter-spacing:.8px;text-transform:uppercase;white-space:nowrap;flex-shrink:0}
.kpi-g .kpi-v{color:var(--grn)}
.kpi{padding:5px 14px;border-right:1px solid var(--brd);min-width:105px}
.kpi:last-child{border-right:none}
.kpi-v{font-size:14px;font-weight:600;line-height:1;display:flex;align-items:baseline;gap:5px}
.kpi-sub{font-size:10px;color:var(--tx2);font-weight:400}
.kpi-l{font-size:9px;color:var(--tx3);margin-top:2px;text-transform:uppercase;letter-spacing:.5px}
.kpi-r .kpi-v{color:var(--red);font-size:22px}
.kpi-y .kpi-v{color:var(--yel);font-size:18px}
.kpi-a .kpi-v{color:#f0a500}
/* ── OC estado colors ── */
.oc-atr .kpi-v{color:var(--red)}
.oc-vcr .kpi-v{color:#f5a623}
.oc-ok  .kpi-v{color:#3ecf8e}
/* ── Main layout ── */
.main{display:flex;flex:1;overflow:hidden}
/* ── Sidebar ── */
.sidebar{width:290px;background:var(--bg1);border-right:1px solid var(--brd);display:flex;flex-direction:column;flex-shrink:0;overflow:visible}
.filter-panel{overflow-y:auto;flex-shrink:0;max-height:55vh}
.filter-bar{display:flex;padding:8px;gap:4px;border-bottom:1px solid var(--brd)}
.filter-btn{flex:1;background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:4px 0;cursor:pointer;font-size:11px;font-family:'Inter',sans-serif;transition:.15s}
.filter-btn:hover{border-color:var(--blu);color:var(--tx1)}
.filter-btn.active{background:var(--blu2);border-color:var(--blu);color:var(--tx1)}
.filter-imp-btn{flex:1;background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:3px 0;cursor:pointer;font-size:10px;font-family:'Inter',sans-serif;transition:.15s}
.filter-imp-btn:hover{border-color:#7c5cbf;color:var(--tx1)}
.filter-imp-btn.active{background:#7c5cbf22;border-color:#7c5cbf;color:#b39ddb}
.search-row{display:flex;align-items:center;gap:6px;padding:6px 8px;border-bottom:1px solid var(--brd)}
.search-input{flex:1;background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx1);padding:5px 8px;font-size:12px;font-family:'Inter',sans-serif;outline:none}
.search-input:focus{border-color:var(--blu)}
.sort-sel{background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:5px 6px;font-size:11px;font-family:'Inter',sans-serif;outline:none;cursor:pointer}
.filter-count{padding:4px 10px;font-size:10px;color:var(--tx3);border-bottom:1px solid var(--brd)}
.mat-list{flex:1;overflow-y:auto;padding:4px}
.mat-item{display:flex;align-items:center;gap:8px;padding:5px 8px;border-radius:var(--r2);cursor:pointer;transition:.1s;border-left:2px solid transparent}
.mat-item:hover{background:var(--bg2)}
.mat-item.active{background:var(--bg3);border-left-color:var(--ora)}
.dot{width:8px;height:8px;border-radius:50%;flex-shrink:0}
.dot-r{background:var(--red)}.dot-y{background:var(--yel)}.dot-g{background:var(--grn)}
.mat-info{flex:1;min-width:0}
.mat-code{font-family:'DM Mono',monospace;font-size:11px;color:var(--tx2);display:flex;align-items:center;gap:4px}
.mat-um{background:var(--bg3);border-radius:3px;padding:0 4px;font-size:10px}
.mat-desc{font-size:10px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;color:var(--tx1)}
.mat-right{text-align:right;flex-shrink:0}
.mat-cob{font-size:12px;font-weight:600;font-family:'DM Mono',monospace}
.dot-r-text{color:var(--red)}.dot-y-text{color:var(--yel)}.dot-g-text{color:var(--grn)}
.empty{color:var(--tx3);text-align:center;padding:20px;font-size:12px}
/* ── Content wrapper ── */
.content-wrapper{flex:1;display:flex;flex-direction:column;overflow:hidden}
.tab-bar{background:var(--bg1);border-bottom:1px solid var(--brd);padding:0 16px;display:flex;gap:2px;flex-shrink:0}
.tab-btn{background:none;border:none;border-bottom:2px solid transparent;color:var(--tx2);padding:9px 16px;cursor:pointer;font-size:12px;font-family:'Inter',sans-serif;white-space:nowrap;transition:.15s}
.tab-btn:hover{color:var(--tx1)}
.tab-btn.active{color:var(--tx1);border-bottom-color:var(--blu)}
.tab-content{flex:1;overflow:hidden;display:flex;flex-direction:column}
.tab-content.hidden{display:none}
/* ── Table wrap ── */
.table-wrap{flex:1;overflow:auto}
/* ── Inventory table ── */
.inv-table{border-collapse:collapse;white-space:nowrap;width:100%}
.inv-table th,.inv-table td{padding:6px 10px;border-bottom:1px solid var(--brd);text-align:left;font-size:12px}
.inv-table th{background:var(--bg2);color:var(--tx2);font-weight:500;font-size:11px;text-transform:uppercase;letter-spacing:.4px;position:sticky;top:0;z-index:3;cursor:pointer;user-select:none}
.inv-table th:hover{color:var(--tx1)}
.inv-table tbody tr:hover{background:var(--bg2)}
.inv-table tbody tr.sel{background:var(--bg3) !important;border-left:3px solid var(--ora)}
.inv-table tbody tr{cursor:pointer;transition:.1s}
.inv-table tfoot tr{background:#181818 !important;cursor:default}
.inv-table tfoot td{border-top:2px solid var(--ora);border-bottom:none;font-size:11px;font-weight:600;position:sticky;bottom:0;z-index:2}
.tot-label{color:var(--tx3);font-size:10px;letter-spacing:.4px}
.tot-um{color:var(--ora);font-weight:700}
.tot-v{color:var(--tx1)}
/* Sticky columns */
.sc{position:sticky;background:var(--bg1);z-index:2}
.inv-table th.sc{z-index:4}
.col-mat{left:0;min-width:90px;max-width:90px;border-right:none}
.col-desc{left:90px;width:var(--desc-w,220px);min-width:80px;border-right:1px solid var(--brd) !important;overflow:hidden;text-overflow:ellipsis;white-space:nowrap}
.col-desc-resize{position:absolute;right:0;top:0;bottom:0;width:6px;cursor:col-resize;background:transparent;z-index:5}
.col-desc-resize:hover,.col-desc-resize.dragging{background:var(--blu)}
.inv-table tbody tr:hover .sc{background:var(--bg2)}
.inv-table tbody tr.sel .sc{background:var(--bg3)}
/* Numeric cells */
.num{text-align:right;font-family:'DM Mono',monospace;font-size:12px}
.mono{font-family:'DM Mono',monospace}
.fw{font-weight:600}
.col-ing{color:var(--grn)}
.col-con{color:var(--red)}
.col-sal{color:var(--tx1)}
/* Coverage cells */
.cob-cell{font-weight:600;font-size:12px;border-radius:3px}
.cob-cell.crit{color:var(--red)}
.cob-cell.risk{color:var(--yel)}
/* Saturday cells */
.sat-r{background:var(--red2);color:var(--red);font-weight:600}
.sat-y{background:var(--yel2);color:var(--yel);font-weight:600}
.sat-g{background:var(--grn2);color:var(--grn)}
.sat-b{background:var(--sky2);color:var(--sky)}
.sat-k{background:var(--bg3);color:var(--tx3);font-weight:600}
.th-sat{min-width:58px;text-align:right}
.inv-table thead tr:nth-child(2) th{top:22px}
/* Zone badge */
.zona{text-align:center}
.zona-badge{padding:2px 7px;border-radius:10px;font-size:10px;font-weight:600}
.zona-r{background:var(--red3);color:var(--red)}
.zona-y{background:var(--yel3);color:var(--yel)}
.zona-g{background:var(--grn3);color:var(--grn)}
.zona-def{background:var(--bg3);color:var(--tx2)}
/* ── Tab OC summary mini-strip ── */
.tab-oc-summary{display:flex;align-items:center;gap:0;padding:8px 16px;background:var(--bg1);border-bottom:1px solid var(--brd);flex-shrink:0;flex-wrap:wrap;gap:6px}
.tos-group{display:flex;align-items:center;gap:6px}
.tos-lbl{font-size:10px;font-weight:700;color:var(--tx3);text-transform:uppercase;letter-spacing:.6px;white-space:nowrap;padding-right:4px}
.tos-card{display:flex;align-items:baseline;gap:5px;background:var(--bg2);border-radius:var(--r2);padding:4px 10px;border:1px solid var(--brd)}
.tos-v{font-size:15px;font-weight:700;line-height:1}
.tos-l{font-size:10px;color:var(--tx3)}
.tos-r .tos-v{color:var(--red)}
.tos-y .tos-v{color:var(--yel)}
.tos-g .tos-v{color:#3ecf8e}
.tos-sep{width:1px;height:32px;background:var(--brd);margin:0 8px}
/* ── OC Tabs ── */
.oc-tabs{display:flex;gap:4px;padding:8px 12px;border-bottom:1px solid var(--brd);flex-shrink:0;background:var(--bg1)}
.oc-tab-btn{background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:6px 14px;cursor:pointer;font-size:12px;font-family:'Inter',sans-serif;display:flex;align-items:center;gap:6px;transition:.15s}
.oc-tab-btn:hover{border-color:var(--blu);color:var(--tx1)}
.oc-tab-btn.active{background:var(--blu2);border-color:var(--blu);color:var(--tx1)}
.cnt-badge{background:var(--bg3);border-radius:10px;padding:1px 7px;font-size:11px}
/* ── OC Table ── */
.oc-table{border-collapse:collapse;white-space:nowrap;width:100%}
.oc-table th,.oc-table td{padding:7px 10px;border-bottom:1px solid var(--brd);font-size:12px}
.oc-table th{background:var(--bg2);color:var(--tx2);font-weight:500;font-size:11px;text-transform:uppercase;letter-spacing:.4px;position:sticky;top:0}
.oc-table tbody tr:hover{background:var(--bg2)}
.oc-table .num{text-align:right;font-family:'DM Mono',monospace}
.row-atraso td{border-left:3px solid var(--red);background:rgba(255,77,77,0.07)}
.row-sinlib td{border-left:3px solid var(--yel)}
.atraso-cell{color:var(--red);font-weight:600}
.lib-badge{padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
.badge-sinlib{background:var(--yel3);color:var(--yel)}
.badge-lib{background:var(--grn2);color:var(--grn)}
/* ── Calendar ── */
.cal-table{border-collapse:collapse;white-space:nowrap}
.cal-table th,.cal-table td{padding:4px 6px;border-bottom:1px solid var(--brd);font-size:11px}
.cal-table th{background:var(--bg2);color:var(--tx2);font-size:10px;text-transform:uppercase;position:sticky;top:0}
.cal-mat-cell{min-width:280px;max-width:280px;position:sticky;left:0;background:var(--bg1);z-index:2;display:flex;align-items:center;gap:5px;border-right:1px solid var(--brd) !important}
.cal-table th.cal-mat-cell{z-index:3}
.cal-code{font-size:10px;color:var(--tx2);font-family:'DM Mono',monospace;flex-shrink:0}
.cal-name{font-size:11px;color:var(--tx1);overflow:hidden;text-overflow:ellipsis;white-space:nowrap;flex:1;min-width:0}
.cal-cob{flex-shrink:0;font-family:'DM Mono',monospace;font-size:11px;font-weight:600;margin-left:4px}
.cal-cell{text-align:center;min-width:52px}
/* SD dots semáforo — compara stock vs buffer DDMRP */
.sd{display:inline-flex;align-items:center;justify-content:center;width:18px;height:18px;font-size:12px;cursor:pointer;border-radius:50%;transition:.1s;font-weight:700}
.sd:hover{transform:scale(1.25)}
.sd.G{background:var(--grn2);color:var(--grn)}
.sd.Y{background:var(--yel2);color:var(--yel)}
.sd.R{background:var(--red2);color:var(--red)}
.sd.K{background:#2d0a0a;color:#c0392b;font-size:15px;line-height:1}
.sd.B{background:var(--sky2);color:var(--sky)}
/* Calendar month header */
.cal-month-hdr{text-align:center;font-size:10px;color:var(--ora);border-bottom:1px solid var(--brd);padding:3px 2px;letter-spacing:.5px;font-weight:600}
.cal-table tbody tr:hover .cal-mat-cell{background:var(--bg2)}
.sel-row{background:var(--bg3)}
/* Semana num en header */
.cal-wk-num{font-size:9px;color:var(--ora);font-weight:700;display:block;line-height:1.2;letter-spacing:.3px}
/* Barra diagnóstico por tab */
.tab-diag-bar{display:flex;gap:6px;padding:6px 16px;background:var(--bg1);border-bottom:1px solid var(--brd);flex-shrink:0;flex-wrap:wrap;align-items:center}
.tdb-chip{padding:3px 10px;border-radius:20px;font-size:11px;font-weight:600;white-space:nowrap}
.tdb-r{background:var(--red2);color:var(--red);border:1px solid var(--red3)}
.tdb-y{background:var(--yel2);color:var(--yel);border:1px solid var(--yel3)}
.tdb-g{background:var(--grn2);color:var(--grn);border:1px solid var(--grn3)}
/* Filtros sidebar colapsables */
.filter-group-hdr{padding:5px 10px;font-size:10px;font-weight:700;color:var(--tx3);cursor:pointer;display:flex;justify-content:space-between;align-items:center;user-select:none;text-transform:uppercase;letter-spacing:.5px}
.filter-group-hdr:hover{color:var(--tx1);background:var(--bg2)}
.filter-group-body{transition:max-height .2s ease}
.filter-group-body.collapsed{max-height:0 !important;overflow:hidden}
.filter-static-label{font-size:10px;font-weight:700;letter-spacing:.06em;text-transform:uppercase;color:var(--muted,var(--tx3));padding:10px 12px 4px;user-select:none}
.fg-arrow{font-size:9px;transition:transform .2s}
.fg-arrow.collapsed{transform:rotate(-90deg)}
/* Badge filtros activos */
.active-filters-badge{margin:6px 8px 2px;padding:5px 10px;background:var(--ora2);border:1px solid var(--ora);border-radius:var(--r2);font-size:11px;color:var(--ora);display:flex;justify-content:space-between;align-items:center}
.afb-clear{background:none;border:none;color:var(--ora);cursor:pointer;font-size:11px;font-weight:700;padding:0}
/* Toggle vista calendario/lista */
.view-toggle{display:flex;gap:3px;padding:0}
.view-toggle-btn{background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:4px 12px;cursor:pointer;font-size:11px;font-family:'Inter',sans-serif;transition:.15s}
.view-toggle-btn:hover{border-color:var(--ora);color:var(--tx1)}
.view-toggle-btn.active{background:var(--ora2);border-color:var(--ora);color:var(--tx1)}
/* ── Estado ALERTA ── */
.dot-a{background:var(--ora)}
.dot-a-text{color:var(--ora)}
.kpi-a .kpi-v{color:var(--ora)}
.badge-a{background:var(--ora2);color:var(--ora);padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
/* ── Estado SIN_CONSUMO ── */
.dot-sc{background:#4a4f6a}
.dot-sc-text{color:#6b7299}
.kpi-sc .kpi-v{color:#6b7299}
.badge-sc{background:#2a2e3d;color:#6b7299;padding:2px 8px;border-radius:10px;font-size:11px;font-weight:600}
/* ── Sem-panel (detalle inferior) ── */
.sem-panel{border-top:2px solid var(--brd);background:var(--bg1);flex-shrink:0;display:flex;flex-direction:column;max-height:210px}
.sem-panel.hidden{display:none}
.sdp-header{display:flex;align-items:center;gap:10px;padding:5px 12px;border-bottom:1px solid var(--brd);font-size:12px;flex-shrink:0;background:var(--bg2)}
.sdp-title{font-family:'DM Mono',monospace;font-weight:500;color:var(--tx1)}
.sdp-info{color:var(--tx2);font-size:11px;flex:1}
.sdp-close{margin-left:auto;background:none;border:none;color:var(--tx3);cursor:pointer;font-size:15px;padding:0 6px;line-height:1}
.sdp-close:hover{color:var(--red)}
.sdp-tabs{display:flex;gap:4px;padding:4px 8px;border-bottom:1px solid var(--brd);flex-shrink:0;background:var(--bg0)}
.sdp-tab{background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:3px 12px;cursor:pointer;font-size:11px;font-family:'Inter',sans-serif;transition:.15s}
.sdp-tab:hover{border-color:var(--blu);color:var(--tx1)}
.sdp-tab.active{background:var(--blu2);border-color:var(--blu);color:var(--tx1)}
.sdp-table-wrap{overflow-y:auto;flex:1}
/* ── Mini-dots sidebar ── */
.mini-row{display:flex;gap:2px;margin-top:3px}
.mini-d{width:7px;height:7px;border-radius:50%;flex-shrink:0}
.mini-d.G{background:var(--grn)}.mini-d.Y{background:var(--yel)}
.mini-d.R{background:var(--red)}.mini-d.K{background:#c0392b}
.mini-d.B{background:var(--sky)}
.riesgo-tag{font-size:9px;color:var(--red);background:var(--red2);border-radius:3px;padding:1px 5px;margin-left:2px;white-space:nowrap}
.quiebre-tag{font-size:9px;color:#fff;background:#8b0000;border-radius:3px;padding:1px 5px;margin-left:2px;white-space:nowrap;font-weight:700}
.quiebre-lt-tag{font-size:9px;color:#fff;background:#7a4f00;border-radius:3px;padding:1px 5px;margin-left:2px;white-space:nowrap;font-weight:700}
.imp-tag{font-size:9px;color:#fff;background:#7c5cbf;border-radius:3px;padding:1px 5px;margin-left:2px;white-space:nowrap;font-weight:700}
.nal-tag{font-size:9px;color:#3ecf8e;background:#1e3a2e;border-radius:3px;padding:1px 5px;margin-left:2px;white-space:nowrap;font-weight:700}
.linea-tag{font-size:9px;color:var(--tx3);background:var(--bg3);border-radius:3px;padding:1px 5px;margin-left:4px;white-space:nowrap}
.mat-impact{font-size:9px;color:var(--tx3);white-space:nowrap;overflow:hidden;text-overflow:ellipsis;margin-top:1px}
.mat-impact .impact-stop{color:var(--red);font-weight:600}
.mat-impact .impact-ok{color:var(--grn)}
.sdp-impact{padding:4px 12px 2px;display:flex;gap:12px;flex-wrap:wrap;border-bottom:1px solid var(--brd);background:var(--bg2)}
.sdp-impact-chip{font-size:11px;display:flex;align-items:center;gap:5px}
.sdp-impact-chip b{color:var(--tx2);font-weight:500;font-size:10px}
.sdp-impact-chip span{color:var(--tx1)}
.sdp-impact-chip .ichip-stop{color:var(--red);font-weight:700}
.sdp-impact-chip .ichip-ok{color:var(--grn);font-weight:700}
.trend-up{color:#3ecf8e;font-size:10px;font-weight:700}
.trend-down{color:#e05c5c;font-size:10px;font-weight:700}
.trend-flat{color:#555;font-size:10px}
.trend-ico{font-size:9px;margin-right:3px;vertical-align:middle}
/* ── Multi-centro badges ── */
.col-centro{white-space:nowrap;padding-right:6px}
.centro-badge{font-size:10px;background:var(--bg3);color:var(--tx2);border-radius:3px;padding:1px 6px;white-space:nowrap}
.cal-centro{font-size:9px;color:var(--tx3);margin-left:4px}
.centro-chk{font-size:10px;color:var(--grn);margin-left:3px;font-weight:700}
/* ── Buffer chart modal ── */
.buf-modal-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.7);z-index:3000;align-items:center;justify-content:center}
.buf-modal-overlay.open{display:flex}
.buf-modal{background:var(--bg1);border:1px solid var(--brd);border-radius:10px;width:min(860px,95vw);max-height:90vh;display:flex;flex-direction:column;overflow:hidden}
.buf-modal-hdr{display:flex;align-items:center;gap:10px;padding:10px 16px;border-bottom:1px solid var(--brd);background:var(--bg2);flex-shrink:0}
.buf-modal-title{font-family:'DM Mono',monospace;font-size:13px;font-weight:500;color:var(--tx1);flex:1}
.buf-modal-close{background:none;border:none;color:var(--tx3);cursor:pointer;font-size:16px;padding:0 6px;line-height:1}
.buf-modal-close:hover{color:var(--red)}
.buf-modal-body{padding:16px;flex:1;overflow:auto;display:flex;flex-direction:column;gap:12px}
.buf-centro-tabs{display:flex;gap:6px;flex-wrap:wrap}
.buf-ctab{background:var(--bg3);border:1px solid var(--brd);border-radius:6px;padding:4px 12px;font-size:11px;color:var(--tx2);cursor:pointer}
.buf-ctab.active{background:var(--grn2);border-color:var(--grn);color:var(--grn)}
.buf-canvas-wrap{position:relative;height:340px;background:var(--bg0);border-radius:8px;padding:8px}
.buf-legend{display:flex;gap:16px;flex-wrap:wrap;font-size:11px;color:var(--tx2);padding:4px 0}
.buf-leg-item{display:flex;align-items:center;gap:5px}
.buf-leg-dot{width:10px;height:10px;border-radius:2px;flex-shrink:0}
/* ── Multi-select dropdown ── */
.msel-wrap{position:relative;width:100%}
.msel-btn{width:100%;background:var(--bg2);border:1px solid var(--brd);border-radius:var(--r2);color:var(--tx2);padding:6px 10px;font-size:11px;font-family:'Inter',sans-serif;cursor:pointer;text-align:left;display:flex;justify-content:space-between;align-items:center;transition:.15s}
.msel-btn:hover{border-color:var(--blu);color:var(--tx1)}
.msel-btn.has-sel{border-color:var(--blu);color:var(--tx1);background:var(--bg3)}
.msel-arrow{font-size:9px;opacity:.6;flex-shrink:0}
.msel-item{display:flex;align-items:center;gap:8px;padding:6px 10px;cursor:pointer;font-size:11px;color:var(--tx2);transition:.1s}
.msel-item:hover{background:var(--bg3);color:var(--tx1)}
.msel-item input[type=checkbox]{width:13px;height:13px;accent-color:var(--blu);cursor:pointer;flex-shrink:0}
.msel-item.checked{color:var(--tx1);background:#251800}
.msel-clear{padding:5px 10px;border-top:1px solid var(--brd);font-size:10px;color:var(--tx3);cursor:pointer;text-align:center;transition:.1s}
.msel-clear:hover{color:var(--red)}
.msel-search{display:flex;align-items:center;gap:5px;padding:5px 8px;border-bottom:1px solid var(--brd);position:sticky;top:0;background:var(--bg2);z-index:1}
.msel-search-icon{color:var(--tx3);font-size:12px;flex-shrink:0}
.msel-search input{flex:1;background:none;border:none;outline:none;color:var(--tx1);font-size:11px;font-family:'Inter',sans-serif}
.msel-search input::placeholder{color:var(--tx3)}
.msel-panel{position:fixed;background:var(--bg2);border:1px solid var(--blu);border-radius:var(--r2);z-index:9999;max-height:220px;overflow-y:auto;box-shadow:0 8px 24px #0006}
/* ── Scrollbar ── */
::-webkit-scrollbar{width:6px;height:6px}
::-webkit-scrollbar-track{background:var(--bg1)}
::-webkit-scrollbar-thumb{background:var(--brd);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--bg4)}

/* ── Responsive móvil (≤768px) ── */
@media (max-width:768px){
  body{overflow:auto;height:auto;min-height:100vh}
  .main{flex-direction:column;overflow:visible;height:auto}
  .sidebar{width:100%;max-height:0;overflow:hidden;transition:max-height .3s ease;border-right:none;border-bottom:1px solid var(--brd);flex-shrink:0}
  .sidebar.mob-open{max-height:80vh;overflow-y:auto}
  .content-wrapper{overflow:visible;height:auto}
  .tab-content{overflow:visible}
  .table-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch}
  .topbar{flex-wrap:wrap;height:auto;min-height:48px;padding:6px 10px;gap:6px}
  .cut-info{font-size:9px;order:3;width:100%}
  .topbar-right{margin-left:auto;gap:6px}
  .kpi{min-width:90px;padding:6px 10px}
  .kpi-v{font-size:18px}
  .filter-bar{flex-wrap:wrap}
  .filter-btn{flex:1 1 45%;font-size:10px}
  .filter-imp-btn{font-size:9px}
  .sem-panel{max-height:50vh}
  .col-desc{width:120px!important;min-width:80px}
  .mob-hide{display:none}
  .mob-sidebar-toggle{display:inline-flex!important}
}
@media (min-width:769px){
  .mob-sidebar-toggle{display:none!important}
}
</style>
</head>
<body>

<!-- Topbar -->
<header class="topbar">
  <div class="logo">supply<span>.</span>stoplight <span style="color:var(--tx3);font-size:12px">v5</span></div>
  <div class="cut-info">Stock: <b id="meta-saldo"></b> &nbsp;·&nbsp; ERP: <b id="meta-b2wise"></b> &nbsp;·&nbsp; <b id="meta-mes"></b> &nbsp;·&nbsp; <b id="meta-centro-nom"></b></div>
  <div class="topbar-right">
    <button class="theme-toggle" id="theme-toggle" onclick="toggleTheme()" title="Toggle theme"></button>
    <button class="btn mob-sidebar-toggle" id="mob-toggle-btn" onclick="toggleMobSidebar()" style="display:none">&#9776; Materials</button>
    <div class="rpt-menu-wrap" id="rpt-menu-wrap">
      <button class="btn" onclick="toggleRptMenu(event)">&#128196; &#128202; Download Report &#9662;</button>
      <div class="rpt-dropdown" id="rpt-dropdown" style="display:none">
        <div class="rpt-dd-item" onclick="closeRptMenu();openReportModal()">&#128196; PDF / Excel Report</div>
        <div class="rpt-dd-item rpt-dd-item-imp" onclick="closeRptMenu();exportImpactosExcel()">&#128203; Impact Report</div>
      </div>
    </div>
  </div>
</header>

<!-- Country filter bar -->
<div class="country-bar" id="country-bar"></div>
<!-- Centro dropdown filter -->
<div class="centro-filter-bar">
  <span class="centro-filter-lbl">Plant:</span>
  <div class="centro-dd" id="centro-dd">
    <button class="centro-dd-btn" id="centro-dd-btn" onclick="toggleCentroDDPanel(event)">
      <span class="dd-label" id="centro-dd-label">All plants</span>
      <span class="dd-arrow">&#9660;</span>
    </button>
  </div>
  <div class="centro-dd-panel" id="centro-dd-panel">
    <input class="centro-dd-search" id="centro-dd-search" placeholder="&#128269; Search plant..." oninput="centroDDFilter(this.value)">
    <div class="centro-dd-scroll" id="centro-dd-scroll"></div>
  </div>
</div>

<!-- KPI strip -->
<div class="kpi-strip" id="kpi-strip"></div>

<!-- Main -->
<div class="main">
  <!-- Sidebar -->
  <aside class="sidebar">
    <!-- Panel de filtros (scrollable) -->
    <div class="filter-panel" id="filter-panel">
    <!-- Badge filtros activos -->
    <div class="active-filters-badge" id="active-filters-badge" style="display:none">
      <span id="afb-text"></span>
      <button class="afb-clear" onclick="clearAllFilters()" title="Clear all filters">&#10005; Clear</button>
    </div>
    <!-- Grupo: Estado -->
    <div class="filter-group">
      <div class="filter-group-hdr" onclick="toggleFG(this)">Status <span class="fg-arrow">&#9660;</span></div>
      <div class="filter-group-body" id="fg-estado">
        <div class="filter-bar" style="border-top:none">
          <button class="filter-btn active" id="fb-ALL" onclick="setFilter('ALL',this)" title="Show all statuses">All</button>
          <button class="filter-btn" id="fb-CRITICO" onclick="setFilter('CRITICO',this)" style="color:var(--red)" title="CRITICAL: stock in red zone or stockout ≤12 weeks with no reaction time">Critical</button>
          <button class="filter-btn" id="fb-RIESGO" onclick="setFilter('RIESGO',this)" style="color:var(--yel)" title="AT RISK: stock in DDMRP yellow zone">At Risk</button>
          <button class="filter-btn" id="fb-ALERTA" onclick="setFilter('ALERTA',this)" style="color:var(--blu)" title="ALERT: projected stockout ≤12W but Lead Time allows placing PO in time">Alert</button>
          <button class="filter-btn" id="fb-OK" onclick="setFilter('OK',this)" style="color:var(--grn)" title="OK: coverage in DDMRP green zone">OK</button>
          <button class="filter-btn" id="fb-SIN_CONSUMO" onclick="setFilter('SIN_CONSUMO',this)" style="color:#6b7299" title="NO DEMAND: material with no active demand in the period">No demand</button>
        </div>
      </div>
    </div>
    <!-- Origen + Impacto en la misma fila -->
    <div style="display:grid;grid-template-columns:1fr 1fr;gap:0 6px;padding:2px 8px 8px">
      <div>
        <div class="filter-static-label" style="padding:6px 0 3px">Origin</div>
        <div id="fg-origen" style="display:flex;gap:3px">
          <button class="filter-imp-btn active" id="fo-ALL" onclick="setFilterOrigen('ALL',this)" style="font-size:9px;padding:3px 0">All</button>
          <button class="filter-imp-btn" id="fo-IMP" onclick="setFilterOrigen('IMP',this)" style="color:#b39ddb;font-size:9px;padding:3px 0">&#128666; IMP</button>
          <button class="filter-imp-btn" id="fo-NAC" onclick="setFilterOrigen('NAC',this)" style="color:var(--grn);font-size:9px;padding:3px 0">&#127981; DOM</button>
        </div>
      </div>
      <div>
        <div class="filter-static-label" style="padding:6px 0 3px">Impact</div>
        <div id="fg-impacto" style="display:flex;gap:3px">
          <button class="filter-imp-btn active" id="fi-ALL" onclick="setFilterImpacto('ALL',this)" style="font-size:9px;padding:3px 0">All</button>
          <button class="filter-imp-btn" id="fi-PARADA" onclick="setFilterImpacto('PARADA',this)" style="color:var(--red);font-size:9px;padding:3px 0">&#9888; Stop</button>
          <button class="filter-imp-btn" id="fi-NO_PARADA" onclick="setFilterImpacto('NO_PARADA',this)" style="color:var(--grn);font-size:9px;padding:3px 0">&#10003; OK</button>
        </div>
      </div>
    </div>
    <!-- Búsqueda y orden -->
    <div class="search-row">
      <input class="search-input" type="text" id="search-input" placeholder="Search material…" oninput="setSearch(this.value)">
      <select class="sort-sel" onchange="setSort(this.value)">
        <option value="estado">&#8597; Status</option>
        <option value="cobertura">&#8597; Coverage</option>
        <option value="codigo">&#8597; Code</option>
        <option value="nombre">&#8597; Name</option>
      </select>
    </div>
    <!-- Grupo material: inline label + dropdown (estilo Centro:) -->
    <div class="centro-filter-bar" style="border-top:1px solid var(--brd);border-bottom:none;padding:4px 10px">
      <span class="centro-filter-lbl" style="white-space:nowrap">Group:</span>
      <div class="msel-wrap" id="group-wrap" style="flex:1;position:relative">
        <button class="centro-dd-btn" onclick="togglePanel('group-panel')" id="group-btn"
                data-placeholder="All groups" style="width:100%;max-width:none;font-size:11px;padding:4px 8px">
          <span class="dd-label" id="group-label">All groups</span>
          <span class="dd-arrow">&#9660;</span>
        </button>
        <div class="msel-panel" id="group-panel" style="display:none"></div>
      </div>
    </div>
    <!-- Línea de fabricación: inline label + dropdown (estilo Centro:) -->
    <div class="centro-filter-bar" style="border-top:1px solid var(--brd);border-bottom:none;padding:4px 10px">
      <span class="centro-filter-lbl" style="white-space:nowrap">Line:</span>
      <div class="msel-wrap" id="linea-wrap" style="flex:1;position:relative">
        <button class="centro-dd-btn" onclick="togglePanel('linea-panel')" id="linea-btn"
                data-placeholder="All lines" style="width:100%;max-width:none;font-size:11px;padding:4px 8px">
          <span class="dd-label" id="linea-label">All lines</span>
          <span class="dd-arrow">&#9660;</span>
        </button>
        <div class="msel-panel" id="linea-panel" style="display:none"></div>
      </div>
    </div>
    </div><!-- /filter-panel -->
    <div class="filter-count" id="filter-count"></div>
    <div class="mat-list" id="mat-list"></div>
  </aside>

  <!-- Content -->
  <div class="content-wrapper">
    <div class="tab-bar">
      <button class="tab-btn active" data-tab="inv" onclick="setTab('inv',this)">📦 Inventory</button>
      <button class="tab-btn" data-tab="oc" onclick="setTab('oc',this)">🚚 POs &amp; PRs</button>
      <button class="tab-btn" data-tab="cal" onclick="setTab('cal',this)">📅 Stoplight</button>
    </div>
    <div id="tab-inv"  class="tab-content"></div>
    <div id="tab-oc"   class="tab-content hidden"></div>
    <div id="tab-cal"  class="tab-content hidden">
      <div id="cal-diag"></div>
      <div class="table-wrap" id="cal-table-wrap"></div>
    </div>
    <div id="sem-panel" class="sem-panel hidden">
      <div class="sdp-header">
        <span class="sdp-title" id="sdp-mat-title">&#8212;</span>
        <span id="sdp-estado"></span>
        <span class="sdp-info" id="sdp-info"></span>
        <button class="copy-btn" style="margin-left:auto;font-size:11px" onclick="showBufferChart()" title="Ver gráfica de inventario buffer DDMRP">&#128202; Gráfica</button>
        <button class="sdp-close" onclick="closeSemPanel()">&#10005;</button>
      </div>
      <div class="sdp-impact" id="sdp-impact" style="display:none"></div>
      <div class="sdp-tabs">
        <button class="sdp-tab active" id="sdp-tab-oc"  onclick="setSemDTab('oc')">Open POs (0)</button>
        <button class="sdp-tab"        id="sdp-tab-sol" onclick="setSemDTab('sol')">PRs (0)</button>
      </div>
      <div class="sdp-table-wrap" id="sdp-content"></div>
    </div>
  </div>
</div>

<!-- ── Buffer Chart Modal ── -->
<div class="buf-modal-overlay" id="buf-modal-overlay" onclick="if(event.target===this)closeBufModal()">
  <div class="buf-modal">
    <div class="buf-modal-hdr">
      <span class="buf-modal-title" id="buf-modal-title">Buffer Inventory Chart</span>
      <button class="buf-modal-close" onclick="closeBufModal()">&#10005;</button>
    </div>
    <div class="buf-modal-body">
      <div class="buf-centro-tabs" id="buf-centro-tabs"></div>
      <div class="buf-legend">
        <span class="buf-leg-item"><span class="buf-leg-dot" style="background:#3ecf8e"></span>Green Zone</span>
        <span class="buf-leg-item"><span class="buf-leg-dot" style="background:#f5a623"></span>Yellow Zone</span>
        <span class="buf-leg-item"><span class="buf-leg-dot" style="background:#ff4d4d"></span>Red Zone</span>
        <span class="buf-leg-item"><span class="buf-leg-dot" style="background:#fff;border:1px solid #555"></span>Projected stock</span>
        <span class="buf-leg-item"><span class="buf-leg-dot" style="background:#0f5a28;border:1px solid #1a9040"></span>PO receipts</span>
      </div>
      <div class="buf-canvas-wrap"><canvas id="buf-chart-canvas"></canvas></div>
    </div>
  </div>
</div>

<script>
// ── Theme system ──────────────────────────────────────────────────────────────
(function(){
  const saved = localStorage.getItem("sp_theme") || "dark";
  if (saved === "light") document.documentElement.setAttribute("data-theme","light");
})();
function toggleTheme(){
  const isLight = document.documentElement.getAttribute("data-theme") === "light";
  if (isLight) {
    document.documentElement.removeAttribute("data-theme");
    localStorage.setItem("sp_theme","dark");
  } else {
    document.documentElement.setAttribute("data-theme","light");
    localStorage.setItem("sp_theme","light");
  }
  _updateThemeBtn();
}
function _updateThemeBtn(){
  const btn = document.getElementById("theme-toggle");
  if (!btn) return;
  const isLight = document.documentElement.getAttribute("data-theme") === "light";
  btn.textContent = isLight ? "☀️" : "🌙";
  btn.title = isLight ? "Switch to dark mode" : "Switch to light mode";
}
// ── Data ─────────────────────────────────────────────────────────────────────
const DATA = /*DATA_PLACEHOLDER*/;
window._IMPACTOS_XLSX_B64 = "/*IMPACTOS_B64_PLACEHOLDER*/";

// ── State ────────────────────────────────────────────────────────────────────
let centrosSel   = new Set(Object.keys(DATA.centros));  // default: todos seleccionados
let filterPais   = "ALL";
let _centroDDOpen = false;
let filter       = "ALL";
let filterOrigen  = "ALL";   // "ALL" | "IMP" | "NAC"
let filterImpacto = "ALL";   // "ALL" | "PARADA" | "NO_PARADA"
let search     = "";
let sortKey    = "estado";
let mainTab    = "inv";
let selMat     = new Set();   // materiales seleccionados (vacío = todos)
let ocTab      = "oc";
let ocSort     = "atraso_desc";  // default: atrasadas primero
let selSemMat  = null;   // material seleccionado en semáforo
let selSemCk   = null;   // centro key del material seleccionado (null = todos los centros)
let semDTab    = "oc";   // sub-tab activo en sem-panel
let selGroup   = new Set();  // filtro Mat Group activo (Set vacío = todos)
let selLinea   = new Set();  // filtro Línea de fabricación activo

// ── Accessors multi-centro ───────────────────────────────────────────────────
// Compatibilidad: cd() devuelve el primer centro seleccionado
const cd     = () => DATA.centros[[...centrosSel][0]];
// allMats / allOC / allSOL: agregan todos los centros seleccionados, añadiendo _ck y _cn
const allMats = () => [...centrosSel].flatMap(k =>
  (DATA.centros[k]?.materiales || []).map(m => ({...m, _ck: k, _cn: DATA.centros[k].nombre}))
);
const allOC = () => [...centrosSel].flatMap(k =>
  (DATA.centros[k]?.oc || []).map(o => ({...o, _ck: k, _cn: DATA.centros[k].nombre}))
);
const allSOL = () => [...centrosSel].flatMap(k =>
  (DATA.centros[k]?.sol || []).map(s => ({...s, _ck: k, _cn: DATA.centros[k].nombre}))
);
const mats   = () => allMats();
const getOC  = () => allOC();
const getSOL = () => allSOL();
const isMulti = () => centrosSel.size > 1;

function getFiltered() {
  let ms = mats();
  if (filter !== "ALL") ms = ms.filter(m => m.estado === filter);
  if (filterOrigen === "IMP") ms = ms.filter(m => m.importado === true);
  if (filterOrigen === "NAC") ms = ms.filter(m => m.importado === false);
  if (filterImpacto === "PARADA")    ms = ms.filter(m => /^parada/i.test((m.impact||"").trim()));
  if (filterImpacto === "NO_PARADA") ms = ms.filter(m => (m.impact||"").trim() && !/^parada/i.test((m.impact||"").trim()));
  if (selGroup.size) ms = ms.filter(m => selGroup.has(m.mat_group));
  if (selLinea.size) ms = ms.filter(m => selLinea.has(m.linea_fab));
  if (search) {
    const s = search.toLowerCase();
    ms = ms.filter(m => m.mat.includes(s) || m.desc.toLowerCase().includes(s));
  }
  return sortMats(ms);
}

// ── Multi-select helpers ──────────────────────────────────────────────────────
function togglePanel(id) {
  const panel = document.getElementById(id);
  const isOpen = panel.style.display !== "none";
  document.querySelectorAll(".msel-panel").forEach(p => p.style.display = "none");
  if (!isOpen) {
    const btn = panel.previousElementSibling;
    const rect = btn.getBoundingClientRect();
    panel.style.top   = (rect.bottom + 3) + "px";
    panel.style.left  = rect.left + "px";
    panel.style.width = rect.width + "px";
    panel.style.display = "block";
  }
}
// Cerrar paneles al hacer clic fuera o al hacer scroll en el panel de filtros
document.addEventListener("click", e => {
  if (!e.target.closest(".msel-wrap"))
    document.querySelectorAll(".msel-panel").forEach(p => p.style.display = "none");
});
document.getElementById("filter-panel")?.addEventListener("scroll", () => {
  document.querySelectorAll(".msel-panel").forEach(p => p.style.display = "none");
});

function buildMselPanel(panelId, btnId, items, selSet, onChangeFn, searchPh) {
  const panel = document.getElementById(panelId);
  const btn   = document.getElementById(btnId);
  const ph    = searchPh || "Buscar…";
  const allChecked = selSet.size === 0;
  const searchBar = `<div class="msel-search">
    <span class="msel-search-icon">&#128269;</span>
    <input type="text" placeholder="${ph}" oninput="filterMselPanel('${panelId}',this.value)" onclick="event.stopPropagation()">
  </div>`;
  const allRow = `<label class="msel-item${allChecked?" checked":""}" style="border-bottom:1px solid var(--border,#2e2e2e);margin-bottom:2px">
    <input type="checkbox" ${allChecked?"checked":""} onchange="${onChangeFn}(null)">
    <span style="font-weight:600">All</span>
  </label>`;
  const rows = items.map(v =>
    `<label class="msel-item ${selSet.has(v)?"checked":""}">` +
      `<input type="checkbox" value="${v}" ${selSet.has(v)?"checked":""} onchange="${onChangeFn}(this)">` +
      `<span>${v}</span>` +
    `</label>`
  ).join("");
  panel.innerHTML = searchBar + `<div class="msel-items">${allRow}${rows}</div>`;
  // Si el botón usa la estructura centro-dd-btn (tiene .dd-label), actualizar solo el span
  const ddLbl = btn.querySelector(".dd-label");
  if (ddLbl) {
    ddLbl.textContent = selSet.size
      ? `${selSet.size} selected`
      : (btn.dataset.placeholder || "");
  } else {
    btn.className = "msel-btn" + (selSet.size ? " has-sel" : "");
    btn.innerHTML = selSet.size
      ? `${selSet.size} selected <span class="msel-arrow">&#9660;</span>`
      : (btn.dataset.placeholder||"") + ` <span class="msel-arrow">&#9660;</span>`;
  }
}

function filterMselPanel(panelId, q) {
  const lq = q.toLowerCase();
  document.querySelectorAll(`#${panelId} .msel-items .msel-item`).forEach(el => {
    el.style.display = el.querySelector("span").textContent.toLowerCase().includes(lq) ? "" : "none";
  });
}

function toggleGroupItem(el) {
  if (!el) { selGroup.clear(); }
  else if (el.checked) selGroup.add(el.value);
  else selGroup.delete(el.value);
  selMat.clear(); selSemMat = null;
  updateFilterBadge();
  renderGroupSelect(); renderList(); renderMainTab();
}

function toggleLineaItem(el) {
  if (!el) { selLinea.clear(); }
  else if (el.checked) selLinea.add(el.value);
  else selLinea.delete(el.value);
  selMat.clear(); selSemMat = null;
  updateFilterBadge();
  renderLineaSelect(); renderList(); renderMainTab();
}

function renderGroupSelect() {
  const groups = [...new Set([...centrosSel].flatMap(k => DATA.centros[k]?.mat_groups || []))].sort();
  selGroup.forEach(v => { if (!groups.includes(v)) selGroup.delete(v); });
  const btn = document.getElementById("group-btn");
  if (btn) btn.dataset.placeholder = "All groups";
  buildMselPanel("group-panel", "group-btn", groups, selGroup, "toggleGroupItem", "Search group…");
}

function renderLineaSelect() {
  const lineas = [...new Set([...centrosSel].flatMap(k => DATA.centros[k]?.lineas_fab || []))].sort();
  selLinea.forEach(v => { if (!lineas.includes(v)) selLinea.delete(v); });
  const btn = document.getElementById("linea-btn");
  if (btn) btn.dataset.placeholder = "All lines";
  buildMselPanel("linea-panel", "linea-btn", lineas, selLinea, "toggleLineaItem", "Search line…");
}

function sortMats(ms) {
  const cp = [...ms];
  const ord = {CRITICO:0, RIESGO:1, ALERTA:2, OK:3, SIN_CONSUMO:4};
  if (sortKey === "estado")    return cp.sort((a,b) => (ord[a.estado]||2)-(ord[b.estado]||2) || a.cob_hoy-b.cob_hoy);
  if (sortKey === "cobertura") return cp.sort((a,b) => a.cob_hoy - b.cob_hoy);
  if (sortKey === "codigo")    return cp.sort((a,b) => a.mat.localeCompare(b.mat));
  if (sortKey === "nombre")    return cp.sort((a,b) => a.desc.localeCompare(b.desc));
  return cp;
}

// ── Country pills ─────────────────────────────────────────────────────────────
function renderCountryBar() {
  const el = document.getElementById("country-bar");
  if (!el) return;
  const paises = DATA.meta.paises || {};
  const pList  = Object.keys(paises).filter(p => (paises[p].centros||[]).some(k => DATA.centros[k]));
  if (pList.length <= 1) { el.innerHTML = ""; return; }
  const FLAG_EMOJI = {AR:"🇦🇷",BR:"🇧🇷",CO:"🇨🇴",PE:"🇵🇪",CL:"🇨🇱"};
  const totAll = Object.keys(DATA.centros).reduce((s,k)=>s+(DATA.centros[k]?.materiales?.length||0),0);
  const allBtn = `<button class="cpill ${filterPais==="ALL"?"active":""}" onclick="setFilterPais('ALL')">&#127758; ALL <span class="cpill-cnt">${totAll}</span></button>`;
  const paisBtns = pList.map(p => {
    const pm  = paises[p];
    const emo = FLAG_EMOJI[pm.flag] || pm.flag || "";
    const cnt = (pm.centros||[]).filter(k=>DATA.centros[k]).reduce((s,k)=>s+(DATA.centros[k]?.materiales?.length||0),0);
    return `<button class="cpill ${filterPais===p?"active":""}" onclick="setFilterPais('${p}')">${emo} ${p} <span class="cpill-cnt">${cnt}</span></button>`;
  }).join("");
  el.innerHTML = allBtn + paisBtns;
}

function setFilterPais(pais) {
  filterPais = pais;
  if (pais === "ALL") {
    centrosSel = new Set(Object.keys(DATA.centros));
  } else {
    const pm = (DATA.meta.paises||{})[pais];
    if (pm && pm.centros.length) centrosSel = new Set(pm.centros.filter(k => DATA.centros[k]));
  }
  if (_centroDDOpen) renderCentroDDPanel();
  _refreshCentro();
}

// ── Centro dropdown ───────────────────────────────────────────────────────────
function updateCentroLabel() {
  const lbl = document.getElementById("centro-dd-label");
  if (!lbl) return;
  const allKeys = Object.keys(DATA.centros);
  if (centrosSel.size === 0 || centrosSel.size === allKeys.length) {
    lbl.textContent = "All plants";
  } else if (centrosSel.size === 1) {
    const k = [...centrosSel][0];
    lbl.textContent = DATA.centros[k]?.nombre || k;
  } else {
    const names = [...centrosSel].map(k => DATA.centros[k]?.nombre || k);
    lbl.textContent = names.slice(0,2).join(", ") + (names.length > 2 ? ` (+${names.length-2})` : "");
  }
}

function renderCentroDDPanel() {
  const scroll = document.getElementById("centro-dd-scroll");
  if (!scroll) return;
  const paises  = DATA.meta.paises || {};
  const q       = (document.getElementById("centro-dd-search")?.value || "").trim().toLowerCase();
  const allKeys = Object.keys(DATA.centros);
  const allSel  = allKeys.every(k => centrosSel.has(k));
  const FLAG_EMOJI = {AR:"🇦🇷",BR:"🇧🇷",CO:"🇨🇴",PE:"🇵🇪",CL:"🇨🇱"};

  // Countries to show: if a country pill is active, show only that country
  const paisList = filterPais === "ALL"
    ? Object.keys(paises).filter(p => (paises[p].centros||[]).some(k => DATA.centros[k]))
    : [filterPais].filter(p => paises[p]);

  let html = `<label class="centro-dd-all">
    <input type="checkbox" id="dd-all-cb" ${allSel?"checked":""} onchange="toggleAllCentroInDD(this)">
    <span>All plants</span>
    <span class="dd-mat-cnt">${allKeys.length} plants</span>
  </label>`;

  for (const pais of paisList) {
    const pm   = paises[pais] || {};
    const keys = (pm.centros||[]).filter(k => DATA.centros[k]);
    const filt = q ? keys.filter(k => DATA.centros[k].nombre.toLowerCase().includes(q)) : keys;
    if (!filt.length) continue;
    const emo  = FLAG_EMOJI[pm.flag] || pm.flag || "";
    html += `<div class="centro-dd-grp-hdr">${emo} ${pais}</div>`;
    for (const k of filt) {
      const cd    = DATA.centros[k];
      const chk   = centrosSel.has(k);
      const ncrit = cd.materiales.filter(m=>m.estado==="CRITICO").length;
      html += `<label class="centro-dd-item">
        <input type="checkbox" ${chk?"checked":""} onchange="toggleCentroInDD('${k}',this)">
        <span>${cd.nombre}</span>
        ${ncrit>0?`<span class="dd-crit">${ncrit}</span>`:""}
        <span class="dd-mat-cnt">${cd.materiales.length}</span>
      </label>`;
    }
  }
  scroll.innerHTML = html;
}

function toggleCentroDDPanel(e) {
  e && e.stopPropagation();
  const panel = document.getElementById("centro-dd-panel");
  const btn   = document.getElementById("centro-dd-btn");
  _centroDDOpen = !_centroDDOpen;
  if (_centroDDOpen) {
    const rect = btn.getBoundingClientRect();
    panel.style.top  = (rect.bottom + 4) + "px";
    panel.style.left = rect.left + "px";
    panel.classList.add("open");
    btn.classList.add("open");
    renderCentroDDPanel();
    setTimeout(() => document.getElementById("centro-dd-search")?.focus(), 50);
  } else {
    panel.classList.remove("open");
    btn.classList.remove("open");
  }
}

document.addEventListener("click", e => {
  if (_centroDDOpen && !document.getElementById("centro-dd-panel")?.contains(e.target)
      && !document.getElementById("centro-dd-btn")?.contains(e.target)) {
    _centroDDOpen = false;
    document.getElementById("centro-dd-panel")?.classList.remove("open");
    document.getElementById("centro-dd-btn")?.classList.remove("open");
  }
});

function toggleCentroInDD(k, el) {
  if (el.checked) { centrosSel.add(k); }
  else { if (centrosSel.size > 1) centrosSel.delete(k); else { el.checked = true; return; } }
  const allKeys = Object.keys(DATA.centros);
  const allCb = document.getElementById("dd-all-cb");
  if (allCb) allCb.checked = allKeys.every(k => centrosSel.has(k));
  updateCentroLabel();
  _softRefreshCentro();
}

function toggleAllCentroInDD(el) {
  const keys = filterPais === "ALL"
    ? Object.keys(DATA.centros)
    : ((DATA.meta.paises||{})[filterPais]?.centros||[]).filter(k => DATA.centros[k]);
  if (el.checked) { keys.forEach(k => centrosSel.add(k)); }
  else { centrosSel = new Set([keys[0] || Object.keys(DATA.centros)[0]]); }
  renderCentroDDPanel();
  updateCentroLabel();
  _softRefreshCentro();
}

function centroDDFilter() { renderCentroDDPanel(); }

function _softRefreshCentro() {
  // Refresh data without resetting filters or closing dropdown
  selMat = new Set(); selSemMat = null; selSemCk = null;
  const names = [...centrosSel].map(k => DATA.centros[k].nombre).join(" + ");
  document.getElementById("meta-centro-nom").textContent = names;
  renderGroupSelect(); renderLineaSelect(); renderKPIs(); renderList(); renderMainTab();
}

// ── Controls ─────────────────────────────────────────────────────────────────
function _refreshCentro() {
  selMat = new Set(); selSemMat = null; selSemCk = null; selGroup = new Set(); selLinea = new Set();
  filterOrigen = "ALL"; filterImpacto = "ALL";
  document.querySelectorAll("#fg-origen .filter-imp-btn").forEach((b,i) => b.classList.toggle("active", i===0));
  document.querySelectorAll("#fg-impacto .filter-imp-btn").forEach((b,i) => b.classList.toggle("active", i===0));
  const names = [...centrosSel].map(k => DATA.centros[k].nombre).join(" + ");
  document.getElementById("meta-centro-nom").textContent = names;
  renderCountryBar(); updateCentroLabel(); renderGroupSelect(); renderLineaSelect(); renderKPIs(); renderList(); renderMainTab();
}

function centroClick(k, e) {
  if (e && (e.ctrlKey || e.metaKey)) {
    // Ctrl+Clic: agregar/quitar del multi-selección
    if (centrosSel.has(k)) {
      if (centrosSel.size === 1) return;
      centrosSel.delete(k);
    } else {
      centrosSel.add(k);
    }
  } else {
    // Clic simple: seleccionar solo este
    centrosSel = new Set([k]);
  }
  _refreshCentro();
}

function toggleCentro(k) { centrosSel.has(k) ? (centrosSel.size>1 && centrosSel.delete(k)) : centrosSel.add(k); _refreshCentro(); }

function selectAllCentros() { setFilterPais("ALL"); }

// Alias para compatibilidad con cualquier referencia existente
function setCentro(c) { centrosSel = new Set([c]); _refreshCentro(); }

function setFilter(f, el) {
  filter = f;
  document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
  el.classList.add("active");
  updateFilterBadge();
  renderList(); renderMainTab();
}

function setFilterOrigen(f, el) {
  filterOrigen = f;
  document.querySelectorAll("#fg-origen .filter-imp-btn").forEach(b => b.classList.remove("active"));
  el.classList.add("active");
  updateFilterBadge();
  renderList(); renderMainTab();
}

function setFilterImpacto(f, el) {
  filterImpacto = f;
  document.querySelectorAll("#fg-impacto .filter-imp-btn").forEach(b => b.classList.remove("active"));
  el.classList.add("active");
  updateFilterBadge();
  renderList(); renderMainTab();
}

// ── Filtros sidebar: colapsables + badge ─────────────────────────────────────
function toggleFG(hdr) {
  const body  = hdr.nextElementSibling;
  const arrow = hdr.querySelector(".fg-arrow");
  const isCollapsed = body.classList.toggle("collapsed");
  arrow.classList.toggle("collapsed", isCollapsed);
}

function updateFilterBadge() {
  const hasEstado   = filter !== "ALL";
  const hasOrigen   = filterOrigen !== "ALL";
  const hasImpacto  = filterImpacto !== "ALL";
  const hasGrupo    = selGroup.size > 0;
  const hasLinea    = selLinea.size > 0;
  const n = (hasEstado?1:0) + (hasOrigen?1:0) + (hasImpacto?1:0) + (hasGrupo?1:0) + (hasLinea?1:0);
  const badge = document.getElementById("active-filters-badge");
  const txt   = document.getElementById("afb-text");
  if (!badge) return;
  if (n === 0) { badge.style.display = "none"; return; }
  badge.style.display = "flex";
  txt.textContent = n + " filtro" + (n > 1 ? "s" : "") + " activo" + (n > 1 ? "s" : "");
}

function clearAllFilters() {
  filter = "ALL"; filterOrigen = "ALL"; filterImpacto = "ALL";
  selGroup.clear(); selLinea.clear();
  const srchInput = document.getElementById("search-input");
  if (srchInput) { srchInput.value = ""; search = ""; }
  document.querySelectorAll(".filter-btn").forEach(b => b.classList.remove("active"));
  const fbAll = document.getElementById("fb-ALL"); if (fbAll) fbAll.classList.add("active");
  document.querySelectorAll("#fg-origen .filter-imp-btn").forEach(b => b.classList.remove("active"));
  const foAll = document.getElementById("fo-ALL"); if (foAll) foAll.classList.add("active");
  document.querySelectorAll("#fg-impacto .filter-imp-btn").forEach(b => b.classList.remove("active"));
  const fiAll = document.getElementById("fi-ALL"); if (fiAll) fiAll.classList.add("active");
  updateFilterBadge();
  renderList(); renderMainTab();
}

function toggleMobSidebar() {
  const sb  = document.querySelector(".sidebar");
  const btn = document.getElementById("mob-toggle-btn");
  sb.classList.toggle("mob-open");
  btn.textContent = sb.classList.contains("mob-open") ? "\u2715 Cerrar" : "\u2630 Materiales";
}

function setSearch(v) { search = v; renderList(); renderMainTab(); }
function setSort(k)   { sortKey = k; renderList(); renderMainTab(); }

function setTab(t, el) {
  mainTab = t;
  document.querySelectorAll(".tab-btn").forEach(b => b.classList.remove("active"));
  document.querySelectorAll(".tab-content").forEach(c => c.classList.add("hidden"));
  if(el) el.classList.add("active");
  document.getElementById("tab-"+t).classList.remove("hidden");
  renderMainTab();
}

function setOCTab(t) { ocTab = t; renderOCSol(); }

function selectMat(mat) {
  if (selMat.has(mat)) selMat.delete(mat);
  else selMat.add(mat);
  selSemMat = selMat.size === 1 ? [...selMat][0] : null;
  // En móvil: cerrar sidebar al seleccionar material
  if (window.innerWidth <= 768) {
    document.querySelector(".sidebar")?.classList.remove("mob-open");
    const btn = document.getElementById("mob-toggle-btn");
    if (btn) btn.textContent = "\u2630 Materiales";
  }
  renderList();
  renderMainTab();
}

// ── Render: Centro bar (now just updates dropdown label) ─────────────────────
function renderCentroBar() { updateCentroLabel(); }

// ── Render: KPI strip ────────────────────────────────────────────────────────
function renderKPIs() {
  const ms      = mats();
  const ncrit   = ms.filter(m => m.estado==="CRITICO").length;
  const nrisk   = ms.filter(m => m.estado==="RIESGO").length;
  const nok     = ms.filter(m => m.estado==="OK").length;
  const nsc     = ms.filter(m => m.estado==="SIN_CONSUMO").length;
  const q30     = 0;
  const q60     = 0;

  const oc  = allOC();
  const sol = allSOL();
  const {atrasadas, porVencer, vence7, resto} = calcOCStats(oc);
  const solSL  = sol.filter(s => s.sin_liberar).length;
  const solLib = sol.length - solSL;

  document.getElementById("kpi-strip").innerHTML = `
    <div class="kpi-grp-lbl">&#128202; INVENTORY</div>
    <div class="kpi ${ncrit>0?"kpi-r":""}"><div class="kpi-v">${ncrit}</div><div class="kpi-l">CRITICAL</div></div>
    <div class="kpi ${nrisk>0?"kpi-y":""}"><div class="kpi-v">${nrisk}</div><div class="kpi-l">AT RISK</div></div>
    <div class="kpi"><div class="kpi-v">${nok}</div><div class="kpi-l">NO RISK</div></div>
    <div class="kpi ${q30>0?"kpi-r":""}"><div class="kpi-v">${q30}</div><div class="kpi-l">Shortage &le;30d no PO</div></div>
    <div class="kpi ${q60>0?"kpi-y":""}"><div class="kpi-v">${q60}</div><div class="kpi-l">Shortage &le;60d no PO</div></div>
    ${nsc>0?`<div class="kpi kpi-sc"><div class="kpi-v">${nsc}</div><div class="kpi-l">NO DEMAND</div></div>`:""}
    <div class="kpi-sep"></div>
    <div class="kpi-grp-lbl">&#x1F4E6; POs</div>
    <div class="kpi ${atrasadas.length>0?"kpi-r":""}">
      <div class="kpi-v">${atrasadas.length}<span class="kpi-sub"> pos.</span></div>
      <div class="kpi-l">Overdue</div>
    </div>
    <div class="kpi ${porVencer.length>0?"kpi-y":""}">
      <div class="kpi-v">${porVencer.length}<span class="kpi-sub"> pos.</span></div>
      <div class="kpi-l">Due &le;21d &mdash; <b>${vence7.length}</b> in &le;7d</div>
    </div>
    <div class="kpi">
      <div class="kpi-v">${resto}<span class="kpi-sub"> pos.</span></div>
      <div class="kpi-l">&gt;21d</div>
    </div>
    <div class="kpi-sep"></div>
    <div class="kpi-grp-lbl">&#x1F4CB; PRs</div>
    <div class="kpi ${solSL>0?"kpi-r":"kpi-g"}">
      <div class="kpi-v">${solSL}</div>
      <div class="kpi-l">Unreleased</div>
    </div>
    <div class="kpi kpi-g">
      <div class="kpi-v">${solLib}</div>
      <div class="kpi-l">Released</div>
    </div>` ;
}

// ── Helper: referencia de fechas y stats OC ──────────────────────────────────
function getOCDateRef() {
  const ref = new Date(DATA.meta.fecha_saldo.replace(/(\d{2})-([A-Za-z]{3})-(\d{4})/, (_, d, m, y) => {
    const mo = {Ene:0,Feb:1,Mar:2,Abr:3,May:4,Jun:5,Jul:6,Ago:7,Sep:8,Oct:9,Nov:10,Dic:11};
    return `${y}-${String(mo[m]+1).padStart(2,"0")}-${d}`;
  }));
  const d21 = new Date(ref); d21.setDate(d21.getDate() + 21);
  const d7  = new Date(ref); d7.setDate(d7.getDate() + 7);
  return {ref, d7, d21};
}

function calcOCStats(oc) {
  const {ref, d7, d21} = getOCDateRef();
  const atrasadas = oc.filter(o => o.atrasada);
  const porVencer = oc.filter(o => {
    if (!o.fecha_entrega || o.atrasada) return false;
    const fe = new Date(o.fecha_entrega);
    return fe > ref && fe <= d21;
  });
  const vence7 = porVencer.filter(o => new Date(o.fecha_entrega) <= d7);
  return {atrasadas, porVencer, vence7, resto: oc.length - atrasadas.length - porVencer.length};
}


// ── Render: Sidebar list ─────────────────────────────────────────────────────
function renderList() {
  // En multi-centro: deduplicar por código de material mostrando el peor estado
  let ms = getFiltered();
  if (isMulti()) {
    const ord = {CRITICO:0, RIESGO:1, ALERTA:2, OK:3, SIN_CONSUMO:4};
    const dedupMap = {};
    ms.forEach(m => {
      const prev = dedupMap[m.mat];
      if (!prev || (ord[m.estado]??5) < (ord[prev.estado]??5)) dedupMap[m.mat] = m;
    });
    ms = sortMats(Object.values(dedupMap));
  }
  const dc = {CRITICO:"dot-r", RIESGO:"dot-y", ALERTA:"dot-a", OK:"dot-g", SIN_CONSUMO:"dot-sc"};
  const tc = {CRITICO:"dot-r-text", RIESGO:"dot-y-text", ALERTA:"dot-a-text", OK:"dot-g-text", SIN_CONSUMO:"dot-sc-text"};
  document.getElementById("mat-list").innerHTML = ms.map(m => {
    const mini = m.sabados.slice(0,8).map(s => `<span class="mini-d ${s.color}"></span>`).join("");
    const rtag = (m.quiebre_12sem && m.estado==="CRITICO")
      ? `<span class="quiebre-tag">&#9670; SHORTAGE &le;12W</span>`
      : (m.quiebre_12sem && m.estado==="ALERTA")
        ? `<span class="quiebre-lt-tag">&#9650; SHORTAGE + LT OK</span>`
        : (m.riesgo_12sem ? `<span class="riesgo-tag">&#9888; 12w</span>` : "");
    const itag = m.importado ? `<span class="imp-tag">IMP</span>` : `<span class="nal-tag">DOM</span>`;
    const ltag = m.linea_fab    ? `<span class="linea-tag">${m.linea_fab}</span>` : "";
    const isStop = m.impact && /^parada/i.test(m.impact.trim());
    const impLine = m.impact
      ? `<div class="mat-impact" title="${m.action||""}">&#9888; <span class="${isStop?"impact-stop":"impact-ok"}">${m.impact}</span>${m.action ? ` &nbsp;· <span style="color:var(--tx3)">${m.action}</span>` : ""}</div>`
      : "";
    return `<div class="mat-item ${selMat.has(m.mat)?"active":""}" onclick="selectMat('${m.mat}')">
      <span class="dot ${dc[m.estado]||"dot-g"}"></span>
      <div class="mat-info">
        <div class="mat-code">${m.mat} <span class="mat-um">${m.um}</span>${rtag}${itag}</div>
        <div class="mat-desc" title="${m.desc}">${m.desc||"—"}</div>
        <div class="mini-row">${mini}${ltag}</div>
        ${impLine}
      </div>
      <div class="mat-right">
        <div class="mat-cob ${tc[m.estado]||"dot-g-text"}">${m.estado==="SIN_CONSUMO"?"Inv.":m.adu_plan>0?(m.cob_hoy>998?"&#x221e;":m.cob_hoy+"d"):"—"}</div>
      </div>
    </div>`;
  }).join("") || `<div class="empty">No results</div>`;
  const selTxt = selMat.size ? ` · ${selMat.size} selected` : "";
  const multiTxt = isMulti() ? ` <span style="font-size:9px;color:var(--tx3)">(${centrosSel.size} plants)</span>` : "";
  document.getElementById("filter-count").innerHTML =
    `${ms.length} materials${multiTxt}${selTxt}` +
    (selMat.size ? ` <span style="cursor:pointer;color:var(--red);margin-left:6px" title="Clear selection" onclick="selMat=new Set();selSemMat=null;renderList();renderMainTab()">&#10005;</span>` : "");
}

// ── Render: Inventory table ──────────────────────────────────────────────────
function renderInventario() {
  let ms     = getFiltered();
  if (selMat.size) ms = ms.filter(m => selMat.has(m.mat));
  const multi = isMulti();
  const sabs = DATA.meta.sem_weeks;
  const sabH = sabs.map(w => `<th class="th-sat num">${w.label}</th>`).join("");
  // Agrupar semanas por mes para header doble (igual que Semáforo)
  const invMGroups = [];
  sabs.forEach(w => {
    if (!invMGroups.length || invMGroups[invMGroups.length-1].m !== w.month)
      invMGroups.push({m: w.month, cnt: 1});
    else
      invMGroups[invMGroups.length-1].cnt++;
  });
  const invMonthHdr = invMGroups.map(g =>
    `<th colspan="${g.cnt}" class="cal-month-hdr">${g.m}</th>`).join("");
  const rows = ms.map(m => {
    const cc  = m.estado==="CRITICO"?"crit":m.estado==="RIESGO"?"risk":"";
    const scv = m.sabados.map(s => {
      const cls = s.color==="R"?"sat-r":s.color==="Y"?"sat-y":s.color==="K"?"sat-k":s.color==="B"?"sat-b":"sat-g";
      const val = s.color==="K"?"&#10005;":m.adu>0?(s.cob>998?"&#x221e;":s.cob+"d"):"—";
      return `<td class="num ${cls}" title="Stock ${fmt(s.stock)} &middot; ${s.cob}d">${val}</td>`;
    }).join("");
    const zc  = zBadge(m.zona_b2);
    const cCol = multi ? `<td class="col-centro"><span class="centro-badge">${m._cn}</span></td>` : "";
    return `<tr data-mat="${m.mat}" class="${selMat.has(m.mat)?"sel":""}" onclick="selectMat('${m.mat}')">
      ${cCol}
      <td class="sc col-mat mono">${m.mat}</td>
      <td class="sc col-desc" title="${m.desc}">${m.desc||"—"}</td>
      <td>${m.um}</td>
      <td class="num">${fmt(m.inv_ini)}</td>
      <td class="num col-ing">+${fmt(m.ingresos)}</td>
      <td class="num col-con">−${fmt(m.consumos)}</td>
      <td class="num col-sal fw">${fmt(m.saldo)}</td>
      <td class="num" style="color:var(--tx2)">${m.adu>0?fmt1(m.adu):"—"}</td>
      <td class="num" style="color:${m.adu_hist>0?(m.adu>0&&m.adu_hist>m.adu*1.2?'var(--red)':m.adu>0&&m.adu_hist<m.adu*0.8?'var(--grn)':'var(--tx2)'):'var(--tx3)'}" title="Consumo real últimos 3 meses">${m.adu_hist>0?fmt1(m.adu_hist):"—"}</td>
      <td class="num cob-cell ${cc}">${m.estado==="SIN_CONSUMO"?"Inv.":m.adu_plan>0?(m.cob_hoy>998?"∞":m.cob_hoy+"d"):"—"}</td>
      ${scv}
      <td class="zona"><span class="zona-badge ${zc.cls}">${zc.lbl}</span></td>
    </tr>`;
  }).join("") || `<tr><td colspan="100" class="empty">No materials match current filters</td></tr>`;

  // ── Totales por UM ──────────────────────────────────────────────────────────
  const umMap = {};
  ms.forEach(m => {
    const u = m.um || "—";
    if (!umMap[u]) umMap[u] = {n: 0, inv_ini: 0, ingresos: 0, consumos: 0, saldo: 0, adu: 0};
    umMap[u].n++;
    umMap[u].inv_ini  += m.inv_ini;
    umMap[u].ingresos += m.ingresos;
    umMap[u].consumos += m.consumos;
    umMap[u].saldo    += m.saldo;
    umMap[u].adu      += m.adu;
  });
  const footRows = Object.entries(umMap)
    .sort((a, b) => a[0].localeCompare(b[0]))
    .map(([um, t]) => {
      const cob = t.adu > 0 ? Math.round(t.saldo / t.adu) + "d" : "—";
      return `<tr>
        <td class="tot-label">${t.n} materials</td>
        <td class="tot-label">TOTAL</td>
        <td class="tot-um">${um}</td>
        <td class="num tot-v">${fmt(t.inv_ini)}</td>
        <td class="num tot-v col-ing">+${fmt(t.ingresos)}</td>
        <td class="num tot-v col-con">−${fmt(t.consumos)}</td>
        <td class="num tot-v col-sal fw">${fmt(t.saldo)}</td>
        <td class="num tot-v">${t.adu > 0 ? fmt1(t.adu) : "—"}</td>
        <td></td>
        <td class="num tot-v">${cob}</td>
        <td colspan="${sabs.length + 1}"></td>
      </tr>`;
    }).join("");

  // ── Barra diagnóstico inventario ──────────────────────────────────────────
  const allMs  = getFiltered();
  const nCrit  = allMs.filter(m=>m.estado==="CRITICO").length;
  const nRisk  = allMs.filter(m=>m.estado==="RIESGO").length;
  const nQ30   = 0;
  const diagInv =
    `<div class="tab-diag-bar">` +
    (nCrit>0 ? `<span class="tdb-chip tdb-r">&#9660; ${nCrit} CRITICAL</span>` : "") +
    (nRisk>0 ? `<span class="tdb-chip tdb-y">&#9650; ${nRisk} at risk</span>` : "") +
    (nQ30>0  ? `<span class="tdb-chip tdb-r">&#9888; ${nQ30} stockout in &le;30d no PO</span>` : "") +
    (nCrit===0&&nRisk===0 ? `<span class="tdb-chip tdb-g">&#10003; No criticals or risks</span>` : "") +
    `</div>`;

  const cHdr1 = multi ? `<th class="col-centro" rowspan="2">Plant</th>` : "";
  document.getElementById("tab-inv").innerHTML =
    diagInv +
    `<div class="tbl-toolbar"><button class="copy-btn" onclick="copyTable(document.getElementById('tbl-inv'),'Inventory')">⧉ Copy table</button></div>` +
    `<div class="table-wrap"><table id="tbl-inv" class="inv-table">
    <thead>
    <tr>
      ${cHdr1}
      <th class="sc col-mat" rowspan="2" onclick="setSort('codigo')">Material</th>
      <th class="sc col-desc" rowspan="2" style="position:relative" onclick="setSort('nombre')">Description ↕<span class="col-desc-resize" id="desc-resize" onmousedown="startDescResize(event)"></span></th>
      <th rowspan="2">UM</th>
      <th class="num" rowspan="2" title="Inventario inicio de mes">Ini. Inv.</th>
      <th class="num col-ing" rowspan="2" title="Ingresos del mes a la fecha">+ Receipts</th>
      <th class="num col-con" rowspan="2" title="Consumos del mes a la fecha">− Consumption</th>
      <th class="num col-sal" rowspan="2" title="Saldo actual">= Stock</th>
      <th class="num" rowspan="2" title="ADU B2Wise (planeación DDMRP)">ADU B2W</th>
      <th class="num" rowspan="2" title="ADU real: promedio últimos 3 meses de consumo">ADU 3M</th>
      <th class="num" rowspan="2" onclick="setSort('cobertura')" title="Días de cobertura hoy">Cov. Today ↕</th>
      ${invMonthHdr}
      <th rowspan="2">Zona B2W</th>
    </tr>
    <tr>
      ${sabH}
    </tr>
    </thead>
    <tbody>${rows}</tbody>
    <tfoot>${footRows}</tfoot>
  </table></div>`;
  renderSemPanel();
}

function setOCSort(v) { ocSort = v; renderOCSol(); }

// ── Render: OC & SOLPEDs ─────────────────────────────────────────────────────
function renderOCSol() {
  const flt  = search.toLowerCase();
  let oc     = getOC().filter(o => (!selMat.size || selMat.has(o.mat)) && (!flt || o.mat.includes(flt) || o.desc.toLowerCase().includes(flt)));
  let sol    = getSOL().filter(s => (!selMat.size || selMat.has(s.mat)) && (!flt || s.mat.includes(flt) || s.desc.toLowerCase().includes(flt)));
  if (filterSolLib === "SIN_LIB") sol = sol.filter(s => s.sin_liberar);
  if (filterSolLib === "LIB")     sol = sol.filter(s => !s.sin_liberar);
  const ocAtras  = getOC().filter(o => o.atrasada).length;
  const solSL    = getSOL().filter(s => s.sin_liberar).length;

  // Ordenar OC según selección
  if (ocSort === "fecha_asc")        oc = [...oc].sort((a,b) => (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""));
  else if (ocSort === "fecha_desc")  oc = [...oc].sort((a,b) => (b.fecha_entrega||"").localeCompare(a.fecha_entrega||""));
  else if (ocSort === "atraso_desc") oc = [...oc].sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return b.dias_atraso-a.dias_atraso; });
  else if (ocSort === "estado")      oc = [...oc].sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""); });

  const multi = isMulti();
  const ocRows = oc.map(o =>
    `<tr class="${o.atrasada?"row-atraso":""}">
      ${multi?`<td class="col-centro"><span class="centro-badge">${o._cn}</span></td>`:""}
      <td class="mono">${o.doc}</td>
      <td class="mono">${o.mat}</td>
      <td title="${o.desc}">${truncate(o.desc,50)}</td>
      <td class="num">${fmt(o.qty_pedida)} <span style="color:var(--tx3)">${o.um}</span></td>
      <td class="num fw">${fmt(o.qty_pend)} <span style="color:var(--tx3)">${o.um}</span></td>
      <td style="color:var(--tx2)">${o.fecha_doc}</td>
      <td class="${o.atrasada?"atraso-cell":""}">${o.fecha_entrega}${o.atrasada?` <span style="color:var(--tx3);font-size:10px">→ proj. ${o.fecha_proyectada}</span>`:""}</td>
      <td class="num ${o.atrasada?"atraso-cell":""}">${o.atrasada?"⚠ "+o.dias_atraso+"d":"—"}</td>
      <td style="color:var(--tx2)">${o.proveedor}</td>
    </tr>`).join("") || `<tr><td colspan="100" class="empty">No open POs</td></tr>`;

  const solRows = sol.map(s =>
    `<tr class="${s.sin_liberar?"row-sinlib":""}">
      ${multi?`<td class="col-centro"><span class="centro-badge">${s._cn}</span></td>`:""}
      <td class="mono">${s.doc}</td>
      <td class="mono">${s.mat}</td>
      <td title="${s.desc}">${truncate(s.desc,50)}</td>
      <td class="num">${fmt(s.qty)} <span style="color:var(--tx3)">${s.um}</span></td>
      <td style="color:var(--tx2)">${s.fecha_entrega}</td>
      <td><span class="lib-badge ${s.sin_liberar?"badge-sinlib":"badge-lib"}">${s.sin_liberar?"Unreleased":"Released"}</span></td>
    </tr>`).join("") || `<tr><td colspan="100" class="empty">No PRs</td></tr>`;

  // Mini resumen de estado para la pestaña
  const {atrasadas: ocAtrL, porVencer: ocVcrL, vence7: ocV7L, resto: ocRestoL} = calcOCStats(allOC());
  const solTot = allSOL().length;
  const solSLtot = allSOL().filter(s => s.sin_liberar).length;

  document.getElementById("tab-oc").innerHTML = `
    <div class="tab-oc-summary">
      <div class="tos-group">
        <span class="tos-lbl">&#x1F4E6; POs</span>
        <div class="tos-card tos-r"><span class="tos-v">${ocAtrL.length}</span><span class="tos-l">Atrasadas</span></div>
        <div class="tos-card tos-y"><span class="tos-v">${ocVcrL.length}</span><span class="tos-l">Vencen &le;21d <span style="color:var(--tx3)">(${ocV7L.length} en &le;7d)</span></span></div>
        <div class="tos-card tos-g"><span class="tos-v">${ocRestoL}</span><span class="tos-l">&gt;21d</span></div>
      </div>
      <div class="tos-sep"></div>
      <div class="tos-group">
        <span class="tos-lbl">&#x1F4CB; PRsPEDs</span>
        <div class="tos-card"><span class="tos-v">${solTot}</span><span class="tos-l">Abiertas</span></div>
        <div class="tos-card ${solSLtot>0?"tos-r":"tos-g"}"><span class="tos-v">${solSLtot}</span><span class="tos-l">Unreleased</span></div>
        <div class="tos-card tos-g"><span class="tos-v">${solTot - solSLtot}</span><span class="tos-l">Releaseds</span></div>
      </div>
    </div>
    <div class="oc-tabs">
      <button class="oc-tab-btn ${ocTab==="oc"?"active":""}" onclick="setOCTab('oc')">
        🚛 OC Abiertas <span class="cnt-badge">${getOC().length}</span>
        ${ocAtras>0?`<span class="badge-r">${ocAtras} atrasadas</span>`:""}
      </button>
      <button class="oc-tab-btn ${ocTab==="sol"?"active":""}" onclick="setOCTab('sol')">
        📋 SOLPEDs <span class="cnt-badge">${getSOL().length}</span>
        ${solSL>0?`<span class="badge-y">${solSL} sin lib.</span>`:""}
      </button>
    </div>
    ${ocTab==="oc"
    ? `<div class="tbl-toolbar">
         <button class="copy-btn" onclick="copyTable(document.getElementById('tbl-oc'),'OC Abiertas')">⧉ Copy POs</button>
         <button class="copy-btn" style="background:#1e6b3a;border-color:#1e6b3a;color:#fff" onclick="openOCExcelModal()">&#128202; Excel</button>
         <span style="margin-left:auto;display:flex;align-items:center;gap:6px;font-size:11px;color:var(--tx3)">
           Ordenar:
           <select class="sort-sel" onchange="setOCSort(this.value)">
             <option value="atraso_desc" ${ocSort==="atraso_desc"?"selected":""}>Atrasadas primero</option>
             <option value="estado"      ${ocSort==="estado"?"selected":""}>Estado (atr → OK)</option>
             <option value="fecha_asc"   ${ocSort==="fecha_asc"?"selected":""}>Fecha entrega ↑</option>
             <option value="fecha_desc"  ${ocSort==="fecha_desc"?"selected":""}>Fecha entrega ↓</option>
           </select>
         </span>
       </div>
       <div class="table-wrap"><table id="tbl-oc" class="oc-table">
        <thead><tr>
          ${multi?"<th class='col-centro'>Centro</th>":""}
          <th>N° OC</th><th>Material</th><th>Descripción</th>
          <th class="num">Cant. Pedida</th><th class="num">Pendiente</th>
          <th>Fecha Doc.</th><th>Fecha Entrega</th>
          <th class="num">Días Atraso</th><th>Proveedor</th>
        </tr></thead>
        <tbody>${ocRows}</tbody>
      </table></div>`
    : `<div class="tbl-toolbar">
         <button class="copy-btn" onclick="copyTable(document.getElementById('tbl-sol'),'SOLPEDs')">⧉ Copy PRs</button>
         <button class="copy-btn" style="background:#1e6b3a;border-color:#1e6b3a;color:#fff" onclick="openOCExcelModal()">&#128202; Excel</button>
         <span style="margin-left:12px;display:flex;align-items:center;gap:4px">
           <button class="sol-lib-btn filter-imp-btn ${filterSolLib==='ALL'?'active':''}" onclick="setFilterSolLib('ALL',this)">📋 Todas</button>
           <button class="sol-lib-btn filter-imp-btn ${filterSolLib==='SIN_LIB'?'active':''}" onclick="setFilterSolLib('SIN_LIB',this)" style="color:#f5c518">⚠ Unreleased</button>
           <button class="sol-lib-btn filter-imp-btn ${filterSolLib==='LIB'?'active':''}" onclick="setFilterSolLib('LIB',this)" style="color:#3ecf8e">✓ Releaseds</button>
         </span>
       </div>
       <div class="table-wrap"><table id="tbl-sol" class="oc-table">
        <thead><tr>
          ${multi?"<th class='col-centro'>Centro</th>":""}
          <th>N° SOLPED</th><th>Material</th><th>Descripción</th>
          <th class="num">Cantidad</th><th>Fecha Entrega</th><th>Estado Lib.</th>
        </tr></thead>
        <tbody>${solRows}</tbody>
      </table></div>`}`;
  renderSemPanel();
}

// ── Render: Semáforo calendario ───────────────────────────────────────────────
function setCalView(v) {
  calView = v;
  renderCal();
}

function renderCal() {
  let ms      = getFiltered();
  if (selMat.size) ms = ms.filter(m => selMat.has(m.mat));

  // ── Barra diagnóstico ────────────────────────────────────────────────────
  const nRojo   = ms.filter(m => m.sabados[0]?.color==="R").length;
  const nAmar   = ms.filter(m => m.sabados[0]?.color==="Y").length;
  const diagCal =
    `<div class="tab-diag-bar">` +
    (nRojo>0 ? `<span class="tdb-chip tdb-r">&#9660; ${nRojo} in red next week</span>` : "") +
    (nAmar>0 ? `<span class="tdb-chip tdb-y">&#9650; ${nAmar} in yellow next week</span>` : "") +
    (nRojo===0&&nAmar===0 ? `<span class="tdb-chip tdb-g">&#10003; No criticals next week</span>` : "") +
    `<div style="flex:1"></div>` +
    `<div class="view-toggle">` +
    `<button class="view-toggle-btn ${calView==="cal"?"active":""}" onclick="setCalView('cal')">&#128197; Calendar</button>` +
    `<button class="view-toggle-btn ${calView==="list"?"active":""}" onclick="setCalView('list')">&#128202; List</button>` +
    `</div></div>`;
  document.getElementById("cal-diag").innerHTML = diagCal;

  if (calView === "list") { renderCalList(ms); renderSemPanel(); return; }

  const weeks = DATA.meta.sem_weeks;
  const dc    = {CRITICO:"dot-r", RIESGO:"dot-y", ALERTA:"dot-a", OK:"dot-g"};
  const tc    = {CRITICO:"dot-r-text", RIESGO:"dot-y-text", ALERTA:"dot-a-text", OK:"dot-g-text"};

  // Agrupar semanas por mes para header doble
  const monthGroups = [];
  weeks.forEach(w => {
    if (!monthGroups.length || monthGroups[monthGroups.length-1].m !== w.month)
      monthGroups.push({m: w.month, cnt: 1});
    else monthGroups[monthGroups.length-1].cnt++;
  });
  const monthHdr = monthGroups.map(g =>
    `<th colspan="${g.cnt}" class="cal-month-hdr">${g.m}</th>`).join("");
  const weekHdr = weeks.map(w =>
    `<th class="cal-cell"><span class="cal-wk-num">S${w.week_num}</span>${w.label}</th>`).join("");

  // Símbolo por estado de dot
  const SYM = {G:"&#9679;", Y:"&#9679;", R:"&#9660;", K:"&#9888;", B:"&#9679;"};

  const multi = isMulti();
  const rows = ms.map(m => {
    const dots = m.sabados.map((s, i) =>
      `<td class="cal-cell" title="${weeks[i]?weeks[i].date:s.fecha}: stock ${fmt(s.stock)} &middot; ${s.cob}d">
        <span class="sd ${s.color}" onclick="semClick('${m.mat}')">${SYM[s.color]||SYM.G}</span></td>`
    ).join("");
    const sel = (m.mat===selSemMat && (!m._ck || !selSemCk || m._ck===selSemCk))?"sel-row":"";
    const cnTag = multi ? `<span class="cal-centro">${m._cn}</span>` : "";
    return `<tr>
      <td class="cal-mat-cell ${sel}" onclick="semClick('${m.mat}','${m._ck||""}')" style="cursor:pointer">
        <span class="dot ${dc[m.estado]||"dot-g"}"></span>
        <span class="cal-code">${m.mat}</span>${cnTag}
        <span class="cal-name" title="${m.desc}">${truncate(m.desc,20)}</span>
        <span class="cal-cob ${tc[m.estado]||"dot-g-text"}">${m.adu_plan>0?(m.cob_hoy>998?"&#x221e;":m.cob_hoy+"d"):"—"}</span>
        <span class="trend-ico ${trendArrow(m).cls}" title="${trendArrow(m).title}">${trendArrow(m).sym}</span>
      </td>${dots}</tr>`;
  }).join("") || `<tr><td colspan="100" class="empty">No materials</td></tr>`;

  document.getElementById("cal-table-wrap").innerHTML =
    `<table class="cal-table">
      <thead>
        <tr><th class="cal-mat-cell"></th>${monthHdr}</tr>
        <tr><th class="cal-mat-cell">Material &middot; Cov. Today</th>${weekHdr}</tr>
      </thead>
      <tbody>${rows}</tbody>
    </table>`;
  renderSemPanel();
}

// ── Vista lista del semáforo ─────────────────────────────────────────────────
function badgeEstado(e) {
  const map = {
    CRITICO: "<span class='badge-r' style='padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;color:#fff'>CR&Iacute;T.</span>",
    RIESGO:  "<span style='padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;background:var(--yel3);color:var(--yel)'>RIESGO</span>",
    ALERTA:  "<span style='padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;background:var(--blu2);color:var(--blu)'>ALERTA</span>",
    OK:      "<span style='padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;background:var(--grn2);color:var(--grn)'>OK</span>",
    SIN_CONSUMO:"<span style='padding:2px 6px;border-radius:10px;font-size:10px;font-weight:700;background:var(--bg3);color:var(--tx3)'>INV.SC</span>",
  };
  return map[e] || "";
}

function renderCalList(ms) {
  const validCob = ms.map(m=>m.cob_hoy).filter(v=>v>0&&v<998);
  const maxCob   = validCob.length ? Math.min(Math.max(...validCob), 120) : 60;
  const barColor = cob => cob<=0 ? "var(--red)" : cob<30 ? "var(--red)" : cob<60 ? "var(--yel)" : "var(--grn)";

  const sortedMs = [...ms].sort((a,b) => {
    if (a.cob_hoy<=0 && b.cob_hoy>0) return -1;
    if (a.cob_hoy>0 && b.cob_hoy<=0) return 1;
    if (a.cob_hoy>=998 && b.cob_hoy<998) return 1;
    if (a.cob_hoy<998 && b.cob_hoy>=998) return -1;
    return a.cob_hoy - b.cob_hoy;
  });

  const rows = sortedMs.map(m => {
    const cob    = m.cob_hoy;
    const pct    = cob<=0 ? 2 : cob>=998 ? 100 : Math.min(cob/maxCob*100, 100);
    const cobLbl = cob<=0 ? "Sin ADU" : cob>=998 ? "&#x221e;" : cob+"d";
    const t      = trendArrow(m);
    return `<tr style="border-bottom:1px solid var(--brd)">
      <td style="padding:5px 8px;font-family:'DM Mono',monospace;font-size:11px;color:var(--tx2);white-space:nowrap;width:110px">${m.mat}</td>
      <td style="padding:5px 8px;font-size:11px;max-width:200px;overflow:hidden;text-overflow:ellipsis;white-space:nowrap" title="${m.desc}">${truncate(m.desc,28)}</td>
      <td style="padding:5px 8px;width:220px">
        <div style="background:var(--bg3);border-radius:4px;height:10px;overflow:hidden">
          <div style="width:${pct}%;height:100%;background:${barColor(cob)};border-radius:4px;transition:.3s"></div>
        </div>
      </td>
      <td style="padding:5px 8px;width:55px;text-align:right;font-family:'DM Mono',monospace;font-size:12px;font-weight:700;color:${barColor(cob)}">${cobLbl}</td>
      <td class="${t.cls}" title="${t.title}" style="padding:5px 8px;width:24px;text-align:center;font-size:12px">${t.sym}</td>
      <td style="padding:5px 8px;width:90px">${badgeEstado(m.estado)}</td>
    </tr>`;
  }).join("") || `<tr><td colspan="6" class="empty">No materials</td></tr>`;

  document.getElementById("cal-table-wrap").innerHTML =
    `<table style="border-collapse:collapse;width:100%;font-size:12px">
      <thead><tr style="border-bottom:2px solid var(--brd)">
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0">C&oacute;digo</th>
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0">Descripci&oacute;n</th>
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0">Cobertura (m&aacute;x ${maxCob}d)</th>
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0;text-align:right">D&iacute;as</th>
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0">Tend.</th>
        <th style="padding:6px 8px;font-size:10px;text-transform:uppercase;letter-spacing:.4px;color:var(--tx3);font-weight:500;background:var(--bg2);position:sticky;top:0">Estado</th>
      </tr></thead>
      <tbody>${rows}</tbody>
    </table>`;
}

function renderMainTab() {
  if (mainTab==="inv") renderInventario();
  else if (mainTab==="oc") renderOCSol();
  else if (mainTab==="cal") renderCal();
}

// ── Semáforo: click + panel de detalle ───────────────────────────────────────
function semClick(mat, ck) {
  selSemMat = mat; selSemCk = ck || null; selMat = new Set([mat]);
  renderList();
  renderCal();   // re-renderiza tabla con fila seleccionada + llama renderSemPanel
}

function closeSemPanel() {
  selSemMat = null; selSemCk = null;
  selMat.clear();
  document.getElementById("sem-panel").classList.add("hidden");
  renderList();
  renderMainTab();
}

function setSemDTab(t) {
  semDTab = t;
  renderSemPanel();
}

function renderSemPanel() {
  const panel = document.getElementById("sem-panel");
  if (!selSemMat) { panel.classList.add("hidden"); return; }

  // En multi-centro: buscar el material en todos los centros seleccionados
  const mEntries = [...centrosSel]
    .map(k => ({ ck: k, cn: DATA.centros[k].nombre, m: (DATA.centros[k]?.materiales||[]).find(x => x.mat === selSemMat) }))
    .filter(x => x.m);
  if (!mEntries.length) { panel.classList.add("hidden"); return; }
  panel.classList.remove("hidden");

  // Para header/estado usar la peor instancia (o la del centro seleccionado en semáforo)
  const ord = {CRITICO:0, RIESGO:1, ALERTA:2, OK:3, SIN_CONSUMO:4};
  const m = mEntries.reduce((best, e) => (ord[e.m.estado]??5) < (ord[best.m.estado]??5) ? e : best).m;
  const multi = mEntries.length > 1;

  const eBadge = {CRITICO:"badge-r", RIESGO:"badge-y", ALERTA:"badge-a", OK:"badge-lib", SIN_CONSUMO:"badge-sc"};
  const ec     = eBadge[m.estado] || "badge-lib";
  const cobStr = m.adu_plan>0 ? (m.cob_hoy>998?"&#x221e;":m.cob_hoy+"d") : "—";

  // Contar OC y SOLPEDs de todos los centros
  const allOcList  = mEntries.flatMap(e => (e.m.oc_list||[]).map(o=>({...o,_cn:e.cn,_imp:e.m.importado,_um:e.m.um})));
  const allSolList = mEntries.flatMap(e => (e.m.sol_list||[]).map(s=>({...s,_cn:e.cn,_um:e.m.um})));

  document.getElementById("sdp-mat-title").textContent = m.mat + " \u2014 " + truncate(m.desc, 42);
  const quiebreLabel = (m.quiebre_12sem && m.estado==="CRITICO")
    ? ` <span class="quiebre-tag" style="font-size:10px">&#9670; SHORTAGE &le;12W</span>`
    : (m.quiebre_12sem && m.estado==="ALERTA")
      ? ` <span class="quiebre-lt-tag" style="font-size:10px">&#9650; SHORTAGE + LT OK${m.primer_quiebre_dias?` (${m.primer_quiebre_dias}d)`:""}</span>`
      : "";
  document.getElementById("sdp-estado").innerHTML =
    `<span class="lib-badge ${ec}">${m.estado}</span>${quiebreLabel}`;
  const ltStr = (m.lt && m.lt > 0) ? `&nbsp;&middot;&nbsp; LT: <b>${m.lt}d</b>` : "";
  // Info: si multi, mostrar stocks por centro
  const stockInfo = multi
    ? mEntries.map(e=>`<span style="color:var(--tx3);font-size:10px">${e.cn}:</span> <b>${fmt(e.m.saldo)}</b>`).join(" &nbsp;|&nbsp; ")
    : `Stock: <b>${fmt(m.saldo)}</b> ${m.um}`;
  document.getElementById("sdp-info").innerHTML =
    `${stockInfo} &nbsp;&middot;&nbsp; Cob: <b>${cobStr}</b> &nbsp;&middot;&nbsp; ADU: <b>${fmt1(m.adu)}/d</b>${ltStr} &nbsp;&middot;&nbsp; MRP M+0: <b>${fmt(m.mrp_m0)}</b>`;
  document.getElementById("sdp-tab-oc").textContent  = "OC en firme (" + allOcList.length  + ")";
  document.getElementById("sdp-tab-sol").textContent = "SOLPEDs ("     + allSolList.length + ")";

  // ── Impact / Action block ─────────────────────────────────────────────────
  const impEl = document.getElementById("sdp-impact");
  const hasImpact = mEntries.some(e => e.m.impact || e.m.action);
  if (!hasImpact) {
    impEl.style.display = "none";
  } else {
    impEl.style.display = "flex";
    const chips = mEntries.filter(e => e.m.impact || e.m.action).map(e => {
      const isStop = /^parada/i.test((e.m.impact||"").trim());
      const impLabel = e.m.impact
        ? `<span class="${isStop?"ichip-stop":"ichip-ok"}">${e.m.impact}</span>`
        : `<span style="color:var(--tx3)">Sin impacto</span>`;
      const actLabel = e.m.action
        ? `<span style="color:var(--tx2)">&nbsp;·&nbsp;${e.m.action}</span>`
        : "";
      const centroTag = mEntries.length > 1 ? `<b>${e.cn}:</b> ` : "";
      return `<div class="sdp-impact-chip">${centroTag}${impLabel}${actLabel}</div>`;
    }).join("");
    impEl.innerHTML = `<b style="font-size:10px;color:var(--tx3);margin-right:4px">&#9888; IMPACTO:</b>` + chips;
  }
  document.querySelectorAll(".sdp-tab").forEach(b => b.classList.remove("active"));
  document.getElementById("sdp-tab-" + semDTab).classList.add("active");

  let content = "";
  if (semDTab === "oc") {
    const ocRows = allOcList.map(o => {
      const umDiff = o.oc_um && o.oc_um !== (o._um||"").toUpperCase();
      const qtyLabel = umDiff
        ? `${fmt(o.qty)} <span style="color:var(--yel);font-size:10px">${o.oc_um}</span> <span style="color:var(--tx3);font-size:10px">= ${fmt(o.qty_base)} ${o._um}</span>`
        : `${fmt(o.qty)} <span style="color:var(--tx3)">${o._um}</span>`;
      const origTag = o._imp
        ? `<span class="imp-tag" style="font-size:9px">IMP</span>`
        : `<span style="font-size:9px;background:#2a4a2a;color:#7ecf7e;border-radius:3px;padding:1px 5px;font-weight:700">NAC</span>`;
      return `<tr class="${o.atrasada?"row-atraso":""}">
        ${multi?`<td class="col-centro"><span class="centro-badge">${o._cn}</span></td>`:""}
        <td class="num fw">${qtyLabel}</td>
        <td class="${o.atrasada?"atraso-cell":""}">${o.eta}${o.atrasada?`<br><span style="color:var(--tx3);font-size:10px">→ proy. ${o.eta_proy||""}</span>`:""}</td>
        <td class="mono">${o.doc}</td>
        <td style="color:var(--tx2);max-width:180px;overflow:hidden;text-overflow:ellipsis">${origTag} ${truncate(o.proveedor,30)}</td>
        <td class="${o.atrasada?"atraso-cell":""}">${o.atrasada?"&#9888; "+o.dias_atraso+"d atr.":""}</td>
      </tr>`;
    }).join("") || `<tr><td colspan="100" class="empty">No open POs para este material</td></tr>`;
    content = `<table class="oc-table" style="width:100%">
      <thead><tr>
        ${multi?"<th class='col-centro'>Centro</th>":""}
        <th class="num">Cant. Pend.</th><th>Fecha Entrega</th><th>Orden de Compra</th><th>Origen / Proveedor</th><th>Estado</th>
      </tr></thead><tbody>${ocRows}</tbody></table>`;
  } else {
    const solRows = allSolList.map(s =>
      `<tr class="${s.sin_liberar?"row-sinlib":""}">
        ${multi?`<td class="col-centro"><span class="centro-badge">${s._cn}</span></td>`:""}
        <td class="num fw">${fmt(s.qty)} <span style="color:var(--tx3)">${s._um}</span></td>
        <td>${s.fecha}</td>
        <td class="mono">${s.doc}</td>
        <td><span class="lib-badge ${s.sin_liberar?"badge-sinlib":"badge-lib"}">${s.sin_liberar?"Unreleased":"Released"}</span></td>
      </tr>`
    ).join("") || `<tr><td colspan="100" class="empty">No PRs para este material</td></tr>`;
    content = `<table class="oc-table" style="width:100%">
      <thead><tr>
        ${multi?"<th class='col-centro'>Centro</th>":""}
        <th class="num">Cantidad</th><th>Fecha Entrega</th><th>N&deg; SOLPED</th><th>Estado Lib.</th>
      </tr></thead><tbody>${solRows}</tbody></table>`;
  }
  document.getElementById("sdp-content").innerHTML =
    `<div class="tbl-toolbar"><button class="copy-btn" onclick="copyTable(document.querySelector('#sdp-content table'),document.getElementById('sdp-mat-title').textContent)">⧉ Copiar detalle</button></div>` +
    content;
}

// ── Buffer Chart ─────────────────────────────────────────────────────────────
let _bufChart = null;
let _bufActiveCk = null;

function showBufferChart() {
  if (!selSemMat) return;
  const overlay = document.getElementById("buf-modal-overlay");
  overlay.classList.add("open");
  _bufActiveCk = selSemCk || [...centrosSel][0];
  _renderBufTabs();
  _drawBufChart(_bufActiveCk);
}

function closeBufModal() {
  document.getElementById("buf-modal-overlay").classList.remove("open");
  if (_bufChart) { _bufChart.destroy(); _bufChart = null; }
}

function _renderBufTabs() {
  const tabsEl = document.getElementById("buf-centro-tabs");
  const entries = [...centrosSel]
    .map(k => ({ k, cn: DATA.centros[k].nombre, m: (DATA.centros[k]?.materiales||[]).find(x=>x.mat===selSemMat) }))
    .filter(x => x.m);
  if (entries.length <= 1) { tabsEl.innerHTML = ""; return; }
  tabsEl.innerHTML = entries.map(e =>
    `<button class="buf-ctab ${e.k===_bufActiveCk?"active":""}" onclick="_bufSelectCk('${e.k}')">${e.cn}</button>`
  ).join("");
}

function _bufSelectCk(k) {
  _bufActiveCk = k;
  _renderBufTabs();
  _drawBufChart(k);
}

function _drawBufChart(ck) {
  if (typeof Chart === "undefined") { alert("Chart.js no cargó — verifica conexión a internet."); return; }
  const ct = DATA.centros[ck];
  if (!ct) return;
  const m = (ct.materiales||[]).find(x => x.mat === selSemMat);
  if (!m) return;

  document.getElementById("buf-modal-title").textContent =
    m.mat + " — " + truncate(m.desc, 50) + " · " + ct.nombre;

  const weeks  = DATA.meta.sem_weeks;
  const labels = ["Hoy", ...weeks.map(w => w.label)];
  const stocks = [m.saldo, ...m.sabados.map(s => Math.max(0, s.stock))];
  // Scale: show at least 1.5× TOR so green planning zone is visible; also cover all stock values
  const maxY   = Math.max(m.z_roja * 1.6, m.z_verde * 1.05, ...stocks) * 1.05;

  // Planning zones (planeación): TOR/2 = top of red, TOR = top of yellow
  // Both zones are equal size (TOR/2 each), unlike execution zones
  const zRojaPlan = m.z_roja / 2;  // top of planning red zone
  const zAmPlan   = m.z_roja;      // top of planning yellow zone

  // Zone fills using PLANNING thresholds
  const fill0        = labels.map(() => 0);
  const fillRojaPlan = labels.map(() => zRojaPlan);
  const fillAmPlan   = labels.map(() => zAmPlan);
  const fillTop      = labels.map(() => maxY);

  // Execution zone reference lines (dashed, for context)
  const fillRoja = labels.map(() => m.z_roja);
  const fillAm   = labels.map(() => m.z_am);
  const fillVerd = labels.map(() => m.z_verde);

  const hex2rgba = (h, a) => {
    const r=parseInt(h.slice(1,3),16), g=parseInt(h.slice(3,5),16), b=parseInt(h.slice(5,7),16);
    return `rgba(${r},${g},${b},${a})`;
  };

  const dotColors = stocks.map((_, i) => {
    if (i===0) return "#ffffff";
    const c = m.sabados[i-1]?.color;
    return c==="K"?"#c0392b":c==="R"?"#ff4d4d":c==="Y"?"#f5a623":c==="B"?"#4fa3e0":"#3ecf8e";
  });

  // ── Inflows OC por semana (barras de ingreso) ────────────────────────
  const weekDates = DATA.meta.sem_weeks.map(w => new Date(w.date + "T12:00:00"));
  const inflowSlots = labels.map(() => ({qty: 0, docs: []}));
  (m.oc_list || []).forEach(o => {
    const etaStr = o.eta_proy || o.eta;
    if (!etaStr) return;
    const eta = new Date(etaStr + "T12:00:00");
    let slot = -1;
    for (let i = 0; i < weekDates.length; i++) {
      if (eta <= weekDates[i]) { slot = i + 1; break; }
    }
    if (slot < 1) slot = 1;  // mínimo primera semana
    if (slot >= inflowSlots.length) slot = inflowSlots.length - 1;
    const qty = o.qty_base != null ? o.qty_base : o.qty;
    inflowSlots[slot].qty += qty;
    inflowSlots[slot].docs.push(o.doc);
  });
  const inflowData = inflowSlots.map(s => s.qty > 0 ? Math.round(s.qty) : null);
  const inflowDocs = inflowSlots.map(s => s.docs);

  const canvas = document.getElementById("buf-chart-canvas");
  if (_bufChart) { _bufChart.destroy(); _bufChart = null; }

  _bufChart = new Chart(canvas, {
    type: "line",
    data: {
      labels,
      datasets: [
        // Zone backgrounds — PLANEACIÓN (red=TOR/2, yellow=TOR/2, equal zones)
        { label:"_r",  data: fillRojaPlan, fill:"origin", backgroundColor:"rgba(255,77,77,0.28)",  borderWidth:0, pointRadius:0, tension:0, order:10 },
        { label:"_y",  data: fillAmPlan,   fill:"-1",     backgroundColor:"rgba(245,166,35,0.25)", borderWidth:0, pointRadius:0, tension:0, order:10 },
        { label:"_g",  data: fillTop,      fill:"-1",     backgroundColor:"rgba(62,207,142,0.13)", borderWidth:0, pointRadius:0, tension:0, order:10 },
        // Planning zone boundary lines
        { label:"TOR/2", data: fillRojaPlan, fill:false, borderColor:"rgba(255,77,77,0.8)",   borderWidth:1.5, borderDash:[4,3], pointRadius:0, tension:0, order:5 },
        { label:"TOR",   data: fillAmPlan,   fill:false, borderColor:"rgba(245,166,35,0.8)",  borderWidth:1.5, borderDash:[4,3], pointRadius:0, tension:0, order:5 },
        // Execution TOR as thin reference line (shows buffer execution threshold)
        { label:"TOR ejec.", data: fillRoja, fill:false, borderColor:"rgba(255,77,77,0.30)", borderWidth:1, borderDash:[2,5], pointRadius:0, tension:0, order:6 },
        // Ingresos OC por semana (barras verde oscuro)
        {
          type: "bar",
          label: "Ingresos OC",
          data: inflowData,
          backgroundColor: "rgba(15, 90, 40, 0.82)",
          borderColor: "rgba(30, 160, 75, 0.90)",
          borderWidth: 1,
          borderRadius: 3,
          barPercentage: 0.45,
          categoryPercentage: 0.8,
          order: 3
        },
        // Stock projection line
        {
          label: "Stock proyectado",
          data: stocks,
          fill: false,
          borderColor: "#e2e8f0",
          borderWidth: 2.5,
          tension: 0.3,
          pointRadius: stocks.map((_,i) => i===0?6:4),
          pointBackgroundColor: dotColors,
          pointBorderColor: "#1a1f2e",
          pointBorderWidth: 1,
          order: 1
        }
      ]
    },
    options: {
      responsive: true,
      maintainAspectRatio: false,
      animation: { duration: 300 },
      scales: {
        x: {
          grid: { color: "rgba(255,255,255,0.05)" },
          ticks: { color: "#8898aa", font: { size: 10 }, maxRotation: 45 }
        },
        y: {
          min: 0, max: maxY,
          grid: { color: "rgba(255,255,255,0.05)" },
          ticks: {
            color: "#8898aa", font: { size: 10 },
            callback: v => v>=1000 ? (v/1000).toFixed(1)+"k" : v.toFixed(0)
          },
          title: { display: true, text: m.um, color: "#8898aa", font: { size: 11 } }
        }
      },
      plugins: {
        legend: { display: false },
        tooltip: {
          callbacks: {
            label: ctx => {
              if (!ctx.dataset.label || ctx.dataset.label.startsWith("_")) return null;
              const fmt = v => v>=1000?(v/1000).toFixed(1)+"k":v.toFixed(0);
              if (ctx.dataset.label === "Stock proyectado") {
                const v = ctx.parsed.y;
                return ` Stock: ${v>=1000?(v/1000).toFixed(2)+"k":v.toFixed(0)} ${m.um}`;
              }
              if (ctx.dataset.label === "Ingresos OC") {
                const v = ctx.parsed.y;
                const docs = inflowDocs[ctx.dataIndex] || [];
                const lines = [` Ingreso: ${fmt(v)} ${m.um}`];
                if (docs.length) lines.push(` PO${docs.length>1?"s":""}: ${docs.join(", ")}`);
                return lines;
              }
              if (ctx.dataset.label === "TOR/2")     return ` TOR/2 (plan): ${fmt(ctx.parsed.y)} ${m.um}`;
              if (ctx.dataset.label === "TOR")       return ` TOR (plan): ${fmt(ctx.parsed.y)} ${m.um}`;
              if (ctx.dataset.label === "TOR ejec.") return ` TOR (ejec.): ${fmt(ctx.parsed.y)} ${m.um}`;
              return null;
            },
            title: ctx => ctx[0]?.label
          },
          backgroundColor: "#1e2536",
          titleColor: "#e2e8f0",
          bodyColor: "#a0aec0",
          borderColor: "#2d3748",
          borderWidth: 1
        }
      }
    }
  });
}

// ── Resize columna Descripción ───────────────────────────────────────────────
function startDescResize(e) {
  e.preventDefault(); e.stopPropagation();
  const handle = document.getElementById("desc-resize");
  if (!handle) return;
  handle.classList.add("dragging");
  const startX   = e.clientX;
  const startW   = parseInt(getComputedStyle(document.documentElement).getPropertyValue("--desc-w")) || 220;
  function onMove(ev) {
    const newW = Math.max(80, startW + ev.clientX - startX);
    document.documentElement.style.setProperty("--desc-w", newW + "px");
  }
  function onUp() {
    handle.classList.remove("dragging");
    document.removeEventListener("mousemove", onMove);
    document.removeEventListener("mouseup", onUp);
  }
  document.addEventListener("mousemove", onMove);
  document.addEventListener("mouseup", onUp);
}

// ── Copy to clipboard ─────────────────────────────────────────────────────────
function copyTable(tableEl, label) {
  if (!tableEl) return;

  // ── Estilos inline para HTML (compatibles con Outlook / Excel) ──────────────
  const TH   = "padding:6px 10px;background:#1e3a5f;color:#ffffff;font-family:Arial,sans-serif;font-size:12px;font-weight:bold;border:1px solid #b0b8cc;white-space:nowrap";
  const TH_R = TH + ";text-align:right";
  const TD   = "padding:5px 10px;background:#ffffff;color:#111111;font-family:Arial,sans-serif;font-size:12px;border:1px solid #d8dde8";
  const TD_R = TD + ";text-align:right";
  const TF   = "padding:5px 10px;background:#eef2f8;color:#111111;font-family:Arial,sans-serif;font-size:12px;border:1px solid #d8dde8;font-weight:bold";
  const TF_R = TF + ";text-align:right";
  const BG_ATRASO = "#fff0f0";
  const BG_SINLIB = "#fffbee";

  // ── Generar TSV (texto plano) ────────────────────────────────────────────────
  const tsvRows = [];
  tableEl.querySelectorAll("thead tr").forEach(tr =>
    tsvRows.push([...tr.querySelectorAll("th")].map(th => th.textContent.trim()).join("\t"))
  );
  tableEl.querySelectorAll("tbody tr").forEach(tr => {
    const cells = [...tr.querySelectorAll("td")].map(td => td.textContent.trim().replace(/\s+/g," "));
    if (cells.some(c => c && c !== "—")) tsvRows.push(cells.join("\t"));
  });
  tableEl.querySelectorAll("tfoot tr").forEach(tr => {
    tsvRows.push([...tr.querySelectorAll("td")].map(td => td.textContent.trim()).join("\t"));
  });
  const tsv = tsvRows.join("\n");

  // ── Generar HTML con formato ─────────────────────────────────────────────────
  let html = "<table style='border-collapse:collapse;font-family:Arial,sans-serif;font-size:12px'>";

  tableEl.querySelectorAll("thead tr").forEach(tr => {
    html += "<tr>";
    tr.querySelectorAll("th").forEach(th =>
      html += `<th style="${th.classList.contains("num") ? TH_R : TH}">${th.textContent.trim()}</th>`
    );
    html += "</tr>";
  });

  tableEl.querySelectorAll("tbody tr").forEach(tr => {
    const tds = tr.querySelectorAll("td");
    if (!tds.length) return;
    const txt0 = tds[0].textContent.trim();
    if (!txt0 || txt0.startsWith("Sin ")) return;
    const bg = tr.classList.contains("row-atraso") ? BG_ATRASO
              : tr.classList.contains("row-sinlib")  ? BG_SINLIB : "#ffffff";
    html += "<tr>";
    tds.forEach(td => {
      const isR = td.classList.contains("num") || td.classList.contains("tot-v");
      const st  = (isR ? TD_R : TD).replace("background:#ffffff", "background:" + bg);
      html += `<td style="${st}">${td.textContent.trim().replace(/\s+/g," ")}</td>`;
    });
    html += "</tr>";
  });

  tableEl.querySelectorAll("tfoot tr").forEach(tr => {
    html += "<tr>";
    tr.querySelectorAll("td").forEach(td => {
      const isR = td.classList.contains("num") || td.classList.contains("tot-v");
      const cp  = td.getAttribute("colspan");
      html += `<td style="${isR ? TF_R : TF}"${cp ? ` colspan="${cp}"` : ""}>${td.textContent.trim()}</td>`;
    });
    html += "</tr>";
  });

  html += "</table>";

  // ── Copiar al portapapeles (HTML + texto plano simultáneo) ──────────────────
  const lbl = label || "Tabla";
  if (window.ClipboardItem) {
    navigator.clipboard.write([new ClipboardItem({
      "text/html":  new Blob([html], { type: "text/html"  }),
      "text/plain": new Blob([tsv],  { type: "text/plain" })
    })]).then(() => showToast("⧉ " + lbl + " copiada"))
       .catch(() => navigator.clipboard.writeText(tsv).then(() => showToast("⧉ " + lbl + " copiada")));
  } else {
    navigator.clipboard.writeText(tsv).then(() => showToast("⧉ " + lbl + " copiada"));
  }
}
function showToast(msg) {
  let t = document.getElementById("copy-toast");
  if (!t) { t = document.createElement("div"); t.id = "copy-toast"; document.body.appendChild(t); }
  t.textContent = msg; t.className = "copy-toast show";
  setTimeout(() => { t.className = "copy-toast"; }, 2000);
}

// ── Utilities ─────────────────────────────────────────────────────────────────
function fmt(n) {
  if (n==null||n===undefined||n==="") return "—";
  const v = Number(n);
  if (isNaN(v)) return "—";
  return Math.round(v).toLocaleString("es-CO");
}
function fmt1(n) {
  const v = Number(n);
  if (isNaN(v)||!v) return "—";
  return v.toLocaleString("es-CO",{maximumFractionDigits:2,minimumFractionDigits:0});
}
function truncate(s, n) {
  return s && s.length>n ? s.substring(0,n)+"…" : (s||"");
}
function zBadge(z) {
  if (!z) return {cls:"zona-def",lbl:"—"};
  const zu = z.toUpperCase();
  if (zu.includes("DR")||zu.startsWith("10")) return {cls:"zona-r",lbl:"DR"};
  if (zu.startsWith("20")||zu.includes(" R")) return {cls:"zona-r",lbl:"Rojo"};
  if (zu.startsWith("30")||zu.includes(" Y")) return {cls:"zona-y",lbl:"Amarillo"};
  if (zu.startsWith("40")||zu.includes(" G")) return {cls:"zona-g",lbl:"Verde"};
  return {cls:"zona-def",lbl:z.substring(0,8)};
}

// ── Modal selección de centros para PDF ──────────────────────────────────────
let _rptPais = "ALL";
let _rptCentroSel = new Set();
let _rptPanelOpen = false;

function openReportModal() {
  _rptPais = "ALL";
  _rptCentroSel = new Set(Object.keys(DATA.centros));
  document.getElementById("report-modal").style.display = "flex";
  renderRptCountryBar();
  updateRptCentroLabel();
}
function closeReportModal() {
  _rptPanelOpen = false;
  const p = document.getElementById("rpt-centro-panel");
  if (p) p.classList.remove("open");
  document.getElementById("report-modal").style.display = "none";
}

// ── Report modal: country bar ─────────────────────────────────────────────────
const _RPT_FLAG = {AR:"🇦🇷",BR:"🇧🇷",CO:"🇨🇴",PE:"🇵🇪",CL:"🇨🇱"};

function renderRptCountryBar() {
  const bar = document.getElementById("rpt-country-bar");
  if (!bar) return;
  const paises = DATA.meta.paises || {};
  const pList = Object.keys(paises).filter(p => (paises[p].centros||[]).some(k => DATA.centros[k]));
  let html = `<button class="rpt-cpill${_rptPais==="ALL"?" active":""}" onclick="setRptPais('ALL')">&#127758; ALL</button>`;
  pList.forEach(p => {
    const f = paises[p].flag || "";
    const emo = _RPT_FLAG[f] || f;
    html += `<button class="rpt-cpill${_rptPais===p?" active":""}" onclick="setRptPais('${p}')">${emo} ${p}</button>`;
  });
  bar.innerHTML = html;
}

function setRptPais(pais) {
  _rptPais = pais;
  if (pais === "ALL") {
    _rptCentroSel = new Set(Object.keys(DATA.centros));
  } else {
    const pm = (DATA.meta.paises||{})[pais];
    _rptCentroSel = new Set((pm?.centros||[]).filter(k => DATA.centros[k]));
  }
  renderRptCountryBar();
  updateRptCentroLabel();
  if (_rptPanelOpen) renderRptCentroPanel();
}

// ── Report modal: centro dropdown ─────────────────────────────────────────────
function _rptVisibleCentros() {
  if (_rptPais === "ALL") return Object.keys(DATA.centros);
  const pm = (DATA.meta.paises||{})[_rptPais];
  return (pm?.centros||[]).filter(k => DATA.centros[k]);
}

function updateRptCentroLabel() {
  const lbl = document.getElementById("rpt-centro-label");
  if (!lbl) return;
  const visible = _rptVisibleCentros();
  const n = [..._rptCentroSel].filter(k => visible.includes(k)).length;
  if (n === 0) { lbl.textContent = "Ningún centro"; return; }
  if (n >= visible.length) { lbl.textContent = "All plants"; return; }
  const names = [..._rptCentroSel].filter(k => visible.includes(k)).map(k => DATA.centros[k]?.nombre || k);
  lbl.textContent = names.slice(0,2).join(", ") + (names.length > 2 ? ` (+${names.length-2})` : "");
}

function toggleRptCentroPanel(e) {
  e && e.stopPropagation();
  const panel = document.getElementById("rpt-centro-panel");
  const btn   = document.getElementById("rpt-centro-btn");
  _rptPanelOpen = !_rptPanelOpen;
  if (_rptPanelOpen) {
    const rect = btn.getBoundingClientRect();
    panel.style.top   = (rect.bottom + 4) + "px";
    panel.style.left  = rect.left + "px";
    panel.style.width = rect.width + "px";
    panel.classList.add("open");
    renderRptCentroPanel();
    setTimeout(() => document.getElementById("rpt-centro-search")?.focus(), 50);
  } else {
    panel.classList.remove("open");
  }
}

function renderRptCentroPanel() {
  const scroll = document.getElementById("rpt-centro-scroll");
  if (!scroll) return;
  const q = (document.getElementById("rpt-centro-search")?.value||"").trim().toLowerCase();
  const paises = DATA.meta.paises || {};
  const paisList = _rptPais === "ALL"
    ? Object.keys(paises).filter(p => (paises[p].centros||[]).some(k => DATA.centros[k]))
    : [_rptPais].filter(p => paises[p]);
  const visible = _rptVisibleCentros();
  const allSel  = visible.length > 0 && visible.every(k => _rptCentroSel.has(k));

  let html = `<div class="rpt-msel-item rpt-msel-all-row">
    <input type="checkbox" id="rpt-all-cb" ${allSel?"checked":""} onchange="rptToggleAll(this)">
    <span>All plants</span>
    <span class="rpt-mat-cnt">${visible.length} centros</span>
  </div>`;

  for (const pais of paisList) {
    const pm   = paises[pais] || {};
    const keys = (pm.centros||[]).filter(k => DATA.centros[k]);
    const filt = q ? keys.filter(k => DATA.centros[k].nombre.toLowerCase().includes(q)) : keys;
    if (!filt.length) continue;
    if (_rptPais === "ALL") {
      const emo = _RPT_FLAG[pm.flag||""] || pm.flag || "";
      html += `<div class="rpt-msel-grp-hdr">${emo} ${pais}</div>`;
    }
    for (const k of filt) {
      const cd    = DATA.centros[k];
      const ncrit = cd.materiales.filter(m=>m.estado==="CRITICO").length;
      const nrisk = cd.materiales.filter(m=>m.estado==="RIESGO").length;
      html += `<div class="rpt-msel-item">
        <input type="checkbox" ${_rptCentroSel.has(k)?"checked":""} onchange="rptToggleCentro('${k}',this)">
        <span>${cd.nombre}</span>
        ${ncrit>0?`<span class="rpt-badge-r">&#128308; ${ncrit}</span>`:""}
        ${nrisk>0?`<span class="rpt-badge-y">&#128993; ${nrisk}</span>`:""}
        <span class="rpt-mat-cnt">(${cd.materiales.length})</span>
      </div>`;
    }
  }
  scroll.innerHTML = html;
}

function rptToggleAll(el) {
  const keys = _rptVisibleCentros();
  if (el.checked) keys.forEach(k => _rptCentroSel.add(k));
  else            keys.forEach(k => _rptCentroSel.delete(k));
  updateRptCentroLabel();
  renderRptCentroPanel();
}

function rptToggleCentro(k, el) {
  if (el.checked) _rptCentroSel.add(k);
  else            _rptCentroSel.delete(k);
  const visible = _rptVisibleCentros();
  const allCb = document.getElementById("rpt-all-cb");
  if (allCb) allCb.checked = visible.every(k => _rptCentroSel.has(k));
  updateRptCentroLabel();
}

document.addEventListener("click", e => {
  if (_rptPanelOpen
      && !document.getElementById("rpt-centro-panel")?.contains(e.target)
      && !document.getElementById("rpt-centro-btn")?.contains(e.target)) {
    _rptPanelOpen = false;
    document.getElementById("rpt-centro-panel")?.classList.remove("open");
  }
});
// ── Menú desplegable Download Report ────────────────────────────────────────
function toggleRptMenu(e) {
  e.stopPropagation();
  const dd = document.getElementById("rpt-dropdown");
  dd.style.display = dd.style.display === "none" ? "block" : "none";
}
function closeRptMenu() {
  document.getElementById("rpt-dropdown").style.display = "none";
}
document.addEventListener("click", function(e) {
  if (!document.getElementById("rpt-menu-wrap").contains(e.target)) closeRptMenu();
});
// ── Impact Report (Excel por países) ──────────────────────────────
function exportImpactosExcel() {
  const b64 = window._IMPACTOS_XLSX_B64;
  if (!b64) { alert("File not available. Regenerate dashboard."); return; }
  const bin = atob(b64);
  const arr = new Uint8Array(bin.length);
  for (let i = 0; i < bin.length; i++) arr[i] = bin.charCodeAt(i);
  const blob = new Blob([arr], {type: "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  const fecha = DATA.meta.fecha_saldo.replace(/[^a-zA-Z0-9]/g,"-");
  a.download = "Reporte_Impactos_Planners_" + fecha + ".xlsx";
  a.click();
}
function generateEmailFromModal() {
  const {selected, estados, origen} = _rptReadFilters();
  closeReportModal();
  if (!selected.length) { alert("Selecciona al menos un centro."); return; }
  try {
  const meta = DATA.meta;
  const fecha = meta.fecha_saldo;
  const C = {
    bg:"#f5f7fa",bg1:"#ffffff",bg2:"#eef1f6",brd:"#d1d8e2",
    tx1:"#1a1d24",tx2:"#5a6478",
    red:"#dc2626",red2:"#fef2f2",
    yel:"#c97b0a",yel2:"#fffbeb",
    grn:"#0a7a52",grn2:"#f0fdf9",
    ora:"#e05a1e",
  };
  const inclSet = estados ? new Set(Object.entries(estados).filter(([,v])=>v).map(([k])=>k)) : null;
  const origenFilter = m => origen==="ALL" ? true : origen==="IMP" ? m.importado===true : m.importado===false;
  function getNextSupply(m){
    let best=null,bQty=0,bUm=m.um||"";
    (m.oc_list||[]).forEach(o=>{
      const eta=o.eta_proy||o.eta; if(!eta) return;
      const d=new Date(eta);
      if(!best||d<best){best=d;bQty=o.qty_base||o.qty||0;bUm=m.um||"";}
    });
    (m.sol_list||[]).forEach(s=>{
      if(s.sin_liberar||!s.fecha) return;
      const d=new Date(s.fecha);
      if(!best||d<best){best=d;bQty=s.qty||0;bUm=s.um||m.um||"";}
    });
    return {date:best,qty:bQty,um:bUm};
  }
  function isParadaReal(m,fechaSaldo){
    const qd=m.primer_quiebre_dias; if(qd==null) return false;
    const stockout=new Date(fechaSaldo); stockout.setDate(stockout.getDate()+Math.round(qd));
    const ns=getNextSupply(m).date;
    if(!ns) return true;
    return stockout<=ns;
  }
  function fmtSupplyDate(d){ if(!d) return "N/P"; return d.toLocaleDateString("es-CO",{day:"2-digit",month:"short",year:"numeric"}); }
  function chip(label,count,color,bg,minWidth){
    return `<td style="padding:6px 10px;text-align:center;background:${bg};border-radius:6px;border:1px solid ${color};min-width:${minWidth||48}px">
      <div style="font-size:18px;font-weight:700;color:${color}">${count}</div>
      <div style="font-size:9px;color:${color};opacity:.8;letter-spacing:.5px">${label}</div></td>`;
  }
  function tbl(headers, rows, colWidths){
    const hRow=headers.map((h,i)=>`<th style="padding:5px 8px;text-align:left;background:#1a1d24;color:#ffffff;font-size:10px;font-weight:700;letter-spacing:.5px;white-space:nowrap;${colWidths&&colWidths[i]?'width:'+colWidths[i]:''}">${h}</th>`).join("");
    return `<table style="width:100%;border-collapse:collapse;font-size:12px;margin-top:8px">
      <thead><tr>${hRow}</tr></thead><tbody>${rows}</tbody></table>`;
  }
  const {ref, d7, d21} = getOCDateRef();
  const centroOrder = Object.keys(DATA.centros).map(k => ({
    key: k,
    nombre: DATA.centros[k]?.nombre || k,
    flag: DATA.centros[k]?.flag || "",
  }));
  const entries = centroOrder.filter(c=>selected.includes(c.key));
  const centroNames = entries.map(c=>c.nombre).join(", ");
  let totalCrit=0, totalRiesgo=0, totalParadas=0;
  const centroBlocks = entries.map(({key,nombre,flag}) => {
    const ct = DATA.centros[key]; if(!ct) return "";
    const allMats = ct.materiales||[];
    const mats = (inclSet ? allMats.filter(m=>inclSet.has(m.estado)) : allMats).filter(origenFilter);
    const oc = ct.oc||[];
    const sol = ct.sol||[];
    const matMap = {}; allMats.forEach(m=>matMap[m.mat]=m);

    // — Estado counts (full portfolio) —
    const cr=allMats.filter(m=>m.estado==="CRITICO").length;
    const ri=allMats.filter(m=>m.estado==="RIESGO").length;
    const al=allMats.filter(m=>m.estado==="ALERTA").length;
    const ok=allMats.filter(m=>m.estado==="OK").length;
    const sc=allMats.filter(m=>m.estado==="SIN_CONSUMO").length;
    totalCrit+=cr; totalRiesgo+=ri;

    // — OC analysis —
    const ocAtr = oc.filter(o=>o.atrasada);
    const ocVcr = oc.filter(o=>{ if(!o.fecha_entrega||o.atrasada) return false; const fe=new Date(o.fecha_entrega); return fe>ref&&fe<=d21; });
    const ocImpAtr = ocAtr.filter(o=>matMap[o.mat]?.importado===true);
    const ocNacAtr = ocAtr.filter(o=>matMap[o.mat]?.importado!==true);
    const ocImpVcr = ocVcr.filter(o=>matMap[o.mat]?.importado===true);
    const ocNacVcr = ocVcr.filter(o=>matMap[o.mat]?.importado!==true);
    const solSL = sol.filter(s=>s.sin_liberar);
    const fechaSaldoDate = new Date(DATA.meta.fecha_saldo);
    function resolveParada(m,fechaSaldo){
      if(m.impact){const t=m.impact.trim();return /^parada/i.test(t);}
      return isParadaReal(m,fechaSaldo);
    }
    const paradas = allMats.filter(m=>resolveParada(m,fechaSaldoDate));
    totalParadas+=paradas.length;

    // — Critical/Risk materials table —
    const critRisk = mats.filter(m=>m.estado==="CRITICO"||m.estado==="RIESGO")
      .sort((a,b)=>(a.estado==="CRITICO"?0:1)-(b.estado==="CRITICO"?0:1)||(a.cob_hoy-b.cob_hoy));
    const altRows = mats.filter(m=>m.estado==="ALERTA").length;
    const altExtra = altRows>0?`<tr><td colspan="6" style="padding:4px 8px;font-size:11px;color:${C.tx2};font-style:italic">+ ${altRows} material(es) en ALERTA (ver PDF adjunto)</td></tr>`:"";

    const critRiskRows = critRisk.slice(0,25).map(m=>{
      const estColor=m.estado==="CRITICO"?C.red:C.yel;
      const cobStr=m.adu_plan>0?(m.cob_hoy>998?"&#8734;":m.cob_hoy+"d"):"&mdash;";
      const orig=m.importado?"<b style='color:#4fa3e0'>IMP</b>":"<b style='color:#2ecc71'>NAC</b>";
      const parada=resolveParada(m,fechaSaldoDate);
      const paradaHtml=parada
        ?`<span style="color:${C.red};font-weight:700;font-size:10px">&#128308; PARADA</span>`
        :`<span style="color:${C.grn};font-size:10px">&#128994; NO PARADA</span>`;
      const {date:nsDate,qty:nsQty,um:nsUm}=getNextSupply(m);
      const nsStr=fmtSupplyDate(nsDate);
      const nsQtyStr=nsQty?Math.round(nsQty).toLocaleString("es-CO")+" "+(nsUm||""):"&mdash;";
      const accion=m.action||(m.sin_tiempo&&!nsDate?"EMITIR OC URGENTE":(m.sol_list||[]).some(s=>s.sin_liberar)?"LIBERAR SOLPED":(m.oc_list||[]).some(o=>o.atrasada)?"GESTIONAR OC":"MONITOREAR");
      return `<tr style="border-bottom:1px solid ${C.brd}">
        <td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${m.mat}</td>
        <td style="padding:5px 8px;color:${C.tx2};font-size:11px">${(m.desc||"").substring(0,32)}</td>
        <td style="padding:5px 8px;text-align:center">${orig}</td>
        <td style="padding:5px 8px;text-align:center;font-weight:700;color:${estColor};font-size:12px">${cobStr}</td>
        <td style="padding:5px 8px;text-align:center">${paradaHtml}</td>
        <td style="padding:5px 8px;text-align:center;font-size:11px;color:${C.tx1}">${nsStr}</td>
        <td style="padding:5px 8px;text-align:right;font-size:11px;color:${C.tx2}">${nsQtyStr}</td>
        <td style="padding:5px 8px;font-size:10px;color:${C.ora};font-weight:600">${accion}</td>
      </tr>`;
    }).join("")+altExtra;

    const critRiskTbl = critRisk.length ? tbl(
      ["C&#243;digo","Descripci&#243;n","Origen","Cob.","Impacto","Pr&#243;x. suministro","Cantidad","Acci&#243;n"],
      critRiskRows, ["85px","","55px","50px","90px","95px","80px",""]
    ) : `<div style="font-size:12px;color:${C.grn};padding:8px 0">&#10003; No critical or at-risk materials with current filters</div>`;

    // — OC section —
    const ocImpRow=(label,list,color)=>list.length?`<tr style="border-bottom:1px solid ${C.brd}">
      <td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${list[0]?.doc||""}</td>
      <td style="padding:5px 8px;color:${C.tx2};font-size:11px">${(list[0]?.desc||"").substring(0,30)}</td>
      <td style="padding:5px 8px;color:${C.tx2};font-size:11px">${list[0]?.proveedor||""}</td>
      <td style="padding:5px 8px;text-align:center;color:${color};font-weight:700;font-size:11px">${list[0]?.fecha_entrega||""}</td>
      <td style="padding:5px 8px;text-align:right;color:${C.tx1};font-size:11px">${Math.round(list[0]?.qty_pend||0).toLocaleString("es-CO")} ${list[0]?.um||""}</td>
    </tr>`:"";
    const buildOcSection=(label,atrList,vcrList,color)=>{
      if(!atrList.length&&!vcrList.length) return "";
      const allRows=[...atrList.slice(0,5),...vcrList.slice(0,3)];
      const rows=allRows.map(o=>`<tr style="border-bottom:1px solid ${C.brd}">
        <td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${o.doc||""}</td>
        <td style="padding:5px 8px;color:${C.tx2};font-size:11px;max-width:160px">${(o.desc||"").substring(0,32)}</td>
        <td style="padding:5px 8px;color:${C.tx2};font-size:10px">${(o.proveedor||"").substring(0,25)}</td>
        <td style="padding:5px 8px;text-align:center;color:${o.atrasada?C.red:C.yel};font-weight:700;font-size:11px">${o.fecha_entrega||""}</td>
        <td style="padding:5px 8px;text-align:center;font-size:10px">${o.atrasada?`<span style="color:${C.red}">ATRASADA</span>`:`<span style="color:${C.yel}">&le;21d</span>`}</td>
        <td style="padding:5px 8px;text-align:right;color:${C.tx1};font-size:11px">${Math.round(o.qty_pend||0).toLocaleString("es-CO")} ${o.um||""}</td>
      </tr>`).join("");
      const extra=(atrList.length+vcrList.length>8)?`<tr><td colspan="6" style="padding:4px 8px;font-size:10px;color:${C.tx2};font-style:italic">+ ${atrList.length+vcrList.length-8} OC adicionales &mdash; ver Excel adjunto</td></tr>`:"";
      return `<div style="margin-bottom:12px">
        <div style="font-size:11px;font-weight:700;color:${color};padding:6px 0;letter-spacing:.5px">${label} &mdash; ${atrList.length} atrasada(s) &middot; ${vcrList.length} vence &le;21d</div>
        ${tbl(["N&#176; OC","Descripci&#243;n","Proveedor","Fecha Entrega","Estado","Cant. Pend."],rows+extra,["80px","","","85px","65px","80px"])}
      </div>`;
    };
    const ocImpSec = buildOcSection("&#128666; IMPORTACIONES",ocImpAtr,ocImpVcr,C.red);
    const ocNacSec = buildOcSection("&#127981; NACIONALES",ocNacAtr,ocNacVcr,C.red);

    // — SOLPEDs sin liberar —
    const solRows = solSL.slice(0,10).map(s=>`<tr style="border-bottom:1px solid ${C.brd}">
      <td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${s.doc}</td>
      <td style="padding:5px 8px;color:${C.tx2};font-size:11px;max-width:160px">${(s.desc||"").substring(0,35)}</td>
      <td style="padding:5px 8px;text-align:center;font-size:11px;color:${C.yel}">${s.fecha_entrega||"&mdash;"}</td>
      <td style="padding:5px 8px;text-align:right;color:${C.tx1};font-size:11px">${Math.round(s.qty||0).toLocaleString("es-CO")} ${s.um||""}</td>
    </tr>`).join("");
    const solExtra=solSL.length>10?`<tr><td colspan="4" style="padding:4px 8px;font-size:10px;color:${C.tx2};font-style:italic">+ ${solSL.length-10} SOLPEDs adicionales</td></tr>`:"";
    const solSec = solSL.length?`
      <div style="margin-bottom:12px">
        <div style="font-size:11px;font-weight:700;color:${C.yel};padding:6px 0;letter-spacing:.5px">&#128196; SOLPEDS SIN LIBERAR (${solSL.length})</div>
        ${tbl(["N&#176; SOLPED","Descripci&#243;n","Fecha Entrega","Cantidad"],solRows+solExtra,["80px","","80px","80px"])}
      </div>`:"";

    // — Paradas de línea —
    const paradaRows=paradas.slice(0,10).map(p=>`<tr style="border-bottom:1px solid #2a0a0a">
      <td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${p.mat}</td>
      <td style="padding:5px 8px;color:${C.tx2};font-size:11px">${(p.desc||"").substring(0,35)}</td>
      <td style="padding:5px 8px;color:${C.red};font-size:11px;font-weight:700">${p.impact}</td>
      <td style="padding:5px 8px;color:${C.tx2};font-size:11px">${(p.action||"").substring(0,50)}</td>
    </tr>`).join("");
    const paradaSec=paradas.length?`
      <div style="margin-bottom:12px;border:1px solid ${C.red};border-radius:6px;overflow:hidden">
        <div style="background:${C.red2};padding:8px 12px;font-size:11px;font-weight:700;color:${C.red}">&#9888;&#65039; IMPACTO PARADA DE L&#205;NEA (${paradas.length} material(es))</div>
        <div style="padding:0 8px 4px">
          ${tbl(["C&#243;digo","Descripci&#243;n","Impacto","Acci&#243;n del Planeador"],paradaRows,["90px","","120px",""])}
        </div>
      </div>`:"";

    const hdrColor=cr>0?C.red:ri>0?C.yel:C.grn;
    return `
    <div style="background:${C.bg1};border:1px solid ${hdrColor};border-radius:8px;margin-bottom:18px;overflow:hidden">
      <div style="background:#e05a1e;padding:10px 16px;border-bottom:1px solid ${C.brd}">
        <span style="font-size:16px;font-weight:700;color:#ffffff">${flag} ${nombre}</span>
        <span style="font-size:11px;color:rgba(255,255,255,.8);float:right">${allMats.length} materiales en portafolio</span>
      </div>
      <div style="padding:14px 16px">
        <div style="background:${C.bg2};border:1px solid ${C.brd};border-radius:6px;padding:8px 12px;margin-bottom:8px">
          <div style="font-size:9px;font-weight:700;color:${C.tx2};letter-spacing:1px;margin-bottom:6px">ESTADO DE INVENTARIO &mdash; ${allMats.length} materiales en portafolio</div>
          <table style="border-collapse:separate;border-spacing:4px"><tr>
            ${chip("CR&#205;TICO",cr,C.red,C.red2)}
            ${chip("RIESGO",ri,C.yel,C.yel2)}
            ${chip("ALERTA",al,C.ora,"#fff3e0")}
            ${chip("OK",ok,C.grn,C.grn2)}
            ${chip("INV.SC",sc,C.tx2,C.bg2)}
          </tr></table>
        </div>
        <div style="background:${C.bg2};border-left:3px solid ${C.ora};border-radius:6px;padding:8px 12px;margin-bottom:14px">
          <div style="font-size:9px;font-weight:700;color:${C.yel};letter-spacing:1px;margin-bottom:4px">POSICIONES DE ABASTECIMIENTO</div>
          <div style="font-size:12px;color:${C.tx1}">
            <b style="color:${ocAtr.length>0?C.red:C.tx1}">${ocAtr.length}</b> posiciones OC atrasadas
            &nbsp;&middot;&nbsp;
            <b style="color:${solSL.length>0?C.yel:C.tx1}">${solSL.length}</b> posiciones SOLPED sin liberar
          </div>
        </div>
        <div style="font-size:10px;font-weight:700;color:${C.tx2};letter-spacing:.8px;margin-bottom:4px">MATERIALES CR&#205;TICOS Y EN RIESGO</div>
        ${critRiskTbl}
        ${sol.length?`<div style="font-size:10px;font-weight:700;color:${C.tx2};letter-spacing:.8px;margin:12px 0 4px">SOLPEDs &mdash; ${sol.length} posiciones</div>${tbl(["N&#176; SOLPED","Material","Descripci&#243;n","Cantidad","Fecha Esp.","Estado"],sol.sort((a,b)=>(b.sin_liberar?1:0)-(a.sin_liberar?1:0)||(a.fecha_entrega||"").localeCompare(b.fecha_entrega||"")).map(s=>`<tr style="border-bottom:1px solid ${C.brd}"><td style="padding:5px 8px;font-family:monospace;color:${C.tx1};font-size:11px">${s.doc}</td><td style="padding:5px 8px;font-family:monospace;color:${C.tx2};font-size:11px">${s.mat||""}</td><td style="padding:5px 8px;color:${C.tx2};font-size:11px">${(s.desc||"").substring(0,32)}</td><td style="padding:5px 8px;text-align:right;color:${C.tx1};font-size:11px">${Math.round(s.qty||0).toLocaleString("es-CO")} ${s.um||""}</td><td style="padding:5px 8px;text-align:center;color:${C.tx2};font-size:11px">${s.fecha_entrega||"&mdash;"}</td><td style="padding:5px 8px;text-align:center">${s.sin_liberar?`<span style="color:${C.red};font-size:9px;font-weight:700">SIN LIBERAR</span>`:`<span style="color:${C.grn};font-size:9px">LIBERADA</span>`}</td></tr>`).join(""),["80px","70px","","90px","80px","80px"])}`:""}
        <div style="font-size:11px;color:${C.tx2};margin-top:10px;padding-top:8px;border-top:1px solid ${C.brd}">&#128206; Detalle de &Oacute;rdenes de Compra disponible en el PDF adjunto.</div>
      </div>
    </div>`;
  }).join("");

  const globalColor=totalCrit>0?C.red:totalRiesgo>0?C.yel:C.grn;
  const globalIcon=totalCrit>0?"&#128308;":totalRiesgo>0?"&#128993;":"&#128994;";
  const globalText=totalCrit>0?`${totalCrit} materiales CR&#205;TICOS`:totalRiesgo>0?`Sin cr&#237;ticos &middot; ${totalRiesgo} at risk`:"Sin cr&#237;ticos ni riesgos activos";
  const paradaGlobal=totalParadas>0?`<div style="background:${C.red2};border:1px solid ${C.red};border-radius:6px;padding:10px 16px;margin-bottom:16px;font-size:12px;color:${C.red}">
    &#9888;&#65039; <b>${totalParadas} material(es) con riesgo de PARADA DE L&#205;NEA</b> &mdash; revisar secci&#243;n de impactos en cada planta
  </div>`:"";
  const origenLabel=origen==="IMP"?" &middot; Imported only":origen==="NAC"?" &middot; Domestic only":"";
  const html=`<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:${C.bg};font-family:Arial,Helvetica,sans-serif;color:${C.tx1}">
<table width="100%" cellpadding="0" cellspacing="0" style="background:${C.bg}">
<tr><td align="center" style="padding:20px 12px">
<table width="680" cellpadding="0" cellspacing="0" style="max-width:680px;width:100%">
<tr><td style="background:#1a1d24;border-radius:8px 8px 0 0;border:1px solid ${C.brd};border-bottom:none;padding:18px 20px">
  <div style="font-size:10px;color:rgba(255,255,255,.55);letter-spacing:1.5px;margin-bottom:4px">SUPPLY STOPLIGHT &mdash; WEEKLY INVENTORY REPORT</div>
  <div style="font-size:20px;font-weight:700;color:#ffffff;margin-bottom:6px">${centroNames}</div>
  <div style="font-size:12px;color:rgba(255,255,255,.65)">Corte: <b style="color:#ffffff">${fecha}</b>${origenLabel}</div>
</td></tr>
<tr><td style="background:${globalColor}22;border-left:4px solid ${globalColor};border:1px solid ${C.brd};border-top:none;border-bottom:none;padding:12px 20px">
  <span style="font-size:14px;font-weight:700;color:${globalColor}">${globalIcon} ${globalText}</span>
  <span style="font-size:11px;color:${C.tx2};margin-left:12px">Attached: domestic POs, import POs, and unreleased PRs requiring action.</span>
</td></tr>
<tr><td style="background:${C.bg1};border:1px solid ${C.brd};border-top:none;border-radius:0 0 8px 8px;padding:16px 20px">
  ${paradaGlobal}${centroBlocks}
  <div style="margin-top:16px;font-size:11px;color:${C.tx2};text-align:center;border-top:1px solid ${C.brd};padding-top:12px">
    Generated by Supply Stoplight v5 &mdash; ${fecha}<br>
    For full stoplight and detailed analysis, see the attached interactive dashboard.
  </div>
</td></tr>
</table></td></tr></table>
</body></html>`;
  const subject=`Supply Stoplight \u2014 ${centroNames} | ${fecha} \u00b7 Weekly Inventory Report`;
  const subjectEnc=`=?UTF-8?B?${btoa(unescape(encodeURIComponent(subject)))}?=`;
  const rawB64=btoa(unescape(encodeURIComponent(html)));
  const wrappedB64=rawB64.match(/.{1,76}/g).join("\r\n");
  const eml=
    "MIME-Version: 1.0\r\n"+
    `Subject: ${subjectEnc}\r\n`+
    "Content-Type: text/html; charset=UTF-8\r\n"+
    "Content-Transfer-Encoding: base64\r\n"+
    "\r\n"+
    wrappedB64;
  const blob=new Blob([eml],{type:"message/rfc822"});
  const a=document.createElement("a");
  a.href=URL.createObjectURL(blob);
  a.download=`Supply_Reporte_${fecha.replace(/[^a-zA-Z0-9]/g,"-")}_${centroNames.replace(/[^a-zA-Z0-9]/g,"_")}.eml`;
  a.click();
  } catch(err) { alert("Error generando email: " + err.message); console.error(err); }
}
function _rptReadFilters() {
  const selected = [..._rptCentroSel];
  const sections = {
    semaforo: document.getElementById("rpt-sec-sem").checked,
    ocSol:    document.getElementById("rpt-sec-ocsol").checked,
  };
  const todosEst = document.getElementById("rpt-est-todos").checked;
  const estados = todosEst ? null : {
    CRITICO:     document.getElementById("rpt-est-crit").checked,
    RIESGO:      document.getElementById("rpt-est-riesgo").checked,
    ALERTA:      document.getElementById("rpt-est-alerta").checked,
    OK:          document.getElementById("rpt-est-ok").checked,
    SIN_CONSUMO: document.getElementById("rpt-est-sc").checked,
  };
  const origen = document.querySelector('input[name="rpt-origen"]:checked')?.value || "ALL";
  return {selected, sections, estados, origen};
}

function generateSelectedReport() {
  const {selected, sections, estados, origen} = _rptReadFilters();
  closeReportModal();
  if (!selected.length) return;
  exportReport(selected, sections, estados, origen);
}

// ── Tendencia de cobertura (cob_hoy vs última semana proyectada) ─────────────
function trendArrow(m) {
  const valid = (m.sabados || []).filter(s => s.cob < 9999 && s.cob > 0);
  if (!valid.length || m.cob_hoy <= 0) return {sym:"—", cls:"trend-flat", title:"No data"};
  const fin = valid[valid.length - 1].cob;
  if (fin > m.cob_hoy * 1.1)  return {sym:"▲", cls:"trend-up",   title:"Mejora → "+fin+"d"};
  if (fin < m.cob_hoy * 0.9)  return {sym:"▼", cls:"trend-down", title:"Empeora → "+fin+"d"};
  return {sym:"—", cls:"trend-flat", title:"Estable → "+fin+"d"};
}

// ── Filtro SOLPEDs interactivas ───────────────────────────────────────────────
let filterSolLib = "ALL"; // "ALL" | "SIN_LIB" | "LIB"
let calView      = "cal"; // "cal" | "list"
function setFilterSolLib(f, el) {
  filterSolLib = f;
  document.querySelectorAll(".sol-lib-btn").forEach(b => b.classList.remove("active"));
  el.classList.add("active");
  renderOCSol();
}

// ── Reporte PDF estratégico por centro ───────────────────────────────────────
function exportReport(centroKeys, sections, estados, origen) {
  sections = sections || {semaforo: true, ocSol: true};
  origen = origen || "ALL";
  const inclSet = estados
    ? new Set(Object.entries(estados).filter(([,v])=>v).map(([k])=>k))
    : null;
  const origenFilter = m => origen === "ALL" ? true : origen === "IMP" ? m.importado === true : m.importado === false;

  const semanas = DATA.meta.sem_weeks.slice(0, 12);
  let sectionsHTML = "";

  const entries = centroKeys
    ? Object.entries(DATA.centros).filter(([k]) => centroKeys.includes(k))
    : Object.entries(DATA.centros);

  const {ref, d7, d21} = getOCDateRef();
  const estOrd = {CRITICO:0, RIESGO:1, ALERTA:2, OK:3, SIN_CONSUMO:4};

  entries.forEach(([key, ct], idx) => {
    const mats        = ct.materiales;
    const filteredMats = (inclSet ? mats.filter(m => inclSet.has(m.estado)) : mats).filter(origenFilter);
    const criticos = filteredMats.filter(m => m.estado === "CRITICO").sort((a,b) => a.cob_hoy - b.cob_hoy);
    const riesgos  = filteredMats.filter(m => m.estado === "RIESGO" ).sort((a,b) => a.cob_hoy - b.cob_hoy);
    const alertas  = filteredMats.filter(m => m.estado === "ALERTA");
    const oks      = filteredMats.filter(m => m.estado === "OK");
    const sinConsOrig = mats.filter(m => m.estado === "SIN_CONSUMO");
    const semMats  = [...filteredMats].sort((a,b) =>
      ((estOrd[a.estado]??5) - (estOrd[b.estado]??5)) || (a.cob_hoy - b.cob_hoy));
    const oc       = ct.oc  || [];
    const sol      = ct.sol || [];
    const filteredMatSet = new Set(filteredMats.map(m => m.mat));

    // ── 1. Resumen ejecutivo — siempre totales del portafolio completo ────────
    const critTotal  = mats.filter(m => m.estado === "CRITICO").length;
    const riesTotal  = mats.filter(m => m.estado === "RIESGO").length;
    const alertTotal = mats.filter(m => m.estado === "ALERTA").length;
    const okTotal    = mats.filter(m => m.estado === "OK").length;
    const scTotal    = mats.filter(m => m.estado === "SIN_CONSUMO").length;

    const ocAtr  = oc.filter(o => o.atrasada);
    const ocVcr  = oc.filter(o => {
      if (!o.fecha_entrega || o.atrasada) return false;
      const fe = new Date(o.fecha_entrega); return fe > ref && fe <= d21;
    });
    const ocOk   = oc.length - ocAtr.length - ocVcr.length;
    const solSL  = sol.filter(s => s.sin_liberar);
    const solLib = sol.length - solSL.length;

    const resumen =
      "<div class='resumen'>" +
        "<div class='res-group'><div class='res-lbl'>Materiales (total portafolio)</div>" +
          "<div class='res-card res-r'><span class='res-v'>" + critTotal  + "</span><span class='res-l'>CR&Iacute;TICO</span></div>" +
          "<div class='res-card res-y'><span class='res-v'>" + riesTotal  + "</span><span class='res-l'>RIESGO</span></div>" +
          "<div class='res-card'><span class='res-v'>"       + alertTotal + "</span><span class='res-l'>ALERTA</span></div>" +
          "<div class='res-card res-g'><span class='res-v'>" + okTotal    + "</span><span class='res-l'>OK</span></div>" +
          "<div class='res-card'><span class='res-v'>"       + scTotal    + "</span><span class='res-l'>INV.SC</span></div>" +
          "<div class='res-card'><span class='res-v'>"       + mats.length + "</span><span class='res-l'>TOTAL</span></div>" +
        "</div>" +
        "<div class='res-group'><div class='res-lbl'>&#211;rdenes de Compra (total)</div>" +
          "<div class='res-card res-r'><span class='res-v'>" + ocAtr.length + "</span><span class='res-l'>Atrasadas</span></div>" +
          "<div class='res-card res-y'><span class='res-v'>" + ocVcr.length + "</span><span class='res-l'>Vencen &le;21d</span></div>" +
          "<div class='res-card res-g'><span class='res-v'>" + ocOk + "</span><span class='res-l'>&gt;21d</span></div>" +
          "<div class='res-card'><span class='res-v'>" + oc.length + "</span><span class='res-l'>Total</span></div>" +
        "</div>" +
        "<div class='res-group'><div class='res-lbl'>SOLPEDs (total)</div>" +
          "<div class='res-card'><span class='res-v'>" + sol.length + "</span><span class='res-l'>Total</span></div>" +
          "<div class='res-card " + (solSL.length>0?"res-r":"res-g") + "'><span class='res-v'>" + solSL.length + "</span><span class='res-l'>Unreleased</span></div>" +
          "<div class='res-card res-g'><span class='res-v'>" + solLib + "</span><span class='res-l'>Releaseds</span></div>" +
        "</div>" +
        (inclSet ? "<div class='res-group'><div class='res-lbl'>Tabla (filtrada)</div>" +
          "<div class='res-card'><span class='res-v'>" + filteredMats.length + "</span><span class='res-l'>Materiales</span></div>" +
        "</div>" : "") +
      "</div>";

    // ── 2. Semáforo agrupado por línea ──────────────────────────────────────
    const dotHdr = semanas.map(w => '<th class="td-dot">' + w.label + '</th>').join("");
    const semRow = m => {
      const clsMap = {CRITICO:"row-crit",RIESGO:"row-risk",ALERTA:"row-alert",OK:"",SIN_CONSUMO:"row-sc"};
      const cls    = clsMap[m.estado] || "";
      const badgeMap = {
        CRITICO: '<span class="badge badge-r">CR&Iacute;T.</span>',
        RIESGO:  '<span class="badge badge-y">RIESGO</span>',
        ALERTA:  '<span class="badge badge-a">ALERTA</span>',
        OK:      '<span class="badge badge-lib">OK</span>',
        SIN_CONSUMO: '<span class="badge badge-sc">INV.SC</span>',
      };
      const badge  = badgeMap[m.estado] || "";
      const cobStr = m.adu_plan > 0 ? (m.cob_hoy > 998 ? "&infin;" : m.cob_hoy + "d") : "&mdash;";
      const cobClsMap = {CRITICO:"cob-r",RIESGO:"cob-y",ALERTA:"cob-a"};
      const cobCls = cobClsMap[m.estado] || "";
      const dots   = m.sabados.slice(0, 12).map(s =>
        '<td class="td-dot"><span class="dot-s ' + s.color + '"></span></td>').join("");
      const ocPend = (m.oc_list || []).reduce((a,o)=>a+o.qty,0);
      const ocAtrM = (m.oc_list || []).filter(o=>o.atrasada).length;
      const imp    = m.importado ? "<span class='imp-tag'>IMP</span>" : "<span class='nal-tag'>NAL</span>";
      const impactoTag = m.impact
        ? (/^parada/i.test(m.impact.trim())
            ? "<span class='parada-tag'>PARADA</span>"
            : "<span class='noparada-tag'>NO PAR.</span>")
        : "";
      const aduRpt = m.adu_plan > 0 ? m.adu_plan : (m.adu_hist > 0 ? m.adu_hist : m.adu);
      const consSem = aduRpt > 0 ? Math.round(aduRpt * 7).toLocaleString("es-CO") : "&mdash;";
      const t = trendArrow(m);
      return "<tr class='" + cls + "'>" +
        "<td class='mono'>" + m.mat + imp + impactoTag + "</td>" +
        "<td class='desc'>" + m.desc + "</td>" +
        "<td>" + m.um + "</td>" +
        "<td class='num'>" + Math.round(m.saldo).toLocaleString("es-CO") + "</td>" +
        "<td class='num'>" + (aduRpt>0?aduRpt.toFixed(1):"&mdash;") + "</td>" +
        "<td class='num'>" + consSem + "</td>" +
        "<td class='num " + cobCls + "'>" + cobStr + "</td>" +
        dots +
        "<td class='" + t.cls + "' title='" + t.title + "'>" + t.sym + "</td>" +
        "<td class='num'>" + (ocPend>0?Math.round(ocPend).toLocaleString("es-CO")+(ocAtrM>0?' <span class="badge badge-r">'+ocAtrM+" atr.</span>":""):"&mdash;") + "</td>" +
        "<td>" + badge + "</td></tr>";
    };

    // Agrupar por línea
    const lineas = {};
    semMats.forEach(m => {
      const l = m.linea_fab || "Sin línea asignada";
      if (!lineas[l]) lineas[l] = [];
      lineas[l].push(m);
    });
    const semTablas = Object.entries(lineas).sort((a,b)=>a[0].localeCompare(b[0])).map(([lin, ms]) =>
      "<h4>L&iacute;nea: " + lin + " &nbsp;<span style='font-weight:normal'>(" + ms.length + " materiales)</span></h4>" +
      "<table><thead><tr>" +
        "<th>C&oacute;digo</th><th>Descripci&oacute;n</th><th>UM</th>" +
        "<th class='num'>Saldo</th><th class='num'>ADU/d</th><th class='num'>Cons.Sem.</th><th class='num'>Cob.Hoy</th>" +
        dotHdr +
        "<th>Tend.</th><th class='num'>OC Pend.</th><th>Estado</th>" +
      "</tr></thead><tbody>" + ms.map(semRow).join("") + "</tbody></table>"
    ).join("");

    // ── 3. OC filtradas: atrasadas + materiales en estados/origen seleccionados ────
    const ocOrigenOk = o => { const m = mats.find(x=>x.mat===o.mat); return m ? origenFilter(m) : true; };
    const ocTodas = oc.filter(o => (o.atrasada || filteredMatSet.has(o.mat)) && ocOrigenOk(o)).sort((a,b) => {
      if (a.atrasada && !b.atrasada) return -1;
      if (!a.atrasada && b.atrasada) return  1;
      if (a.atrasada && b.atrasada)  return b.dias_atraso - a.dias_atraso;
      return (a.fecha_entrega||"").localeCompare(b.fecha_entrega||"");
    });
    const matEstado = m => {
      const obj = mats.find(x=>x.mat===m);
      if (!obj) return "";
      if (obj.estado==="CRITICO")     return "<span class='badge badge-r'>CR&Iacute;T.</span>";
      if (obj.estado==="RIESGO")      return "<span class='badge badge-y'>RIESGO</span>";
      if (obj.estado==="ALERTA")      return "<span class='badge badge-a'>ALERTA</span>";
      if (obj.estado==="OK")          return "<span class='badge badge-lib'>OK</span>";
      if (obj.estado==="SIN_CONSUMO") return "<span class='badge badge-sc'>INV.SC</span>";
      return "";
    };
    const matImportado = m => {
      const obj = mats.find(x=>x.mat===m);
      if (!obj) return "";
      return obj.importado
        ? "<span class='badge' style='background:#7c5cbf;color:#fff'>IMP</span>"
        : "<span class='badge' style='background:#5a8a5a;color:#fff'>NAC</span>";
    };
    const ocToRow = o =>
      "<tr class='" + (o.atrasada?"row-crit":"") + "'>" +
        "<td class='mono'>" + o.doc + "</td>" +
        "<td class='mono'>" + o.mat + "</td>" +
        "<td class='desc'>" + o.desc + "</td>" +
        "<td>" + o.fecha_entrega + (o.atrasada?" <span style='color:#999;font-size:7px'>→ proy: " + (o.fecha_proyectada||"") + "</span>":"") + "</td>" +
        "<td class='num'>" + Math.round(o.qty_pend).toLocaleString("es-CO") + " " + o.um + "</td>" +
        "<td class='num'>" + (o.atrasada?"<span class='badge badge-r'>"+o.dias_atraso+"d</span>":"&mdash;") + "</td>" +
        "<td>" + o.proveedor + "</td>" +
        "<td>" + matImportado(o.mat) + "</td>" +
        "<td>" + matEstado(o.mat) + "</td></tr>";
    const ocHdrRow =
      "<thead><tr>" +
        "<th>N&deg; OC</th><th>Material</th><th>Descripci&oacute;n</th>" +
        "<th>Fecha Entrega</th><th class='num'>Cant. Pend.</th>" +
        "<th class='num'>D&iacute;as Atraso</th><th>Proveedor</th><th>Origen</th><th>Estado Mat.</th>" +
      "</tr></thead>";
    const ocImpRows = ocTodas.filter(o => { const m = mats.find(x=>x.mat===o.mat); return m && m.importado; });
    const ocNacRows = ocTodas.filter(o => { const m = mats.find(x=>x.mat===o.mat); return !m || !m.importado; });
    const ocTableImp =
      "<table>" + ocHdrRow + "<tbody>" +
      (ocImpRows.map(ocToRow).join("") || "<tr><td colspan='9' style='color:#999;text-align:center'>Sin OC importadas</td></tr>") +
      "</tbody></table>";
    const ocTableNac =
      "<table>" + ocHdrRow + "<tbody>" +
      (ocNacRows.map(ocToRow).join("") || "<tr><td colspan='9' style='color:#999;text-align:center'>Sin OC nacionales</td></tr>") +
      "</tbody></table>";

    // ── 4. SOLPEDs: sin liberar + materiales en estados/origen seleccionados ────────
    const solOrigenOk = s => { const m = mats.find(x=>x.mat===s.mat); return m ? origenFilter(m) : true; };
    const solTodas = sol.filter(s => (s.sin_liberar || filteredMatSet.has(s.mat)) && solOrigenOk(s))
      .sort((a,b) => (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""));
    const solAllRows = solTodas.map(s =>
      "<tr class='" + (s.sin_liberar?"row-sinlib":"") + "'>" +
        "<td class='mono'>" + s.doc + "</td>" +
        "<td class='mono'>" + s.mat + "</td>" +
        "<td class='desc'>" + s.desc + "</td>" +
        "<td class='num'>" + Math.round(s.qty).toLocaleString("es-CO") + " " + s.um + "</td>" +
        "<td>" + s.fecha_entrega + "</td>" +
        "<td><span class='badge " + (s.sin_liberar?"badge-sl":"badge-lib") + "'>" + (s.sin_liberar?"Unreleased":"Released") + "</span></td>" +
        "<td>" + matEstado(s.mat) + "</td></tr>"
    ).join("") || "<tr><td colspan='7' style='color:#999;text-align:center'>No PRs</td></tr>";

    // ── 4b. Impactos y acciones ────────────────────────────────────────────────
    const impMats = filteredMats.filter(m => m.impact || m.action);
    const impSection = impMats.length > 0
      ? "<h3>&#9888; Impactos y Acciones Planeador &mdash; " + impMats.length + " material" + (impMats.length>1?"es":"") + "</h3>" +
        "<table class='imp-action-tbl'><thead><tr>" +
          "<th>C&oacute;digo</th><th>Descripci&oacute;n</th><th>Estado</th><th>Impacto</th><th>Acci&oacute;n</th>" +
        "</tr></thead><tbody>" +
        impMats.map(m => {
          const isStop = /^parada/i.test((m.impact||"").trim());
          const rowCls = isStop ? "row-crit" : "row-risk";
          return "<tr class='" + rowCls + "'>" +
            "<td class='mono'>" + m.mat + "</td>" +
            "<td class='desc'>" + m.desc + "</td>" +
            "<td>" + matEstado(m.mat) + "</td>" +
            "<td><b>" + (m.impact||"&mdash;") + "</b></td>" +
            "<td>" + (m.action||"&mdash;") + "</td></tr>";
        }).join("") +
        "</tbody></table>"
      : "";

    const breakClass = idx > 0 ? " section-break" : "";
    const origenLabel = origen === "IMP" ? " · Imported only" : origen === "NAC" ? " · Domestic only" : "";
    const estLabel = (inclSet ? Array.from(inclSet).join(" · ") : "Todos los estados") + origenLabel;
    const semSection = sections.semaforo
      ? "<h3>Sem&aacute;foro por l&iacute;nea &mdash; " + estLabel + " (" + semMats.length + " mat.)</h3>" +
        (semMats.length ? semTablas : "<p style='color:#999;font-size:8px;margin:4px 8px'>No materials for the selected statuses</p>")
      : "";
    const ocSolSection = sections.ocSol
      ? "<h3>&#127758; Importados &mdash; " + ocImpRows.length + " OC</h3>" + ocTableImp +
        "<h3>&#127981; Nacionales / Locales &mdash; " + ocNacRows.length + " OC</h3>" + ocTableNac +
        "<h3>SOLPEDs (" + solTodas.length + " de " + sol.length + ")</h3>" +
        "<table><thead><tr>" +
          "<th>N&deg; SOLPED</th><th>Material</th><th>Descripci&oacute;n</th>" +
          "<th class='num'>Cantidad</th><th>Fecha Entrega</th><th>Estado Lib.</th><th>Estado Mat.</th>" +
        "</tr></thead><tbody>" + solAllRows + "</tbody></table>"
      : "";
    sectionsHTML +=
      "<div class='section" + breakClass + "'>" +
      "<h2>" + ct.nombre +
        "<span>CR&Iacute;TICOS: " + critTotal + " &nbsp;|&nbsp; RIESGO: " + riesTotal +
        " &nbsp;|&nbsp; ALERTA: " + alertTotal +
        " &nbsp;|&nbsp; OK: " + okTotal +
        " &nbsp;|&nbsp; INV.SC: " + scTotal +
        " &nbsp;|&nbsp; TOTAL: " + mats.length + " mat." +
        " &nbsp;&middot;&nbsp; OC: " + oc.length + " &nbsp;&middot;&nbsp; SP: " + sol.length +
        (inclSet ? " &nbsp;<span style='opacity:.7'>(tabla filtrada: " + filteredMats.length + " mat.)</span>" : "") +
        "</span></h2>" +
      resumen + semSection + impSection + ocSolSection +
      "</div>";
  });  // end entries.forEach

  const css =
    "@page{size:A4 landscape;margin:12mm}" +
    "body{font-family:Arial,sans-serif;font-size:9px;color:#111;-webkit-print-color-adjust:exact;print-color-adjust:exact}" +
    // Header
    ".rpt-header{display:flex;justify-content:space-between;align-items:flex-end;margin-bottom:10px;border-bottom:2px solid #c0450d;padding-bottom:6px}" +
    ".rpt-title{font-size:14px;font-weight:700;color:#c0450d}" +
    ".rpt-meta{font-size:8px;color:#666}" +
    // Sección — sin page-break-inside en .section para no dejar leyenda sola
    ".section{margin-bottom:16px}" +
    ".section-break{page-break-before:always}" +
    "h2{font-size:11px;background:#c0450d;color:white;padding:5px 10px;margin:0 0 6px 0;page-break-after:avoid;page-break-inside:avoid;display:flex;justify-content:space-between;align-items:center}" +
    "h2 span{font-weight:normal;font-size:9px;opacity:.85}" +
    "h3{font-size:9px;background:#f0f2f5;color:#3a3f4b;padding:3px 8px;margin:10px 0 2px 0;border-left:3px solid #c0450d;page-break-after:avoid;font-weight:600}" +
    "h4{font-size:8px;background:#f5f0ff;color:#5a3a9a;padding:2px 8px;margin:6px 0 2px 0;border-left:3px solid #7c5cbf;page-break-after:avoid;font-weight:600}" +
    // Resumen ejecutivo cards
    ".resumen{display:flex;gap:6px;margin-bottom:10px;flex-wrap:wrap}" +
    ".res-group{display:flex;gap:4px;padding-right:10px;border-right:1px solid #ccc}" +
    ".res-group:last-child{border-right:none}" +
    ".res-lbl{font-size:7px;color:#888;text-transform:uppercase;letter-spacing:.5px;margin-bottom:4px;font-weight:600}" +
    ".res-card{background:#f8f9fc;border:1px solid #dde;border-radius:4px;padding:4px 8px;min-width:52px;text-align:center}" +
    ".res-v{font-size:14px;font-weight:700;line-height:1;display:block}" +
    ".res-l{font-size:7px;color:#666;margin-top:1px;display:block}" +
    ".res-r .res-v{color:#cc2222}" +
    ".res-y .res-v{color:#b07000}" +
    ".res-g .res-v{color:#1a9e6a}" +
    ".res-b .res-v{color:#3355bb}" +
    // Tablas
    "table{border-collapse:collapse;width:100%;margin-bottom:2px}" +
    "th{background:#f0f2f8;font-size:8px;padding:3px 5px;border:1px solid #bbb;white-space:nowrap;text-align:left}" +
    "td{font-size:8px;padding:2px 5px;border:1px solid #ddd;white-space:nowrap}" +
    ".desc{max-width:160px;overflow:hidden;text-overflow:ellipsis}" +
    ".row-crit td{background:#fff0f0}.row-risk td{background:#fffae6}.row-alert td{background:#fff8ee}.row-sc td{background:#f4f4f8}.row-sinlib td{background:#fffff0}" +
    ".badge{padding:1px 4px;border-radius:5px;font-size:7px;font-weight:700;white-space:nowrap}" +
    ".badge-r{background:#ff4d4d;color:#fff}.badge-y{background:#f5a623;color:#fff}" +
    ".badge-a{background:#e8870a;color:#fff}.badge-sc{background:#7c8aaa;color:#fff}" +
    ".badge-sl{background:#e0a800;color:#fff}.badge-lib{background:#3ecf8e;color:#fff}" +
    ".num{text-align:right;font-family:monospace}.mono{font-family:monospace;font-size:8px}" +
    ".cob-r{color:#cc0000;font-weight:700}.cob-y{color:#b07000;font-weight:700}.cob-a{color:#e8870a;font-weight:700}" +
    ".dot-s{display:inline-block;width:9px;height:9px;border-radius:50%}" +
    ".td-dot{text-align:center;padding:2px;min-width:26px}" +
    ".G{background:#3ecf8e}.Y{background:#f5a623}.R{background:#ff4d4d}.K{background:#999}.B{background:#4fa3e0}" +
    ".imp-tag{font-size:7px;background:#7c5cbf;color:#fff;border-radius:3px;padding:0 3px;margin-left:2px}" +
    ".nal-tag{font-size:7px;background:#1e3a2e;color:#3ecf8e;border-radius:3px;padding:0 3px;margin-left:2px}" +
    ".parada-tag{font-size:7px;background:#cc2222;color:#fff;border-radius:3px;padding:0 3px;margin-left:2px;font-weight:700}" +
    ".noparada-tag{font-size:7px;background:#1a9e6a;color:#fff;border-radius:3px;padding:0 3px;margin-left:2px}" +
    ".imp-action-tbl td:nth-child(4){color:#cc2222;font-weight:700}" +
    ".trend-up{color:#3ecf8e;font-size:10px;font-weight:700;text-align:center}" +
    ".trend-down{color:#e05c5c;font-size:10px;font-weight:700;text-align:center}" +
    ".trend-flat{color:#888;font-size:10px;text-align:center}" +
    ".leyenda{font-size:7px;padding:3px 8px;background:#1a1a2e;color:#ccc;margin-bottom:6px;border-radius:4px;display:flex;gap:6px;align-items:center;flex-wrap:nowrap;white-space:nowrap;page-break-after:avoid}" +
    ".leyenda .dot-s{vertical-align:middle;width:8px;height:8px;flex-shrink:0}";

  const _fechaSaldo = new Date(DATA.meta.fecha_saldo);
  const _hoy = new Date();
  const _diasAnt = Math.floor((_hoy - _fechaSaldo) / 86400000);
  const antLabel = _diasAnt > 2
    ? "<span style='color:#e05c5c;font-weight:700'>&#9888; Datos con " + _diasAnt + "d de antig&uuml;edad</span>"
    : "<span style='color:#3ecf8e'>&#10003; Datos frescos (" + _diasAnt + "d)</span>";

  const leyenda =
    "<div class='leyenda'>" +
      "<span class='dot-s R'></span>&nbsp;<b>CR&Iacute;TICO:</b> quiebre activo o &lt; zona roja &nbsp;|&nbsp;" +
      "<span class='dot-s Y'></span>&nbsp;<b>RIESGO:</b> zona amarilla DDMRP &nbsp;|&nbsp;" +
      "<span class='dot-s B'></span>&nbsp;<b>OK+:</b> cob &gt; 90d &nbsp;|&nbsp;" +
      "<span class='dot-s G'></span>&nbsp;<b>OK:</b> adequate coverage &nbsp;|&nbsp;" +
      "<span class='dot-s K'></span>&nbsp;<b>INV.SC:</b> no active demand" +
    "</div>";

  const html =
    "<!DOCTYPE html><html><head><meta charset='utf-8'>" +
    "<title>Supply Planning &mdash; Reporte Cr&iacute;ticos</title>" +
    "<style>" + css + "</style></head><body>" +
    "<div class='rpt-header'>" +
      "<div><div class='rpt-title'>Supply Stoplight &mdash; Critical &amp; At-Risk Materials</div>" +
      "<div class='rpt-meta' style='margin-top:2px'>Stock: <b>" + DATA.meta.fecha_saldo +
      "</b> &nbsp;&middot;&nbsp; B2Wise: <b>" + DATA.meta.fecha_b2wise +
      "</b> &nbsp;&middot;&nbsp; M+0=<b>" + DATA.meta.mes_m0 + "</b>" +
      "&nbsp;&middot;&nbsp;" + antLabel + "</div></div>" +
      "<div class='rpt-meta'>Generado: " + new Date().toLocaleDateString("es-CO") + "</div>" +
    "</div>" + leyenda + sectionsHTML + "</body></html>";

  const w = window.open("", "_blank", "width=1200,height=800");
  w.document.write(html);
  w.document.close();
  w.focus();
  setTimeout(() => w.print(), 600);
}

function exportCSV() {
  const ms    = getFiltered();
  const weeks = DATA.meta.sem_weeks;
  const hdr   = ["Material","Description","UM","Ini.Inv","Receipts","Consumption","Stock","ADU",
                  "Cov.Today",...weeks.map(w=>w.label),"B2W Zone"].join(",");
  const rows  = ms.map(m => [
    m.mat, `"${m.desc}"`, m.um,
    m.inv_ini, m.ingresos, m.consumos, m.saldo, m.adu_plan, m.cob_hoy,
    ...m.sabados.map(s=>s.cob), m.zona_b2
  ].join(",")).join("\n");
  const blob = new Blob(["\uFEFF"+hdr+"\n"+rows], {type:"text/csv;charset=utf-8"});
  const a = document.createElement("a");
  a.href = URL.createObjectURL(blob);
  a.download = `supply_stoplight_${DATA.meta.fecha_saldo.replace(/-/g,"")}.csv`;
  a.click();
}

// ── Export OC / SOLPED directo desde tabla ────────────────────────────────────
function exportOCTabXLSX() {
  if (typeof XLSX === "undefined") { alert("SheetJS not loaded. Check your internet connection."); return; }
  const flt = search.toLowerCase();
  let oc = getOC().filter(o => (!selMat.size || selMat.has(o.mat)) && (!flt || o.mat.includes(flt) || o.desc.toLowerCase().includes(flt)));
  if (ocSort === "fecha_asc")        oc = [...oc].sort((a,b) => (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""));
  else if (ocSort === "fecha_desc")  oc = [...oc].sort((a,b) => (b.fecha_entrega||"").localeCompare(a.fecha_entrega||""));
  else if (ocSort === "atraso_desc") oc = [...oc].sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return b.dias_atraso-a.dias_atraso; });
  else if (ocSort === "estado")      oc = [...oc].sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""); });

  const getImp = mat => { const m = cd().materiales.find(x=>x.mat===mat); return m?(m.importado?"Importado":"Nacional"):""; };
  const getEst = mat => { const m = cd().materiales.find(x=>x.mat===mat); return m?m.estado:""; };

  const hdr = ["N° OC","Material","Descripción","Cant. Pedida","Pendiente","UM","Fecha Doc.","Fecha Entrega (SAP)","Fecha Proyectada","Días Atraso","Proveedor","Origen","Estado Mat."];
  const rows = oc.map(o => [
    o.doc, o.mat, o.desc,
    o.qty_pedida, o.qty_pend, o.um,
    o.fecha_doc, o.fecha_entrega,
    o.atrasada ? o.fecha_proyectada : null,
    o.atrasada ? o.dias_atraso : null,
    o.proveedor, getImp(o.mat), getEst(o.mat)
  ]);
  const ws = XLSX.utils.aoa_to_sheet([hdr, ...rows]);
  ws["!cols"] = [12,12,40,12,12,6,12,14,14,10,30,12,12].map(w=>({wch:w}));
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, "Open POs");
  XLSX.writeFile(wb, `OC_${cd().nombre}_${DATA.meta.fecha_saldo.replace(/-/g,"")}.xlsx`);
}

function exportSOLTabXLSX() {
  if (typeof XLSX === "undefined") { alert("SheetJS not loaded. Check your internet connection."); return; }
  const flt = search.toLowerCase();
  const sol = getSOL().filter(s => (!selMat.size || selMat.has(s.mat)) && (!flt || s.mat.includes(flt) || s.desc.toLowerCase().includes(flt)));

  const getEst = mat => { const m = cd().materiales.find(x=>x.mat===mat); return m?m.estado:""; };

  const hdr = ["N° SOLPED","Material","Descripción","Cantidad","UM","Fecha Entrega","Estado Lib.","Estado Mat."];
  const rows = sol.map(s => [
    s.doc, s.mat, s.desc,
    s.qty, s.um, s.fecha_entrega,
    s.sin_liberar ? "Unreleased" : "Released",
    getEst(s.mat)
  ]);
  const ws = XLSX.utils.aoa_to_sheet([hdr, ...rows]);
  ws["!cols"] = [12,12,40,12,6,12,14,12].map(w=>({wch:w}));
  const wb = XLSX.utils.book_new();
  XLSX.utils.book_append_sheet(wb, ws, "PRs");
  XLSX.writeFile(wb, `SOLPEDs_${cd().nombre}_${DATA.meta.fecha_saldo.replace(/-/g,"")}.xlsx`);
}

// ── Reporte Excel (XLSX) ─────────────────────────────────────────────────────
function generateSelectedXLSX() {
  const {selected, sections, estados, origen} = _rptReadFilters();
  closeReportModal();
  if (!selected.length) return;
  exportReportXLSX(selected, sections, estados, origen);
}

function exportReportXLSX(centroKeys, sections, estados, origen) {
  if (typeof XLSX === "undefined") { alert("SheetJS not loaded. Check your internet connection."); return; }
  sections = sections || {semaforo:true, ocSol:true};
  origen = origen || "ALL";
  const inclSet = estados ? new Set(Object.entries(estados).filter(([,v])=>v).map(([k])=>k)) : null;
  const origenFilter = m => origen === "ALL" ? true : origen === "IMP" ? m.importado === true : m.importado === false;
  const wb = XLSX.utils.book_new();
  const semanas = DATA.meta.sem_weeks.slice(0, 8);
  const entries = centroKeys
    ? Object.entries(DATA.centros).filter(([k]) => centroKeys.includes(k))
    : Object.entries(DATA.centros);
  const estOrd = {CRITICO:0, RIESGO:1, ALERTA:2, OK:3, SIN_CONSUMO:4};
  const {ref, d7, d21} = getOCDateRef();

  // ── Hoja Executive Summary ────────────────────────────────────────────────
  const resHdr1 = ["Supply Stoplight  —  Report  |  Stock: " + DATA.meta.fecha_saldo + "  |  B2Wise: " + DATA.meta.fecha_b2wise + "  |  " + DATA.meta.mes_m0];
  const resHdr2 = [];
  const resHdr3 = [
    "Plant",
    "Critical","At Risk","Alert","OK","Inv.SC","Total Mat.",
    "","Overdue POs","POs Due ≤21d","POs >21d","Total POs",
    "","Unreleased PRs","Released PRs","Total PRs"
  ];
  const resRows = entries.map(([key, ct]) => {
    const mats = ct.materiales;
    const oc   = ct.oc  || [];
    const sol  = ct.sol || [];
    const ocAtr = oc.filter(o => o.atrasada).length;
    const ocVcr = oc.filter(o => { if(!o.fecha_entrega||o.atrasada)return false; const fe=new Date(o.fecha_entrega); return fe>ref&&fe<=d21; }).length;
    const ocOk  = oc.length - ocAtr - ocVcr;
    const solSL = sol.filter(s=>s.sin_liberar).length;
    return [
      ct.nombre,
      mats.filter(m=>m.estado==="CRITICO").length,
      mats.filter(m=>m.estado==="RIESGO").length,
      mats.filter(m=>m.estado==="ALERTA").length,
      mats.filter(m=>m.estado==="OK").length,
      mats.filter(m=>m.estado==="SIN_CONSUMO").length,
      mats.length,
      "",
      ocAtr, ocVcr, ocOk, oc.length,
      "",
      solSL, sol.length-solSL, sol.length
    ];
  });
  const wsRes = XLSX.utils.aoa_to_sheet([resHdr1, resHdr2, resHdr3, ...resRows]);
  wsRes["!cols"] = [18,9,9,9,9,9,10,2,12,14,10,10,2,14,14,10].map(w=>({wch:w}));
  wsRes["!merges"] = [{s:{r:0,c:0},e:{r:0,c:15}}];
  XLSX.utils.book_append_sheet(wb, wsRes, "Executive Summary");

  entries.forEach(([key, ct]) => {
    const mats        = ct.materiales;
    const filteredMats = (inclSet ? mats.filter(m => inclSet.has(m.estado)) : mats).filter(origenFilter);
    const semMats     = [...filteredMats].sort((a,b) => ((estOrd[a.estado]??5)-(estOrd[b.estado]??5)) || (a.cob_hoy-b.cob_hoy));
    const oc          = ct.oc  || [];
    const sol         = ct.sol || [];
    const filteredMatSet = new Set(filteredMats.map(m => m.mat));
    const matImp = m => { const o = mats.find(x=>x.mat===m); return o ? (o.importado?"Importado":"Nacional") : ""; };
    const matEst = m => { const o = mats.find(x=>x.mat===m); return o ? o.estado : ""; };
    const sheetSuffix = ct.nombre.substring(0,20).replace(/[:\\/?*\[\]]/g,"");

    // ── Hoja Semáforo ──────────────────────────────────────────────────────
    if (sections.semaforo && semMats.length) {
      const semTitle = [ct.nombre + " — Semáforo  |  CRÍTICOS: " +
        mats.filter(m=>m.estado==="CRITICO").length + "  |  RIESGO: " +
        mats.filter(m=>m.estado==="RIESGO").length + "  |  ALERTA: " +
        mats.filter(m=>m.estado==="ALERTA").length + "  |  OK: " +
        mats.filter(m=>m.estado==="OK").length + "  |  TOTAL: " + mats.length + " mat."];
      const hdr = ["Material","Descripción","UM","Stock","ADU/d","Cov.Today",
                   ...semanas.map(w=>w.label),"B2W Zone","Estado","Importado","Línea Fab.","OC Pend.",
                   "Impacto Planta","Acción Planeador"];
      const rows = semMats.map(m => {
        const ocPend = (m.oc_list||[]).reduce((a,o)=>a+o.qty,0);
        return [m.mat, m.desc, m.um, m.saldo,
                m.adu_plan>0?m.adu_plan:null, m.adu_plan>0?(m.cob_hoy<9999?m.cob_hoy:null):null,
                ...m.sabados.slice(0,8).map(s => s.cob<9999?s.cob:null),
                m.zona_b2, m.estado, m.importado?"Importado":"Nacional",
                m.linea_fab||"", ocPend>0?ocPend:null,
                m.impact||"", m.action||""];
      });
      const ws = XLSX.utils.aoa_to_sheet([semTitle, [], hdr, ...rows]);
      ws["!cols"] = [10,40,6,12,8,8,...semanas.map(()=>8),10,12,12,20,10,22,40].map(w=>({wch:w}));
      ws["!merges"] = [{s:{r:0,c:0},e:{r:0,c:hdr.length-1}}];
      XLSX.utils.book_append_sheet(wb, ws, (sheetSuffix+" Sem").substring(0,31));
    }

    // ── Hoja OC ────────────────────────────────────────────────────────────
    if (sections.ocSol) {
      const ocOrigenOk = o => { const m = mats.find(x=>x.mat===o.mat); return m ? origenFilter(m) : true; };
      const ocTodas = oc.filter(o => (o.atrasada || filteredMatSet.has(o.mat)) && ocOrigenOk(o))
        .sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return b.dias_atraso-a.dias_atraso; });
      const ocAtr = ocTodas.filter(o=>o.atrasada).length;
      const ocVcr = ocTodas.filter(o=>{ if(!o.fecha_entrega||o.atrasada)return false; const fe=new Date(o.fecha_entrega); return fe>ref&&fe<=d21; }).length;
      const ocTitle = [ct.nombre + " — Órdenes de Compra  |  CRÍTICOS: " +
        mats.filter(m=>m.estado==="CRITICO").length + "  |  RIESGO: " +
        mats.filter(m=>m.estado==="RIESGO").length + "  |  ALERTA: " +
        mats.filter(m=>m.estado==="ALERTA").length + "  |  OK: " +
        mats.filter(m=>m.estado==="OK").length + "  |  TOTAL: " + mats.length + " mat."];
      const ocResumen = ["","Atrasadas","Vencen ≤21d","Con fecha >21d","Total POs"];
      const ocResumenVals = ["",ocAtr,ocVcr,ocTodas.length-ocAtr-ocVcr,ocTodas.length];
      const ocHdr = ["N° OC","Material","Descripción","Fecha Entrega (SAP)","Cant. Pend.","UM","Días Atraso","Proveedor","Origen","Estado Mat."];
      const toOcRow = o => [
        o.doc, o.mat, o.desc, o.fecha_entrega,
        o.qty_pend, o.um,
        o.atrasada ? o.dias_atraso : null,
        o.proveedor, matImp(o.mat), matEst(o.mat)
      ];
      // Separar importados vs nacionales
      const ocImp = ocTodas.filter(o => { const m = mats.find(x=>x.mat===o.mat); return m && m.importado; });
      const ocNac = ocTodas.filter(o => { const m = mats.find(x=>x.mat===o.mat); return !m || !m.importado; });
      const titleImp = ["IMPORTADOS — " + ocImp.length + " OC"];
      const titleNac = ["NACIONALES / LOCALES — " + ocNac.length + " OC"];
      // Calcular filas para merges dinámicos
      // Bloque fijo: title(0) empty(1) resumen(2) resumenVals(3) empty(4) → 5 filas
      const rowTitleImp = 5;
      const rowTitleNac = rowTitleImp + 1 + ocImp.length + 1; // titleImp + ocHdr + rows + empty
      const ncols = 9; // índice última columna (0-based)
      const wsOC = XLSX.utils.aoa_to_sheet([
        ocTitle, [], ocResumen, ocResumenVals, [],
        titleImp, ocHdr, ...ocImp.map(toOcRow), [],
        titleNac, ocHdr, ...ocNac.map(toOcRow)
      ]);
      wsOC["!cols"] = [12,12,40,14,12,6,10,30,12,12].map(w=>({wch:w}));
      wsOC["!merges"] = [
        {s:{r:0,c:0},e:{r:0,c:ncols}},
        {s:{r:rowTitleImp,c:0},e:{r:rowTitleImp,c:ncols}},
        {s:{r:rowTitleNac,c:0},e:{r:rowTitleNac,c:ncols}}
      ];
      XLSX.utils.book_append_sheet(wb, wsOC, (sheetSuffix+" OC").substring(0,31));

      // ── Hoja SOLPEDs ────────────────────────────────────────────────────
      const solOrigenOk = s => { const m = mats.find(x=>x.mat===s.mat); return m ? origenFilter(m) : true; };
      const solTodas = sol.filter(s => (s.sin_liberar || filteredMatSet.has(s.mat)) && solOrigenOk(s))
        .sort((a,b) => (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""));
      const solSL = solTodas.filter(s=>s.sin_liberar).length;
      const solTitle = [ct.nombre + " — SOLPEDs  |  Unreleased: " + solSL + "  |  Releaseds: " + (solTodas.length-solSL) + "  |  Total: " + solTodas.length];
      const solHdr = ["N° SOLPED","Material","Descripción","Cantidad","UM","Fecha Entrega","Estado Lib.","Estado Mat."];
      const solRows = solTodas.map(s => [
        s.doc, s.mat, s.desc, s.qty, s.um, s.fecha_entrega,
        s.sin_liberar?"Unreleased":"Released", matEst(s.mat)
      ]);
      const wsSol = XLSX.utils.aoa_to_sheet([solTitle, [], solHdr, ...solRows]);
      wsSol["!cols"] = [12,12,40,12,6,12,14,12].map(w=>({wch:w}));
      wsSol["!merges"] = [{s:{r:0,c:0},e:{r:0,c:7}}];
      XLSX.utils.book_append_sheet(wb, wsSol, (sheetSuffix+" SOL").substring(0,31));
    }
  });

  if (wb.SheetNames.length <= 1) { alert("No data for the selected filters."); return; }
  const fname = `supply_planner_${DATA.meta.fecha_saldo.replace(/-/g,"")}.xlsx`;
  XLSX.writeFile(wb, fname);
}

// ── Init ──────────────────────────────────────────────────────────────────────
window.onload = () => {
  document.getElementById("meta-saldo").textContent      = DATA.meta.fecha_saldo;
  document.getElementById("meta-b2wise").textContent     = DATA.meta.fecha_b2wise;
  document.getElementById("meta-mes").textContent        = "M+0=" + DATA.meta.mes_m0;
  document.getElementById("meta-centro-nom").textContent = [...centrosSel].map(k=>DATA.centros[k].nombre).join(" + ");
  _updateThemeBtn();
  renderCountryBar();
  updateCentroLabel();
  renderGroupSelect();
  renderLineaSelect();
  renderKPIs();
  renderList();
  renderInventario();
};
</script>

<!-- Modal selección de centros para PDF -->
<div id="report-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;align-items:center;justify-content:center">
  <div style="background:var(--bg1);border:1px solid var(--brd);border-radius:10px;padding:28px 32px;min-width:360px;max-width:500px;max-height:90vh;overflow-y:auto;box-shadow:0 8px 40px #0007">
    <div style="font-size:15px;font-weight:700;color:var(--tx1);margin-bottom:6px">&#128196; Generate Report</div>

    <!-- País -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">Country</div>
    <div class="rpt-country-bar" id="rpt-country-bar"></div>
    <!-- Centros -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:6px">Plants</div>
    <div style="margin-bottom:18px">
      <div class="rpt-msel-wrap">
        <button class="rpt-msel-btn" id="rpt-centro-btn" onclick="toggleRptCentroPanel(event)">
          <span id="rpt-centro-label">All plants</span>
          <span style="color:var(--tx2);font-size:10px">&#9660;</span>
        </button>
        <div class="rpt-msel-panel" id="rpt-centro-panel">
          <div class="rpt-msel-search">
            <span style="color:var(--tx2)">&#128269;</span>
            <input type="text" id="rpt-centro-search" placeholder="Buscar centro&#8230;" oninput="renderRptCentroPanel()" onclick="event.stopPropagation()">
          </div>
          <div class="rpt-msel-scroll" id="rpt-centro-scroll"></div>
        </div>
      </div>
    </div>

    <!-- Secciones -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Sections to include</div>
    <div style="display:flex;flex-direction:column;gap:8px;margin-bottom:18px">
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="rpt-sec-sem" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e">
        &#127939; Material Stoplight
      </label>
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="rpt-sec-ocsol" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e">
        &#128666; Open POs &amp; PRs
      </label>
    </div>

    <!-- Material origin -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Material origin</div>
    <div style="display:flex;gap:16px;margin-bottom:18px">
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="rpt-origen" value="ALL" checked style="accent-color:#3ecf8e;cursor:pointer;width:14px;height:14px">
        <b>All</b>
      </label>
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="rpt-origen" value="IMP" style="accent-color:#7c5cbf;cursor:pointer;width:14px;height:14px">
        <span style="background:#7c5cbf;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">IMPORTED</span>
      </label>
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="rpt-origen" value="NAC" style="accent-color:#3ecf8e;cursor:pointer;width:14px;height:14px">
        <span style="background:#1e6b3a;color:#3ecf8e;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">DOMESTIC</span>
      </label>
    </div>

    <!-- Estados -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Statuses to include</div>
    <div style="display:flex;flex-direction:column;gap:8px;margin-bottom:20px">
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="rpt-est-todos" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e" onchange="rptEstToggleAll(this.checked)">
        <b>All statuses</b>
      </label>
      <div id="rpt-est-detail" style="display:none;flex-direction:column;gap:8px;padding-left:24px">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="rpt-est-crit" checked style="width:14px;height:14px;cursor:pointer;accent-color:#ff4d4d">
          <span style="background:#ff4d4d;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">CR&Iacute;TICO</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="rpt-est-riesgo" checked style="width:14px;height:14px;cursor:pointer;accent-color:#f5a623">
          <span style="background:#f5a623;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">RIESGO</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="rpt-est-alerta" checked style="width:14px;height:14px;cursor:pointer;accent-color:#e8870a">
          <span style="background:#e8870a;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">ALERTA</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="rpt-est-ok" style="width:14px;height:14px;cursor:pointer;accent-color:#3ecf8e">
          <span style="background:#3ecf8e;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">OK</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="rpt-est-sc" style="width:14px;height:14px;cursor:pointer;accent-color:#7c8aaa">
          <span style="background:#7c8aaa;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">NO DEMAND</span>
        </label>
      </div>
    </div>

    <div style="display:flex;gap:8px;align-items:center">
      <button class="btn" style="background:var(--bg3);margin-right:auto" onclick="closeReportModal()">Cancel</button>
      <button class="btn" style="background:#1e6b3a;color:#fff;font-weight:700" onclick="generateSelectedXLSX()">&#128202; Excel</button>
      <button class="btn" style="background:#1a4a7a;color:#fff;font-weight:700" onclick="generateEmailFromModal()">&#128140; Prepare Email</button>
      <button class="btn" style="background:#3ecf8e;color:#0f1117;font-weight:700" onclick="generateSelectedReport()">&#128196; Generate PDF</button>
    </div>
  </div>
</div>
<script>
function rptEstToggleAll(checked) {
  const detail = document.getElementById("rpt-est-detail");
  detail.style.display = checked ? "none" : "flex";
  ["rpt-est-crit","rpt-est-riesgo","rpt-est-alerta","rpt-est-ok","rpt-est-sc"]
    .forEach(id => { document.getElementById(id).checked = true; });
}
</script>

<!-- Modal Export OC Excel -->
<div id="oc-excel-modal" style="display:none;position:fixed;inset:0;background:rgba(0,0,0,.65);z-index:9999;align-items:center;justify-content:center">
  <div style="background:var(--bg1);border:1px solid var(--brd);border-radius:10px;padding:28px 32px;min-width:360px;max-width:500px;max-height:90vh;overflow-y:auto;box-shadow:0 8px 40px #0007">
    <div style="font-size:15px;font-weight:700;color:var(--tx1);margin-bottom:6px">&#128202; Export POs &amp; PRs — Excel</div>

    <!-- Centros -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Plants</div>
    <div id="oc-excel-centro-list" style="display:flex;flex-direction:column;gap:8px;margin-bottom:8px"></div>
    <div style="display:flex;gap:8px;margin-bottom:18px">
      <button class="btn" style="font-size:10px;padding:4px 10px" onclick="document.querySelectorAll('.oc-excel-centro-chk').forEach(c=>c.checked=true)">All</button>
      <button class="btn" style="font-size:10px;padding:4px 10px" onclick="document.querySelectorAll('.oc-excel-centro-chk').forEach(c=>c.checked=false)">None</button>
    </div>

    <!-- Secciones -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Include</div>
    <div style="display:flex;flex-direction:column;gap:8px;margin-bottom:18px">
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="oc-excel-inc-oc" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e">
        &#128666; Purchase Orders (POs)
      </label>
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="oc-excel-inc-sol" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e">
        &#128203; Purchase Requisitions (PRs)
      </label>
    </div>

    <!-- Estados -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Material status</div>
    <div style="display:flex;flex-direction:column;gap:8px;margin-bottom:20px">
      <label style="display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="checkbox" id="oc-excel-est-todos" checked style="width:15px;height:15px;cursor:pointer;accent-color:#3ecf8e" onchange="ocExcelEstToggle(this.checked)">
        <b>All statuses</b>
      </label>
      <div id="oc-excel-est-detail" style="display:none;flex-direction:column;gap:8px;padding-left:24px">
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="oc-excel-est-crit" checked style="width:14px;height:14px;cursor:pointer;accent-color:#ff4d4d">
          <span style="background:#ff4d4d;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">CR&Iacute;TICO</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="oc-excel-est-riesgo" checked style="width:14px;height:14px;cursor:pointer;accent-color:#f5a623">
          <span style="background:#f5a623;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">RIESGO</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="oc-excel-est-alerta" checked style="width:14px;height:14px;cursor:pointer;accent-color:#e8870a">
          <span style="background:#e8870a;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">ALERTA</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="oc-excel-est-ok" checked style="width:14px;height:14px;cursor:pointer;accent-color:#3ecf8e">
          <span style="background:#3ecf8e;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">OK</span>
        </label>
        <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
          <input type="checkbox" id="oc-excel-est-sc" checked style="width:14px;height:14px;cursor:pointer;accent-color:#7c8aaa">
          <span style="background:#7c8aaa;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">NO DEMAND</span>
        </label>
      </div>
    </div>

    <!-- Material origin -->
    <div style="font-size:11px;font-weight:600;color:var(--tx2);text-transform:uppercase;letter-spacing:.5px;margin-bottom:8px">Material origin</div>
    <div style="display:flex;gap:16px;margin-bottom:20px">
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="oc-excel-origen" value="ALL" checked style="accent-color:#3ecf8e;cursor:pointer;width:14px;height:14px">
        <b>All</b>
      </label>
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="oc-excel-origen" value="IMP" style="accent-color:#7c5cbf;cursor:pointer;width:14px;height:14px">
        <span style="background:#7c5cbf;color:#fff;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">IMPORTED</span>
      </label>
      <label style="display:flex;align-items:center;gap:8px;cursor:pointer;color:var(--tx1);font-size:12px">
        <input type="radio" name="oc-excel-origen" value="NAC" style="accent-color:#3ecf8e;cursor:pointer;width:14px;height:14px">
        <span style="background:#1e6b3a;color:#3ecf8e;border-radius:4px;padding:1px 7px;font-size:10px;font-weight:700">DOMESTIC</span>
      </label>
    </div>

    <div style="display:flex;gap:10px;justify-content:flex-end">
      <button class="btn" style="background:var(--bg3)" onclick="document.getElementById('oc-excel-modal').style.display='none'">Cancel</button>
      <button class="btn" style="background:#1e6b3a;color:#fff;font-weight:700" onclick="runOCExcelExport()">&#128202; Download Excel</button>
    </div>
  </div>
</div>
<script>
function ocExcelEstToggle(checked) {
  document.getElementById("oc-excel-est-detail").style.display = checked ? "none" : "flex";
  ["oc-excel-est-crit","oc-excel-est-riesgo","oc-excel-est-alerta","oc-excel-est-ok","oc-excel-est-sc"]
    .forEach(id => { document.getElementById(id).checked = true; });
}
function openOCExcelModal() {
  document.getElementById("oc-excel-modal").style.display = "flex";
}
function runOCExcelExport() {
  if (typeof XLSX === "undefined") { alert("SheetJS not loaded. Check your internet connection."); return; }
  const selCentros = Array.from(document.querySelectorAll(".oc-excel-centro-chk:checked")).map(c=>c.value);
  if (!selCentros.length) return;
  const incOC  = document.getElementById("oc-excel-inc-oc").checked;
  const incSOL = document.getElementById("oc-excel-inc-sol").checked;
  const todosEst = document.getElementById("oc-excel-est-todos").checked;
  const estSet = todosEst ? null : new Set(
    [["oc-excel-est-crit","CRITICO"],["oc-excel-est-riesgo","RIESGO"],
     ["oc-excel-est-alerta","ALERTA"],["oc-excel-est-ok","OK"],["oc-excel-est-sc","SIN_CONSUMO"]]
    .filter(([id])=>document.getElementById(id).checked).map(([,v])=>v)
  );
  document.getElementById("oc-excel-modal").style.display = "none";
  const origenVal = document.querySelector('input[name="oc-excel-origen"]:checked')?.value || "ALL";

  const wb = XLSX.utils.book_new();
  const flags = {"4301":"CO","4321":"CO","PERU":"PE","4403":"PE"};

  selCentros.forEach(key => {
    const ct   = DATA.centros[key];
    if (!ct) return;
    const mats = ct.materiales;
    const getEst = mat => { const m = mats.find(x=>x.mat===mat); return m?m.estado:""; };
    const getImp = mat => { const m = mats.find(x=>x.mat===mat); return m?(m.importado?"Importado":"Nacional"):""; };
    const matOk  = mat => { const est = getEst(mat); return !estSet || estSet.has(est); };
    const origenOk = mat => {
      if (origenVal === "ALL") return true;
      const m = mats.find(x => x.mat === mat);
      if (!m) return true;
      return origenVal === "IMP" ? m.importado === true : m.importado === false;
    };
    const sfx    = ct.nombre.substring(0,15).replace(/[:\\/?*\[\]]/g,"");

    if (incOC) {
      const oc = (ct.oc||[]).filter(o => matOk(o.mat) && origenOk(o.mat))
        .sort((a,b) => { if(a.atrasada&&!b.atrasada)return -1; if(!a.atrasada&&b.atrasada)return 1; return b.dias_atraso-a.dias_atraso; });
      const hdr = ["Centro","N° OC","Material","Descripción","Cant. Pedida","Pendiente","UM","Fecha Doc.","Fecha Entrega (SAP)","Fecha Proyectada","Días Atraso","Proveedor","Origen","Estado Mat."];
      const rows = oc.map(o => [ct.nombre, o.doc, o.mat, o.desc, o.qty_pedida, o.qty_pend, o.um,
        o.fecha_doc, o.fecha_entrega, o.atrasada?o.fecha_proyectada:null, o.atrasada?o.dias_atraso:null, o.proveedor, getImp(o.mat), getEst(o.mat)]);
      const ws = XLSX.utils.aoa_to_sheet([hdr,...rows]);
      ws["!cols"] = [14,12,12,40,12,12,6,12,14,14,10,30,12,12].map(w=>({wch:w}));
      XLSX.utils.book_append_sheet(wb, ws, (sfx+" OC").substring(0,31));
    }
    if (incSOL) {
      const sol = (ct.sol||[]).filter(s => matOk(s.mat) && origenOk(s.mat))
        .sort((a,b) => (a.fecha_entrega||"").localeCompare(b.fecha_entrega||""));
      const hdr = ["Centro","N° SOLPED","Material","Descripción","Cantidad","UM","Fecha Entrega","Estado Lib.","Estado Mat."];
      const rows = sol.map(s => [ct.nombre, s.doc, s.mat, s.desc, s.qty, s.um,
        s.fecha_entrega, s.sin_liberar?"Unreleased":"Released", getEst(s.mat)]);
      const ws = XLSX.utils.aoa_to_sheet([hdr,...rows]);
      ws["!cols"] = [14,12,12,40,12,6,12,14,12].map(w=>({wch:w}));
      XLSX.utils.book_append_sheet(wb, ws, (sfx+" SOL").substring(0,31));
    }
  });

  if (!wb.SheetNames.length) { alert("No datas para los filtros seleccionados."); return; }
  XLSX.writeFile(wb, `OC_SOL_${DATA.meta.fecha_saldo.replace(/-/g,"")}.xlsx`);
}
// Poblar centros en modal OC Excel
(function() {
  const container = document.getElementById("oc-excel-centro-list");
  const flags = {"4301":"🇨🇴","4321":"🇨🇴","PERU":"🇵🇪","4403":"🇵🇪"};
  Object.entries(DATA.centros).forEach(([key, ct]) => {
    const label = document.createElement("label");
    label.style.cssText = "display:flex;align-items:center;gap:10px;cursor:pointer;color:var(--tx1);font-size:12px";
    const cb = document.createElement("input");
    cb.type = "checkbox"; cb.className = "oc-excel-centro-chk"; cb.value = key; cb.checked = true;
    cb.style.cssText = "width:16px;height:16px;cursor:pointer;accent-color:#3ecf8e";
    label.appendChild(cb);
    const span = document.createElement("span");
    span.innerHTML = (flags[key]||"") + " <b>" + ct.nombre + "</b> <span style='color:var(--tx2);font-size:10px'>(" + (ct.oc||[]).length + " OC · " + (ct.sol||[]).length + " SOL)</span>";
    label.appendChild(span);
    container.appendChild(label);
  });
})();
</script>
</body>
</html>
"""


def build_impactos_xlsx_b64(country_centros: list, fecha_saldo: str) -> str:
    """
    Genera el Excel "Impact Report" con hojas por país.

    Parameters
    ----------
    country_centros : list of dicts
        [{"pais": "COLOMBIA", "centros": [col_data, car_data]}, ...]
        donde cada centro_data es el dict retornado por build_centro_data().
    fecha_saldo : str
        Fecha legible ("15-Abr-2026") para la columna FECHA.

    Returns
    -------
    str  Base64-encoded XLSX bytes.
    """
    import io, base64
    from openpyxl import Workbook
    from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                                  numbers)
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.datavalidation import DataValidation

    PAISES_ORDEN = ["COLOMBIA", "PERÚ", "CHILE", "BRASIL", "ARGENTINA"]
    # Colores por estado
    FILL_CRIT = PatternFill("solid", fgColor="FF4D4D")
    FILL_RIES = PatternFill("solid", fgColor="F5A623")
    FILL_HDR  = PatternFill("solid", fgColor="2D3450")
    WHITE     = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    FONT_NORM = Font(size=10, name="Calibri")
    FONT_BOLD = Font(size=10, bold=True, name="Calibri")
    ALIGN_C   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ALIGN_L   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin = Side(style="thin", color="CCCCCC")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

    HEADERS = ["CENTRO", "CÓDIGO", "DESCRIPCIÓN", "FECHA", "IMPACTO", "ACCIÓN"]
    COL_WIDTHS = [16, 13, 40, 14, 26, 40]

    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    country_map = {c["pais"]: c["centros"] for c in country_centros}

    for pais in PAISES_ORDEN:
        ws = wb.create_sheet(title=pais)

        # Data validation dropdown para IMPACTO (col E = 5)
        dv = DataValidation(
            type="list",
            formula1='"Parada de línea,No parada de línea"',
            allow_blank=True,
            showDropDown=False,
        )
        dv.sqref = "E2:E2000"
        ws.add_data_validation(dv)

        # Header row
        for col_idx, (hdr, width) in enumerate(zip(HEADERS, COL_WIDTHS), start=1):
            cell = ws.cell(row=1, column=col_idx, value=hdr)
            cell.font = WHITE
            cell.fill = FILL_HDR
            cell.alignment = ALIGN_C
            cell.border = BORDER
            ws.column_dimensions[get_column_letter(col_idx)].width = width
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = "A2"

        centros_list = country_map.get(pais, [])
        row = 2
        for centro_data in centros_list:
            nombre = centro_data.get("nombre", "")
            mats = [m for m in centro_data.get("materiales", [])
                    if m.get("estado") in ("CRITICO", "RIESGO")]
            # Sort CRITICO first, that risk
            mats.sort(key=lambda m: 0 if m["estado"] == "CRITICO" else 1)
            for m in mats:
                fill = FILL_CRIT if m["estado"] == "CRITICO" else FILL_RIES
                values = [nombre, m["mat"], m["desc"], fecha_saldo, "", ""]
                for col_idx, val in enumerate(values, start=1):
                    cell = ws.cell(row=row, column=col_idx, value=val)
                    cell.border = BORDER
                    cell.alignment = ALIGN_L if col_idx != 4 else ALIGN_C
                    cell.font = FONT_NORM
                    if col_idx in (1, 2):  # CENTRO y CÓDIGO destacan con fill de estado
                        cell.fill = fill
                        cell.font = FONT_BOLD
                row += 1

        if row == 2:  # No datas — agregar fila guía
            ws.cell(row=2, column=1, value="(No critical or at-risk materials)").font = Font(
                italic=True, color="888888", size=9, name="Calibri")

    buf = io.BytesIO()
    wb.save(buf)
    return base64.b64encode(buf.getvalue()).decode("ascii")


def render_html(json_str: str) -> str:
    """
    Inject the serialised DATA JSON into the template placeholder.

    The template contains a single /*DATA_PLACEHOLDER*/ comment inside
    the <script> block where the JS variable DATA is assigned.
    This function replaces it with the real JSON string.

    Parameters
    ----------
    json_str : str
        Compact JSON produced by json_builder.build_json().

    Returns
    -------
    str
        Complete, self-contained HTML document ready to write to disk.
    """
    html = HTML_TEMPLATE.replace("/*DATA_PLACEHOLDER*/", json_str)
    html = html.replace("/*IMPACTOS_B64_PLACEHOLDER*/", "")
    return html


def write_html(json_str: str, out_path: Path, impactos_b64: str = "") -> None:
    """
    Render the template and write the result to *out_path*.

    Parameters
    ----------
    json_str : str
        Compact JSON produced by json_builder.build_json().
    out_path : Path
        Destination file (typically supply_planner_v5.html).
    impactos_b64 : str
        Base64-encoded XLSX for Impact Report.
    """
    html = HTML_TEMPLATE.replace("/*DATA_PLACEHOLDER*/", json_str)
    html = html.replace("/*IMPACTOS_B64_PLACEHOLDER*/", impactos_b64)
    Path(out_path).write_text(html, encoding="utf-8")
